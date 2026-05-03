"""Export aFRR daily revenue (balanced vs aggressive strategies) to Excel.

Three sheets (mirrors the PZU script):
  1. "Summary"          — battery params, headline annual totals, monthly
                          comparison across all 3 strategies, notes.
  2. "aFRR_balanced"    — 366 daily rows at 80% acceptance, capacity multiplier 1.0.
  3. "aFRR_aggressive"  — 366 daily rows at 60% acceptance, capacity multiplier 1.15.

Battery: 10 MW / 20 MWh (per 2026-05-03 user directive). aFRR has two
revenue legs:
  - CAPACITY revenue:  power_mw × 24h × YOUR_BID × acceptance_rate.
    Paid for availability whether activated or not.
  - ACTIVATION revenue: captured_energy × activation_price, where
    captured_energy = min(market_activations × 10% share × acceptance,
    power × 24h, battery_capacity × 1 cycle/day).

The model picks the more profitable of aFRR+ (upward) or aFRR- (downward)
per day. Logic mirrors fr_service.project_annual_revenue at daily
granularity instead of monthly.

Run:
    cd backend
    PYTHONPATH=. arch -arm64 /usr/local/bin/python3 scripts/afrr_daily_excel.py
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional, Tuple

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.config import settings  # noqa: E402
from app.services.fr_service import FRService  # noqa: E402

# ===================================================================
#  Battery + market constants
# ===================================================================
POWER_MW = float(settings.DEFAULT_POWER_MW)             # 10.0
CAPACITY_MWH = float(settings.DEFAULT_CAPACITY_MWH)     # 20.0
HOURS_PER_DAY = 24
MARKET_SHARE = 0.10  # 10% of market activations captured (small entrant)
AFRR_FLOOR_EUR_MW_H = 5.0   # ROMANIAN_FR_PRODUCTS["aFRR"]["capacity_eur_mw_h"]
AFRR_UP_CAP_EUR_MW_H = 60.0
AFRR_DOWN_CAP_EUR_MW_H = 40.0
PRICE_VALID_MAX = 500.0  # filter outlier prices (matches production)

# Strategy configurations (mirror fr_service.project_annual_revenue).
STRATEGIES = {
    "balanced": {"acceptance": 0.80, "capacity_multiplier": 1.00},
    "aggressive": {"acceptance": 0.60, "capacity_multiplier": 1.15},
    "conservative": {"acceptance": 0.90, "capacity_multiplier": 0.85},
}

# ===================================================================
#  Column-name constants
# ===================================================================
COL_DATE = "Date"
COL_WEEKDAY = "Weekday"
COL_UP_AVG = "aFRR+ avg price (€/MWh)"
COL_DOWN_AVG = "aFRR- avg price (€/MWh)"
COL_SELECTED = "Selected"
COL_BID_PCT = "Bid percentile"
COL_PROXY = "Proxy activation price (€/MWh)"
COL_BID = "Recommended bid (€/MW/h)"
COL_POWER = "Power (MW)"
COL_HOURS = "Hours"
COL_CAP_REV = "Capacity revenue (€)"
COL_MKT_ACT = "Market activations (MWh)"
COL_CAPTURED = "Captured energy (MWh)"
COL_ACT_REV = "Activation revenue (€)"
COL_DAILY_REV = "Daily revenue (€)"

PLACEHOLDER_DASH = "—"
SHEET_SUMMARY = "Summary"
SHEET_BALANCED = "aFRR_balanced"
SHEET_AGGRESSIVE = "aFRR_aggressive"


# ===================================================================
#  Daily revenue calc
# ===================================================================
def _filter_valid(prices: pd.Series) -> pd.Series:
    return prices[(prices.abs() > 0) & (prices.abs() < PRICE_VALID_MAX)]


# Real Romanian aFRR capacity prices — read from the CANONICAL resolver
# (app/market_data/capacity_prices.py) so this script returns the same
# numbers as the production fr_service. Default mode is the recent-window
# DAMAS sample mean (currently €6.57 aFRRUp / €3.77 aFRRDown after the
# Aug 2025 - Jan 2026 compression).
from app.market_data.capacity_prices import get_canonical_capacity_price


def _load_damas_capacity_avgs() -> dict:
    return {
        "aFRRUp": get_canonical_capacity_price("aFRRUp").price_eur_mw_h,
        "aFRRDown": get_canonical_capacity_price("aFRRDown").price_eur_mw_h,
    }


_DAMAS_AVGS = _load_damas_capacity_avgs()


def _bid_for_direction(proxy: float, capacity_ratio: float, mult: float, cap: float,
                        damas_clearing: float) -> float:
    """Anchored on real DAMAS clearing × strategy multiplier."""
    raw = damas_clearing * mult
    return float(min(cap, max(AFRR_FLOOR_EUR_MW_H, raw)))


def _direction_revenue(
    proxy: float, bid: float, market_activations: float,
    acceptance: float, capacity_mwh: float,
) -> Tuple[float, float, float, float]:
    """Return (capacity_revenue, captured_energy, activation_revenue, daily_total)."""
    capacity_rev = POWER_MW * HOURS_PER_DAY * bid * acceptance
    our_max_capacity = POWER_MW * HOURS_PER_DAY  # power-limited per day
    unlimited = min(market_activations * MARKET_SHARE, our_max_capacity) * acceptance
    captured = min(unlimited, capacity_mwh)  # battery throughput cap
    activation_rev = captured * abs(proxy)
    return capacity_rev, captured, activation_rev, capacity_rev + activation_rev


def daily_afrr_row(group: pd.DataFrame, day, strategy: str) -> Optional[dict]:
    """Build one daily detail row for `strategy`. Returns None if the day
    has no valid prices (e.g. data outage)."""
    cfg = STRATEGIES[strategy]
    acceptance = cfg["acceptance"]
    mult = cfg["capacity_multiplier"]

    valid_up = _filter_valid(group["afrr_up_price_eur"])
    valid_down = _filter_valid(group["afrr_down_price_eur"])
    if valid_up.empty or valid_down.empty:
        return None

    bid_pct = 1.0 - acceptance
    capacity_ratio = 0.20 + bid_pct * 0.10  # 0.20-0.30

    up_proxy = float(valid_up.quantile(bid_pct))
    down_proxy = float(valid_down.quantile(bid_pct))
    up_bid = _bid_for_direction(up_proxy, capacity_ratio, mult,
                                AFRR_UP_CAP_EUR_MW_H, _DAMAS_AVGS["aFRRUp"])
    down_bid = _bid_for_direction(down_proxy, capacity_ratio, mult,
                                  AFRR_DOWN_CAP_EUR_MW_H, _DAMAS_AVGS["aFRRDown"])

    market_up = float(group["afrr_up_activated_mwh"].sum())
    market_down = float(group["afrr_down_activated_mwh"].sum())

    up_cap, up_cap_e, up_act, up_total = _direction_revenue(
        up_proxy, up_bid, market_up, acceptance, CAPACITY_MWH)
    down_cap, down_cap_e, down_act, down_total = _direction_revenue(
        down_proxy, down_bid, market_down, acceptance, CAPACITY_MWH)

    if up_total >= down_total:
        selected = "aFRR+"
        avg = float(valid_up.mean())
        proxy = up_proxy
        bid = up_bid
        cap_rev, captured, act_rev, total = up_cap, up_cap_e, up_act, up_total
        market = market_up
    else:
        selected = "aFRR-"
        avg = float(valid_down.mean())
        proxy = down_proxy
        bid = down_bid
        cap_rev, captured, act_rev, total = down_cap, down_cap_e, down_act, down_total
        market = market_down

    return {
        COL_DATE: pd.Timestamp(day),
        COL_WEEKDAY: pd.Timestamp(day).day_name(),
        COL_UP_AVG: round(float(valid_up.mean()), 2),
        COL_DOWN_AVG: round(float(valid_down.mean()), 2),
        COL_SELECTED: selected,
        COL_BID_PCT: round(bid_pct, 2),
        COL_PROXY: round(proxy, 2),
        COL_BID: round(bid, 2),
        COL_POWER: POWER_MW,
        COL_HOURS: HOURS_PER_DAY,
        COL_CAP_REV: cap_rev,
        COL_MKT_ACT: round(market, 2),
        COL_CAPTURED: round(captured, 2),
        COL_ACT_REV: act_rev,
        COL_DAILY_REV: total,
    }


def empty_row(day) -> dict:
    return {
        COL_DATE: pd.Timestamp(day),
        COL_WEEKDAY: pd.Timestamp(day).day_name(),
        COL_UP_AVG: None, COL_DOWN_AVG: None,
        COL_SELECTED: PLACEHOLDER_DASH,
        COL_BID_PCT: None, COL_PROXY: None, COL_BID: None,
        COL_POWER: POWER_MW, COL_HOURS: 0,
        COL_CAP_REV: 0.0, COL_MKT_ACT: 0.0,
        COL_CAPTURED: 0.0, COL_ACT_REV: 0.0, COL_DAILY_REV: 0.0,
    }


# ===================================================================
#  Excel formatting helpers
# ===================================================================
EUR_FMT = '"€"#,##0;[Red]"-€"#,##0'
EUR_MWH_FMT = '"€"#,##0.00'
EUR_MW_H_FMT = '"€"#,##0.00" /MW/h"'
MWH_FMT = '#,##0.00" MWh"'
PCT_FMT = '0%'


def _styles():
    from openpyxl.styles import Border, Font, PatternFill, Side
    border = Side(border_style="thin", color="BFBFBF")
    return {
        "bold": Font(bold=True, size=11),
        "bold_white": Font(bold=True, color="FFFFFF", size=11),
        "bold_big": Font(bold=True, size=14, color="FFFFFF"),
        "header_fill": PatternFill("solid", fgColor="1F4E78"),
        "title_fill": PatternFill("solid", fgColor="2E75B6"),
        "total_fill": PatternFill("solid", fgColor="FFEB99"),
        "revenue_fill": PatternFill("solid", fgColor="E2EFDA"),
        "cost_fill": PatternFill("solid", fgColor="FCE4D6"),
        "profit_fill": PatternFill("solid", fgColor="DDEBF7"),
        "box": Border(left=border, right=border, top=border, bottom=border),
    }


def _detail_format(col: str) -> Optional[str]:
    if col in (COL_UP_AVG, COL_DOWN_AVG, COL_PROXY):
        return EUR_MWH_FMT
    if col == COL_BID:
        return EUR_MW_H_FMT
    if "(€)" in col:
        return EUR_FMT
    if "(MWh)" in col:
        return MWH_FMT
    if col == COL_BID_PCT:
        return PCT_FMT
    if col == COL_DATE:
        return "yyyy-mm-dd"
    return None


def _column_width(col: str) -> int:
    if col in (COL_DATE, COL_WEEKDAY, COL_SELECTED):
        return 12
    if "(€)" in col:
        return 18
    if "(MWh)" in col or col == COL_BID:
        return 18
    if "price" in col.lower() or "proxy" in col.lower():
        return 16
    return 14


def _style_header(ws, styles):
    from openpyxl.styles import Alignment
    for cell in ws[1]:
        cell.font = styles["bold_white"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = styles["box"]
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "C2"


def _style_body(ws, df, styles, profit_cols, rev_cols):
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    for col_idx, col_name in enumerate(df.columns, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _column_width(col_name)
        nf = _detail_format(col_name)
        align = Alignment(horizontal="left" if col_name in (COL_DATE, COL_WEEKDAY) else "center")
        for cell in next(ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row)):
            if nf:
                cell.number_format = nf
            cell.alignment = align
            cell.border = styles["box"]
        fill = None
        if col_name in profit_cols:
            fill = styles["profit_fill"]
        elif col_name in rev_cols:
            fill = styles["revenue_fill"]
        if fill:
            for cell in next(ws.iter_cols(min_col=col_idx, max_col=col_idx,
                                          min_row=2, max_row=ws.max_row - 1)):
                cell.fill = fill


def _style_totals(ws, styles):
    for cell in ws[ws.max_row]:
        cell.font = styles["bold"]
        cell.fill = styles["total_fill"]
        cell.border = styles["box"]


def _style_detail(ws, df, styles):
    profit_cols = {COL_DAILY_REV}
    rev_cols = {COL_CAP_REV, COL_ACT_REV}
    _style_header(ws, styles)
    _style_body(ws, df, styles, profit_cols, rev_cols)
    _style_totals(ws, styles)


def _summary_title(ws, styles, first_date, last_date) -> None:
    from openpyxl.styles import Alignment
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = (
        f"aFRR daily revenue — balanced vs aggressive strategy  |  "
        f"{first_date} → {last_date}"
    )
    cell.font = styles["bold_big"]
    cell.fill = styles["title_fill"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    for col_letter, width in zip("ABCDEF", [38, 22, 22, 22, 22, 16]):
        ws.column_dimensions[col_letter].width = width


def _summary_battery(ws, styles, start_row: int) -> int:
    from openpyxl.styles import Alignment
    ws[f"A{start_row}"] = "Battery configuration"
    ws[f"A{start_row}"].font = styles["bold_white"]
    ws[f"A{start_row}"].fill = styles["header_fill"]
    ws.merge_cells(f"A{start_row}:F{start_row}")
    ws[f"A{start_row}"].alignment = Alignment(horizontal="center")
    params = [
        ("Power (MW)", POWER_MW, "Battery rated power"),
        ("Energy capacity (MWh)", CAPACITY_MWH, "Usable per cycle (per user directive)"),
        ("Hours per day", HOURS_PER_DAY, "Capacity availability window"),
        ("Market share assumption", MARKET_SHARE, "Captured share of market aFRR activations (10% = small new entrant)"),
        ("aFRR capacity floor", AFRR_FLOOR_EUR_MW_H, "ROMANIAN_FR_PRODUCTS catalog floor (5 €/MW/h)"),
        ("aFRR+ bid cap", AFRR_UP_CAP_EUR_MW_H, "Hard ceiling on upward bid"),
        ("aFRR- bid cap", AFRR_DOWN_CAP_EUR_MW_H, "Hard ceiling on downward bid"),
        ("Price-validity range", f"|p| ∈ (0, {PRICE_VALID_MAX}) €/MWh", "Outlier filter (matches production)"),
    ]
    for offset, (label, value, note) in enumerate(params, start=1):
        r = start_row + offset
        ws[f"A{r}"] = label
        ws[f"B{r}"] = value
        ws[f"C{r}"] = note
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws[f"A{r}"].font = styles["bold"]
        for col in ("A", "B", "C", "D", "E", "F"):
            ws[f"{col}{r}"].border = styles["box"]
    return start_row + len(params) + 2


def _summary_headline(ws, styles, start_row: int, totals: dict) -> int:
    from openpyxl.styles import Alignment
    ws[f"A{start_row}"] = f"Headline annual totals ({totals['days']} days)"
    ws[f"A{start_row}"].font = styles["bold_white"]
    ws[f"A{start_row}"].fill = styles["header_fill"]
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    ws[f"A{start_row}"].alignment = Alignment(horizontal="center")

    hdr = start_row + 1
    cols = ["Metric", "Conservative", "Balanced", "Aggressive", "Best – Worst"]
    for col_letter, label in zip("ABCDE", cols):
        c = ws[f"{col_letter}{hdr}"]
        c.value = label
        c.font = styles["bold_white"]
        c.fill = styles["title_fill"]
        c.alignment = Alignment(horizontal="center")
        c.border = styles["box"]
    ws.merge_cells(start_row=hdr, start_column=5, end_row=hdr, end_column=6)

    cap = totals["capacity_revenue"]
    act = totals["activation_revenue"]
    daily = totals["daily_revenue"]
    rows = [
        ("Capacity revenue (€)",
         cap["conservative"], cap["balanced"], cap["aggressive"]),
        ("Activation revenue (€)",
         act["conservative"], act["balanced"], act["aggressive"]),
        ("Total annual revenue (€)",
         daily["conservative"], daily["balanced"], daily["aggressive"]),
        ("Avg daily revenue (€)",
         daily["conservative"] / totals["days"], daily["balanced"] / totals["days"],
         daily["aggressive"] / totals["days"]),
    ]
    for offset, (label, c_val, b_val, a_val) in enumerate(rows, start=1):
        r = hdr + offset
        ws[f"A{r}"] = label
        ws[f"B{r}"] = c_val
        ws[f"C{r}"] = b_val
        ws[f"D{r}"] = a_val
        spread = max(c_val, b_val, a_val) - min(c_val, b_val, a_val)
        ws[f"E{r}"] = spread
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws[f"A{r}"].font = styles["bold"]
        ws[f"B{r}"].number_format = EUR_FMT
        ws[f"C{r}"].number_format = EUR_FMT
        ws[f"D{r}"].number_format = EUR_FMT
        ws[f"E{r}"].number_format = EUR_FMT
        for col in ("A", "B", "C", "D", "E", "F"):
            ws[f"{col}{r}"].border = styles["box"]
        if "Total" in label:
            for col in ("A", "B", "C", "D", "E", "F"):
                ws[f"{col}{r}"].fill = styles["total_fill"]
                ws[f"{col}{r}"].font = styles["bold"]
    return hdr + len(rows) + 1


def _summary_monthly(ws, styles, start_row: int, monthly_df: pd.DataFrame) -> int:
    from openpyxl.styles import Alignment
    ws[f"A{start_row}"] = "Monthly comparison"
    ws[f"A{start_row}"].font = styles["bold_white"]
    ws[f"A{start_row}"].fill = styles["header_fill"]
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    ws[f"A{start_row}"].alignment = Alignment(horizontal="center")
    hdr = start_row + 1
    headers = ["Month", "Days", "Conservative (€)", "Balanced (€)", "Aggressive (€)", "Spread (€)"]
    for col_letter, label in zip("ABCDEF", headers):
        c = ws[f"{col_letter}{hdr}"]
        c.value = label
        c.font = styles["bold_white"]
        c.fill = styles["title_fill"]
        c.alignment = Alignment(horizontal="center")
        c.border = styles["box"]
    r = hdr + 1
    for month, row in monthly_df.iterrows():
        ws[f"A{r}"] = month
        ws[f"B{r}"] = int(row["days"])
        ws[f"C{r}"] = float(row["conservative"])
        ws[f"D{r}"] = float(row["balanced"])
        ws[f"E{r}"] = float(row["aggressive"])
        ws[f"F{r}"] = float(row["spread"])
        ws[f"A{r}"].font = styles["bold"]
        for col_letter in ("C", "D", "E", "F"):
            ws[f"{col_letter}{r}"].number_format = EUR_FMT
        for col in ("A", "B", "C", "D", "E", "F"):
            ws[f"{col}{r}"].border = styles["box"]
            ws[f"{col}{r}"].alignment = Alignment(
                horizontal="center" if col != "A" else "left",
            )
        r += 1
    # Totals.
    ws[f"A{r}"] = "TOTAL"
    ws[f"B{r}"] = int(monthly_df["days"].sum())
    ws[f"C{r}"] = float(monthly_df["conservative"].sum())
    ws[f"D{r}"] = float(monthly_df["balanced"].sum())
    ws[f"E{r}"] = float(monthly_df["aggressive"].sum())
    ws[f"F{r}"] = float(monthly_df["spread"].sum())
    for col_letter in ("C", "D", "E", "F"):
        ws[f"{col_letter}{r}"].number_format = EUR_FMT
    for col in ("A", "B", "C", "D", "E", "F"):
        ws[f"{col}{r}"].font = styles["bold"]
        ws[f"{col}{r}"].fill = styles["total_fill"]
        ws[f"{col}{r}"].border = styles["box"]
        ws[f"{col}{r}"].alignment = Alignment(
            horizontal="center" if col != "A" else "left",
        )
    return r + 2


def _summary_notes(ws, styles, start_row: int) -> None:
    from openpyxl.styles import Alignment
    ws[f"A{start_row}"] = "Notes"
    ws[f"A{start_row}"].font = styles["bold_white"]
    ws[f"A{start_row}"].fill = styles["header_fill"]
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    ws[f"A{start_row}"].alignment = Alignment(horizontal="center")
    notes = [
        "Data source: DAMAS (Romanian TSO) 15-min slot data, loaded via fr_service._load_fr_data().",
        "Daily logic mirrors fr_service.project_annual_revenue at daily granularity. Per direction (aFRR+ / aFRR-): bid_percentile = 1 − acceptance_rate; capacity_ratio = 0.20 + bid_percentile × 0.10; bid = max(floor 5, min(cap 60/40, |proxy| × capacity_ratio × strategy_multiplier)).",
        "CAPACITY revenue = power × 24h × bid × acceptance_rate (paid for availability). ACTIVATION revenue = captured_energy × proxy_price, where captured_energy = min(market_activations × 10% share × acceptance, power × 24h, capacity_mwh).",
        "The model picks the more profitable of aFRR+ or aFRR- per day. The detail sheets show only the selected direction's numbers; the other direction's totals are absorbed into the daily pick.",
        "Strategy 'balanced' (acceptance 80%, multiplier 1.0) matches the production default; 'aggressive' (60%, 1.15) bids higher / accepts less; 'conservative' (90%, 0.85) bids lower / accepts more.",
        "Numbers are SIMULATED on real DAMAS data, not bankable. Compliance gates (ANRE / OPCOM / PRE / BRP / BSP / FSE) and tariff exemption math live downstream in investment_service.analyze().",
    ]
    r = start_row
    for note in notes:
        r += 1
        ws[f"A{r}"] = "• " + note
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 32


# ===================================================================
#  Main
# ===================================================================
def _total_row(df: pd.DataFrame) -> dict:
    return {
        COL_DATE: "TOTAL", COL_WEEKDAY: f"{len(df)} days",
        COL_UP_AVG: None, COL_DOWN_AVG: None,
        COL_SELECTED: PLACEHOLDER_DASH,
        COL_BID_PCT: None, COL_PROXY: None, COL_BID: None,
        COL_POWER: POWER_MW, COL_HOURS: int(df[COL_HOURS].sum()),
        COL_CAP_REV: float(df[COL_CAP_REV].sum()),
        COL_MKT_ACT: float(df[COL_MKT_ACT].sum()),
        COL_CAPTURED: float(df[COL_CAPTURED].sum()),
        COL_ACT_REV: float(df[COL_ACT_REV].sum()),
        COL_DAILY_REV: float(df[COL_DAILY_REV].sum()),
    }


def _build_strategy_df(df: pd.DataFrame, strategy: str) -> pd.DataFrame:
    rows = []
    for day, group in df.groupby(df["date"].dt.date):
        row = daily_afrr_row(group, day, strategy)
        rows.append(row if row is not None else empty_row(day))
    return pd.DataFrame(rows).sort_values(COL_DATE).reset_index(drop=True)


def _build_monthly_comparison(strategy_dfs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    months = strategy_dfs["balanced"].copy()
    months["Month"] = months[COL_DATE].dt.strftime("%Y-%m")
    days_per_month = months.groupby("Month").size()
    monthly = pd.DataFrame({"days": days_per_month})
    for strat, sdf in strategy_dfs.items():
        m = sdf.copy()
        m["Month"] = m[COL_DATE].dt.strftime("%Y-%m")
        monthly[strat] = m.groupby("Month")[COL_DAILY_REV].sum()
    monthly["spread"] = monthly[["conservative", "balanced", "aggressive"]].max(axis=1) - \
                        monthly[["conservative", "balanced", "aggressive"]].min(axis=1)
    return monthly


def style_workbook(out_path: Path, balanced_df: pd.DataFrame, aggressive_df: pd.DataFrame,
                   conservative_df: pd.DataFrame, monthly_df: pd.DataFrame,
                   first_date, last_date) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    styles = _styles()

    _style_detail(wb[SHEET_BALANCED], balanced_df, styles)
    _style_detail(wb[SHEET_AGGRESSIVE], aggressive_df, styles)

    totals = {
        "days": len(balanced_df),
        "capacity_revenue": {
            s: df[COL_CAP_REV].sum()
            for s, df in [("conservative", conservative_df),
                          ("balanced", balanced_df),
                          ("aggressive", aggressive_df)]
        },
        "activation_revenue": {
            s: df[COL_ACT_REV].sum()
            for s, df in [("conservative", conservative_df),
                          ("balanced", balanced_df),
                          ("aggressive", aggressive_df)]
        },
        "daily_revenue": {
            s: df[COL_DAILY_REV].sum()
            for s, df in [("conservative", conservative_df),
                          ("balanced", balanced_df),
                          ("aggressive", aggressive_df)]
        },
    }

    ws = wb[SHEET_SUMMARY]
    _summary_title(ws, styles, first_date, last_date)
    next_row = _summary_battery(ws, styles, start_row=3)
    next_row = _summary_headline(ws, styles, next_row, totals)
    next_row = _summary_monthly(ws, styles, next_row, monthly_df)
    _summary_notes(ws, styles, next_row)
    wb.active = wb.sheetnames.index(SHEET_SUMMARY)
    wb.save(out_path)


def _load_fr_window(fr: FRService) -> Tuple[pd.DataFrame, pd.Timestamp, pd.Timestamp]:
    df = fr._load_fr_data().copy()
    last_date = df["date"].max().date()
    first_date = (pd.Timestamp(last_date) - pd.Timedelta(days=365)).date()
    df = df[(df["date"].dt.date >= first_date) & (df["date"].dt.date <= last_date)]
    return df, first_date, last_date


def main() -> None:
    fr = FRService()
    df, first_date, last_date = _load_fr_window(fr)
    print(f"Window: {first_date} → {last_date}  ({df['date'].dt.date.nunique()} days)")
    print(f"Battery: {POWER_MW} MW / {CAPACITY_MWH} MWh")

    balanced_df = _build_strategy_df(df, "balanced")
    aggressive_df = _build_strategy_df(df, "aggressive")
    conservative_df = _build_strategy_df(df, "conservative")

    monthly_df = _build_monthly_comparison({
        "balanced": balanced_df,
        "aggressive": aggressive_df,
        "conservative": conservative_df,
    })

    balanced_out = pd.concat(
        [balanced_df, pd.DataFrame([_total_row(balanced_df)])], ignore_index=True)
    aggressive_out = pd.concat(
        [aggressive_df, pd.DataFrame([_total_row(aggressive_df)])], ignore_index=True)

    out_dir = ROOT / "reports"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"aFRR_daily_balanced_vs_aggressive_{first_date}_to_{last_date}.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        pd.DataFrame().to_excel(xl, sheet_name=SHEET_SUMMARY, index=False)
        balanced_out.to_excel(xl, sheet_name=SHEET_BALANCED, index=False)
        aggressive_out.to_excel(xl, sheet_name=SHEET_AGGRESSIVE, index=False)

    style_workbook(out_path, balanced_df, aggressive_df, conservative_df,
                   monthly_df, first_date, last_date)

    print(f"\nWrote: {out_path}")
    print(f"  Sheet '{SHEET_SUMMARY}'        — battery params + headline + monthly + notes")
    print(f"  Sheet '{SHEET_BALANCED}'   — {len(balanced_df)} daily rows + totals row")
    print(f"  Sheet '{SHEET_AGGRESSIVE}' — {len(aggressive_df)} daily rows + totals row")
    print("\nHeadline annual totals:")
    for strat, sdf in [("conservative", conservative_df),
                       ("balanced", balanced_df),
                       ("aggressive", aggressive_df)]:
        total = sdf[COL_DAILY_REV].sum()
        cap = sdf[COL_CAP_REV].sum()
        act = sdf[COL_ACT_REV].sum()
        print(f"  {strat:12s}: capacity €{cap:>10,.0f}  + activation €{act:>10,.0f}  = €{total:>10,.0f}")


if __name__ == "__main__":
    main()
