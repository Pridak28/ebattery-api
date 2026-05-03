"""Export aFRR daily revenue using REAL per-slot market prices to Excel.

Mirrors the PZU 1-cycle / 2-cycle script but for aFRR. Two key
differences from the proxy-based afrr_daily_excel.py:

  - Capacity revenue is your bid × power × hours × acceptance, same as before.
  - ACTIVATION revenue is computed slot-by-slot using the REAL DAMAS
    afrr_*_price_eur for each 15-min slot the TSO activated, not a
    synthetic 20th-percentile quantile proxy. Daily battery throughput
    cap (20 MWh) is enforced as a running budget across the day.

Three sheets:
  1. "Summary"           — battery params, headline annual totals,
                           monthly comparison, notes.
  2. "aFRR_balanced"     — 366 daily rows at 80% acceptance.
  3. "aFRR_aggressive"   — 366 daily rows at 60% acceptance.

Battery: 10 MW / 20 MWh, RTE 0.97, full 0-100% SOC band (2026-05-03
user directive). aFRR has two revenue legs:
  - Capacity:  power × 24h × bid × acceptance (paid for availability).
  - Activation: Σ_slot (your_MWh_in_slot × actual_settlement_price),
                with your_MWh = min(market_MWh × share, power × 0.25h,
                                    remaining battery capacity).

Run:
    cd backend
    PYTHONPATH=. arch -arm64 /usr/local/bin/python3 scripts/afrr_daily_real_market_excel.py
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional, Tuple

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.config import settings  # noqa: E402
from app.services.fr_service import FRService  # noqa: E402

# ===================================================================
#  Battery + market constants
# ===================================================================
POWER_MW = float(settings.DEFAULT_POWER_MW)             # 10.0
CAPACITY_MWH = float(settings.DEFAULT_CAPACITY_MWH)     # 20.0
HOURS_PER_DAY = 24
SLOT_HOURS = 0.25
SLOTS_PER_DAY = 96
MARKET_SHARE = 0.10                                     # captured share of activations
AFRR_FLOOR_EUR_MW_H = 5.0
AFRR_UP_CAP_EUR_MW_H = 60.0
AFRR_DOWN_CAP_EUR_MW_H = 40.0
PRICE_VALID_MAX = 500.0

STRATEGIES = {
    "balanced":     {"acceptance": 0.80, "capacity_multiplier": 1.00},
    "aggressive":   {"acceptance": 0.60, "capacity_multiplier": 1.15},
    "conservative": {"acceptance": 0.90, "capacity_multiplier": 0.85},
}

# ===================================================================
#  Column constants
# ===================================================================
COL_DATE = "Date"
COL_WEEKDAY = "Weekday"
COL_SELECTED = "Selected"
COL_BID = "Recommended bid (€/MW/h)"
COL_ACCEPT = "Acceptance rate"
COL_HOURS = "Capacity hours"
COL_CAP_REV = "Capacity revenue (€)"
COL_SLOTS_ACTIVATED = "Slots activated (#/96)"
COL_AVG_ACT_PRICE = "Avg activation price (€/MWh)"
COL_MAX_ACT_PRICE = "Max activation price (€/MWh)"
COL_MKT_TOTAL = "Market activations (MWh)"
COL_CAPTURED = "Captured energy (MWh)"
COL_BATTERY_BOUND = "Battery cap binding"
COL_ACT_REV = "Activation revenue (€)"
COL_DAILY_REV = "Daily revenue (€)"

PLACEHOLDER_DASH = "—"
SHEET_SUMMARY = "Summary"
SHEET_BALANCED = "aFRR_balanced"
SHEET_AGGRESSIVE = "aFRR_aggressive"


# ===================================================================
#  Slot-level simulation
# ===================================================================
def _filter_valid(prices: pd.Series) -> pd.Series:
    return prices[(prices.abs() > 0) & (prices.abs() < PRICE_VALID_MAX)]


# Real Romanian aFRR capacity prices — CANONICAL resolver
# (app/market_data/capacity_prices.py). Same numbers as production
# fr_service: recent-window DAMAS sample mean (€6.57 / €3.77 today).
from app.market_data.capacity_prices import get_canonical_capacity_price


def _load_damas_capacity_avgs() -> dict:
    return {
        "aFRRUp": get_canonical_capacity_price("aFRRUp").price_eur_mw_h,
        "aFRRDown": get_canonical_capacity_price("aFRRDown").price_eur_mw_h,
    }


_DAMAS_AVGS = _load_damas_capacity_avgs()


def _bid_for_direction(proxy: float, capacity_ratio: float, mult: float, cap: float,
                        damas_clearing: float) -> float:
    """Compute the bid as a multiple of the real DAMAS capacity clearing price.
    `proxy` and `capacity_ratio` are kept in the signature for backward
    compatibility but no longer drive the bid. `damas_clearing` is the
    real avg clearing price (€/MW/h) for the direction.
    """
    raw = damas_clearing * mult
    return float(min(cap, max(AFRR_FLOOR_EUR_MW_H, raw)))


def _simulate_direction(
    prices: np.ndarray, market_activations: np.ndarray,
    acceptance: float, sign: int,
) -> dict:
    """Walk the day's 96 slots chronologically, accumulate activation
    revenue using REAL per-slot prices. Battery cap (20 MWh) is enforced
    as a running budget — once you've delivered the day's full cycle,
    later slot activations are skipped.

    `sign` = +1 for aFRR+ (price as-is), -1 for aFRR- (we earn on |price|
    when we charge — the sign discussion).

    Returns dict with: slots_activated, captured_mwh, activation_revenue,
    avg_activation_price (over slots that paid us), max_activation_price,
    battery_bound (whether the 20 MWh cap binded today).
    """
    n = len(prices)
    captured_total = 0.0
    activation_revenue = 0.0
    activated_slots = 0
    activation_prices = []
    per_slot_max_mwh = POWER_MW * SLOT_HOURS  # 10 MW × 0.25h = 2.5 MWh max per slot
    battery_bound = False

    for i in range(n):
        market_mwh = float(market_activations[i])
        if market_mwh <= 0:
            continue
        # Our share of the market activation in this slot.
        our_share_mwh = market_mwh * MARKET_SHARE
        constrained = min(our_share_mwh, per_slot_max_mwh) * acceptance
        if constrained <= 0:
            continue
        # Apply daily-cycle battery throughput cap as a running budget.
        remaining = CAPACITY_MWH - captured_total
        if remaining <= 0:
            battery_bound = True
            break
        slot_mwh = min(constrained, remaining)
        if slot_mwh < constrained:
            battery_bound = True
        price = float(prices[i])
        # We earn |price| × MWh on either direction (aFRR-: when negative
        # price means TSO pays you to absorb energy → revenue too).
        slot_rev = slot_mwh * abs(price)
        captured_total += slot_mwh
        activation_revenue += slot_rev
        activated_slots += 1
        activation_prices.append(price)

    return {
        "slots_activated": activated_slots,
        "captured_mwh": captured_total,
        "activation_revenue": activation_revenue,
        "avg_activation_price": (
            float(np.mean(np.abs(activation_prices))) if activation_prices else 0.0
        ),
        "max_activation_price": (
            float(np.max(np.abs(activation_prices))) if activation_prices else 0.0
        ),
        "battery_bound": battery_bound,
        "market_total_mwh": float(np.sum(market_activations)),
    }


def daily_afrr_row(group: pd.DataFrame, day, strategy: str) -> Optional[dict]:
    """Build one daily row using slot-level real-market simulation.
    Picks aFRR+ or aFRR- based on which produces higher total revenue."""
    cfg = STRATEGIES[strategy]
    acceptance = cfg["acceptance"]
    mult = cfg["capacity_multiplier"]

    # Sort by slot to ensure chronological iteration.
    g = group.sort_values("slot")
    up_prices = g["afrr_up_price_eur"].to_numpy()
    down_prices = g["afrr_down_price_eur"].to_numpy()
    up_activations = g["afrr_up_activated_mwh"].to_numpy()
    down_activations = g["afrr_down_activated_mwh"].to_numpy()

    # Recommended bid still derived from quantile of valid prices (this is
    # the bid you SUBMIT; activation revenue uses the REAL settled price).
    valid_up = _filter_valid(g["afrr_up_price_eur"])
    valid_down = _filter_valid(g["afrr_down_price_eur"])
    if valid_up.empty or valid_down.empty:
        return None

    bid_pct = 1.0 - acceptance
    capacity_ratio = 0.20 + bid_pct * 0.10
    up_proxy = float(valid_up.quantile(bid_pct))
    down_proxy = float(valid_down.quantile(bid_pct))
    # Bids anchored on REAL DAMAS clearing prices (not the activation-quantile proxy).
    up_bid = _bid_for_direction(up_proxy, capacity_ratio, mult,
                                AFRR_UP_CAP_EUR_MW_H, _DAMAS_AVGS["aFRRUp"])
    down_bid = _bid_for_direction(down_proxy, capacity_ratio, mult,
                                  AFRR_DOWN_CAP_EUR_MW_H, _DAMAS_AVGS["aFRRDown"])

    up_cap_rev = POWER_MW * HOURS_PER_DAY * up_bid * acceptance
    down_cap_rev = POWER_MW * HOURS_PER_DAY * down_bid * acceptance

    up_sim = _simulate_direction(up_prices, up_activations, acceptance, sign=+1)
    down_sim = _simulate_direction(down_prices, down_activations, acceptance, sign=-1)

    up_total = up_cap_rev + up_sim["activation_revenue"]
    down_total = down_cap_rev + down_sim["activation_revenue"]

    if up_total >= down_total:
        sel = "aFRR+"
        cap_rev, bid, sim = up_cap_rev, up_bid, up_sim
        total = up_total
    else:
        sel = "aFRR-"
        cap_rev, bid, sim = down_cap_rev, down_bid, down_sim
        total = down_total

    return {
        COL_DATE: pd.Timestamp(day),
        COL_WEEKDAY: pd.Timestamp(day).day_name(),
        COL_SELECTED: sel,
        COL_BID: round(bid, 2),
        COL_ACCEPT: acceptance,
        COL_HOURS: HOURS_PER_DAY,
        COL_CAP_REV: cap_rev,
        COL_SLOTS_ACTIVATED: sim["slots_activated"],
        COL_AVG_ACT_PRICE: round(sim["avg_activation_price"], 2),
        COL_MAX_ACT_PRICE: round(sim["max_activation_price"], 2),
        COL_MKT_TOTAL: round(sim["market_total_mwh"], 2),
        COL_CAPTURED: round(sim["captured_mwh"], 2),
        COL_BATTERY_BOUND: "yes" if sim["battery_bound"] else "no",
        COL_ACT_REV: sim["activation_revenue"],
        COL_DAILY_REV: total,
    }


def empty_row(day) -> dict:
    return {
        COL_DATE: pd.Timestamp(day),
        COL_WEEKDAY: pd.Timestamp(day).day_name(),
        COL_SELECTED: PLACEHOLDER_DASH,
        COL_BID: None, COL_ACCEPT: None, COL_HOURS: 0,
        COL_CAP_REV: 0.0, COL_SLOTS_ACTIVATED: 0,
        COL_AVG_ACT_PRICE: None, COL_MAX_ACT_PRICE: None,
        COL_MKT_TOTAL: 0.0, COL_CAPTURED: 0.0,
        COL_BATTERY_BOUND: PLACEHOLDER_DASH,
        COL_ACT_REV: 0.0, COL_DAILY_REV: 0.0,
    }


# ===================================================================
#  Excel formatting helpers
# ===================================================================
EUR_FMT = '"€"#,##0;[Red]"-€"#,##0'
EUR_MWH_FMT = '"€"#,##0.00'
EUR_MW_H_FMT = '"€"#,##0.00" /MW/h"'
MWH_FMT = '#,##0.00" MWh"'
PCT_FMT = '0%'


def _styles():
    from openpyxl.styles import Border, Font, PatternFill, Side
    border = Side(border_style="thin", color="BFBFBF")
    return {
        "bold": Font(bold=True, size=11),
        "bold_white": Font(bold=True, color="FFFFFF", size=11),
        "bold_big": Font(bold=True, size=14, color="FFFFFF"),
        "header_fill": PatternFill("solid", fgColor="1F4E78"),
        "title_fill": PatternFill("solid", fgColor="2E75B6"),
        "total_fill": PatternFill("solid", fgColor="FFEB99"),
        "revenue_fill": PatternFill("solid", fgColor="E2EFDA"),
        "profit_fill": PatternFill("solid", fgColor="DDEBF7"),
        "box": Border(left=border, right=border, top=border, bottom=border),
    }


def _detail_format(col: str) -> Optional[str]:
    if col in (COL_AVG_ACT_PRICE, COL_MAX_ACT_PRICE):
        return EUR_MWH_FMT
    if col == COL_BID:
        return EUR_MW_H_FMT
    if "(€)" in col:
        return EUR_FMT
    if "(MWh)" in col:
        return MWH_FMT
    if col == COL_ACCEPT:
        return PCT_FMT
    if col == COL_DATE:
        return "yyyy-mm-dd"
    return None


def _column_width(col: str) -> int:
    if col in (COL_DATE, COL_WEEKDAY, COL_SELECTED, COL_BATTERY_BOUND):
        return 12
    if "revenue" in col.lower():
        return 18
    if col in (COL_BID, COL_AVG_ACT_PRICE, COL_MAX_ACT_PRICE, COL_MKT_TOTAL, COL_CAPTURED):
        return 18
    if col == COL_SLOTS_ACTIVATED:
        return 14
    return 14


def _style_header(ws, styles):
    from openpyxl.styles import Alignment
    for cell in ws[1]:
        cell.font = styles["bold_white"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = styles["box"]
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "C2"


def _style_body(ws, df, styles, profit_cols, rev_cols):
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    for col_idx, col_name in enumerate(df.columns, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _column_width(col_name)
        nf = _detail_format(col_name)
        align = Alignment(horizontal="left" if col_name in (COL_DATE, COL_WEEKDAY) else "center")
        for cell in next(ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row)):
            if nf:
                cell.number_format = nf
            cell.alignment = align
            cell.border = styles["box"]
        fill = None
        if col_name in profit_cols:
            fill = styles["profit_fill"]
        elif col_name in rev_cols:
            fill = styles["revenue_fill"]
        if fill:
            for cell in next(ws.iter_cols(min_col=col_idx, max_col=col_idx,
                                          min_row=2, max_row=ws.max_row - 1)):
                cell.fill = fill


def _style_totals(ws, styles):
    for cell in ws[ws.max_row]:
        cell.font = styles["bold"]
        cell.fill = styles["total_fill"]
        cell.border = styles["box"]


def _style_detail(ws, df, styles):
    profit_cols = {COL_DAILY_REV}
    rev_cols = {COL_CAP_REV, COL_ACT_REV}
    _style_header(ws, styles)
    _style_body(ws, df, styles, profit_cols, rev_cols)
    _style_totals(ws, styles)


def _summary_title(ws, styles, first_date, last_date) -> None:
    from openpyxl.styles import Alignment
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = (
        f"aFRR daily revenue — REAL DAMAS market simulation (per-slot prices)  |  "
        f"{first_date} → {last_date}"
    )
    cell.font = styles["bold_big"]
    cell.fill = styles["title_fill"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    for col_letter, width in zip("ABCDEF", [38, 22, 22, 22, 22, 16]):
        ws.column_dimensions[col_letter].width = width


def _summary_battery(ws, styles, start_row: int) -> int:
    from openpyxl.styles import Alignment
    ws[f"A{start_row}"] = "Battery configuration"
    ws[f"A{start_row}"].font = styles["bold_white"]
    ws[f"A{start_row}"].fill = styles["header_fill"]
    ws.merge_cells(f"A{start_row}:F{start_row}")
    ws[f"A{start_row}"].alignment = Alignment(horizontal="center")
    params = [
        ("Power (MW)", POWER_MW, "Battery rated power"),
        ("Energy capacity (MWh)", CAPACITY_MWH, "Daily throughput cap (1 cycle/day)"),
        ("Slot length", "15 min", "DAMAS native resolution"),
        ("Slots per day", SLOTS_PER_DAY, "96 quarter-hours"),
        ("Market share", MARKET_SHARE, "Captured share of TSO activations (10% = small entrant)"),
        ("Per-slot max MWh", round(POWER_MW * SLOT_HOURS, 2),
         "= power × 0.25h. Hard limit per 15-min slot regardless of market size"),
        ("aFRR floor", AFRR_FLOOR_EUR_MW_H, "Catalog floor on bid"),
        ("aFRR+ bid cap", AFRR_UP_CAP_EUR_MW_H, "Hard ceiling on upward bid"),
        ("aFRR- bid cap", AFRR_DOWN_CAP_EUR_MW_H, "Hard ceiling on downward bid"),
        ("Price-validity range", f"|p| ∈ (0, {PRICE_VALID_MAX}) €/MWh", "Outlier filter on bid derivation"),
    ]
    for offset, (label, value, note) in enumerate(params, start=1):
        r = start_row + offset
        ws[f"A{r}"] = label
        ws[f"B{r}"] = value
        ws[f"C{r}"] = note
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws[f"A{r}"].font = styles["bold"]
        for col in ("A", "B", "C", "D", "E", "F"):
            ws[f"{col}{r}"].border = styles["box"]
    return start_row + len(params) + 2


def _summary_headline(ws, styles, start_row: int, totals: dict) -> int:
    from openpyxl.styles import Alignment
    ws[f"A{start_row}"] = f"Headline annual totals ({totals['days']} days)"
    ws[f"A{start_row}"].font = styles["bold_white"]
    ws[f"A{start_row}"].fill = styles["header_fill"]
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    ws[f"A{start_row}"].alignment = Alignment(horizontal="center")
    hdr = start_row + 1
    cols = ["Metric", "Conservative", "Balanced", "Aggressive", "Best – Worst"]
    for col_letter, label in zip("ABCDE", cols):
        c = ws[f"{col_letter}{hdr}"]
        c.value = label
        c.font = styles["bold_white"]
        c.fill = styles["title_fill"]
        c.alignment = Alignment(horizontal="center")
        c.border = styles["box"]
    ws.merge_cells(start_row=hdr, start_column=5, end_row=hdr, end_column=6)

    cap = totals["capacity_revenue"]
    act = totals["activation_revenue"]
    daily = totals["daily_revenue"]
    rows = [
        ("Capacity revenue (€)",   cap["conservative"],   cap["balanced"],   cap["aggressive"]),
        ("Activation revenue (€)", act["conservative"],   act["balanced"],   act["aggressive"]),
        ("Total annual revenue (€)", daily["conservative"], daily["balanced"], daily["aggressive"]),
        ("Avg daily revenue (€)",
         daily["conservative"] / totals["days"], daily["balanced"] / totals["days"],
         daily["aggressive"] / totals["days"]),
    ]
    for offset, (label, c_val, b_val, a_val) in enumerate(rows, start=1):
        r = hdr + offset
        ws[f"A{r}"] = label
        ws[f"B{r}"] = c_val
        ws[f"C{r}"] = b_val
        ws[f"D{r}"] = a_val
        ws[f"E{r}"] = max(c_val, b_val, a_val) - min(c_val, b_val, a_val)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws[f"A{r}"].font = styles["bold"]
        for col_letter in ("B", "C", "D", "E"):
            ws[f"{col_letter}{r}"].number_format = EUR_FMT
        for col in ("A", "B", "C", "D", "E", "F"):
            ws[f"{col}{r}"].border = styles["box"]
        if "Total" in label:
            for col in ("A", "B", "C", "D", "E", "F"):
                ws[f"{col}{r}"].fill = styles["total_fill"]
                ws[f"{col}{r}"].font = styles["bold"]
    return hdr + len(rows) + 1


def _summary_monthly(ws, styles, start_row: int, monthly_df: pd.DataFrame) -> int:
    from openpyxl.styles import Alignment
    ws[f"A{start_row}"] = "Monthly comparison"
    ws[f"A{start_row}"].font = styles["bold_white"]
    ws[f"A{start_row}"].fill = styles["header_fill"]
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    ws[f"A{start_row}"].alignment = Alignment(horizontal="center")
    hdr = start_row + 1
    headers = ["Month", "Days", "Conservative (€)", "Balanced (€)", "Aggressive (€)", "Spread (€)"]
    for col_letter, label in zip("ABCDEF", headers):
        c = ws[f"{col_letter}{hdr}"]
        c.value = label
        c.font = styles["bold_white"]
        c.fill = styles["title_fill"]
        c.alignment = Alignment(horizontal="center")
        c.border = styles["box"]
    r = hdr + 1
    for month, row in monthly_df.iterrows():
        ws[f"A{r}"] = month
        ws[f"B{r}"] = int(row["days"])
        ws[f"C{r}"] = float(row["conservative"])
        ws[f"D{r}"] = float(row["balanced"])
        ws[f"E{r}"] = float(row["aggressive"])
        ws[f"F{r}"] = float(row["spread"])
        ws[f"A{r}"].font = styles["bold"]
        for col_letter in ("C", "D", "E", "F"):
            ws[f"{col_letter}{r}"].number_format = EUR_FMT
        for col in ("A", "B", "C", "D", "E", "F"):
            ws[f"{col}{r}"].border = styles["box"]
            ws[f"{col}{r}"].alignment = Alignment(
                horizontal="center" if col != "A" else "left",
            )
        r += 1
    ws[f"A{r}"] = "TOTAL"
    ws[f"B{r}"] = int(monthly_df["days"].sum())
    ws[f"C{r}"] = float(monthly_df["conservative"].sum())
    ws[f"D{r}"] = float(monthly_df["balanced"].sum())
    ws[f"E{r}"] = float(monthly_df["aggressive"].sum())
    ws[f"F{r}"] = float(monthly_df["spread"].sum())
    for col_letter in ("C", "D", "E", "F"):
        ws[f"{col_letter}{r}"].number_format = EUR_FMT
    for col in ("A", "B", "C", "D", "E", "F"):
        ws[f"{col}{r}"].font = styles["bold"]
        ws[f"{col}{r}"].fill = styles["total_fill"]
        ws[f"{col}{r}"].border = styles["box"]
        ws[f"{col}{r}"].alignment = Alignment(
            horizontal="center" if col != "A" else "left",
        )
    return r + 2


def _summary_notes(ws, styles, start_row: int) -> None:
    from openpyxl.styles import Alignment
    ws[f"A{start_row}"] = "Notes"
    ws[f"A{start_row}"].font = styles["bold_white"]
    ws[f"A{start_row}"].fill = styles["header_fill"]
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    ws[f"A{start_row}"].alignment = Alignment(horizontal="center")
    notes = [
        "Data source: DAMAS (Romanian TSO), 15-min slot data, loaded via fr_service._load_fr_data(). 96 slots/day × 366 days = 35,136 slots simulated per strategy.",
        "Per-slot logic: for each 15-min slot the TSO actually activated (afrr_*_activated_mwh > 0), our captured energy = min(market_MWh × 10% share × acceptance_rate, power × 0.25h, remaining battery_capacity). Activation revenue = captured × actual settlement price (afrr_*_price_eur). Battery cap is a RUNNING budget across the day — once 20 MWh delivered, later slot activations are skipped (Battery cap binding column flags this).",
        "Bid (recommended) is still derived from the day's price quantile (1 − acceptance) — that's the price you SUBMIT. Activation revenue uses the REAL settled price the TSO paid for each slot you serviced.",
        "Capacity revenue = power × 24h × bid × acceptance_rate. Standard 'paid for availability' leg, same in production.",
        "Direction selection: model picks aFRR+ or aFRR- per day, whichever yields higher (capacity + activation) revenue. Real operations could split bids, but battery throughput has only one daily budget — committing to one direction is the conservative model.",
        "Bid acceptance is treated as binary at the daily level via acceptance_rate scaling. Production has the same simplification.",
        "Numbers are SIMULATED on real DAMAS data — NOT bankable. Compliance gates (ANRE / OPCOM / PRE / BRP / BSP / FSE) and tariff exemption math live downstream in investment_service.analyze(). DAMAS source_kind is 'unverified_snapshot', not settlement-grade.",
    ]
    r = start_row
    for note in notes:
        r += 1
        ws[f"A{r}"] = "• " + note
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 38


# ===================================================================
#  Main
# ===================================================================
def _total_row(df: pd.DataFrame) -> dict:
    return {
        COL_DATE: "TOTAL", COL_WEEKDAY: f"{len(df)} days",
        COL_SELECTED: PLACEHOLDER_DASH,
        COL_BID: None, COL_ACCEPT: None,
        COL_HOURS: int(df[COL_HOURS].sum()),
        COL_CAP_REV: float(df[COL_CAP_REV].sum()),
        COL_SLOTS_ACTIVATED: int(df[COL_SLOTS_ACTIVATED].sum()),
        COL_AVG_ACT_PRICE: None, COL_MAX_ACT_PRICE: None,
        COL_MKT_TOTAL: float(df[COL_MKT_TOTAL].sum()),
        COL_CAPTURED: float(df[COL_CAPTURED].sum()),
        COL_BATTERY_BOUND: f"{(df[COL_BATTERY_BOUND] == 'yes').sum()} days",
        COL_ACT_REV: float(df[COL_ACT_REV].sum()),
        COL_DAILY_REV: float(df[COL_DAILY_REV].sum()),
    }


def _build_strategy_df(df: pd.DataFrame, strategy: str) -> pd.DataFrame:
    rows = []
    for day, group in df.groupby(df["date"].dt.date):
        row = daily_afrr_row(group, day, strategy)
        rows.append(row if row is not None else empty_row(day))
    return pd.DataFrame(rows).sort_values(COL_DATE).reset_index(drop=True)


def _build_monthly_comparison(strategy_dfs: dict) -> pd.DataFrame:
    base = strategy_dfs["balanced"].copy()
    base["Month"] = base[COL_DATE].dt.strftime("%Y-%m")
    days_per_month = base.groupby("Month").size()
    monthly = pd.DataFrame({"days": days_per_month})
    for strat, sdf in strategy_dfs.items():
        m = sdf.copy()
        m["Month"] = m[COL_DATE].dt.strftime("%Y-%m")
        monthly[strat] = m.groupby("Month")[COL_DAILY_REV].sum()
    monthly["spread"] = (
        monthly[["conservative", "balanced", "aggressive"]].max(axis=1)
        - monthly[["conservative", "balanced", "aggressive"]].min(axis=1)
    )
    return monthly


def style_workbook(out_path: Path, balanced_df: pd.DataFrame, aggressive_df: pd.DataFrame,
                   conservative_df: pd.DataFrame, monthly_df: pd.DataFrame,
                   first_date, last_date) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    styles = _styles()

    _style_detail(wb[SHEET_BALANCED], balanced_df, styles)
    _style_detail(wb[SHEET_AGGRESSIVE], aggressive_df, styles)

    totals = {
        "days": len(balanced_df),
        "capacity_revenue": {s: df[COL_CAP_REV].sum() for s, df in [
            ("conservative", conservative_df), ("balanced", balanced_df),
            ("aggressive", aggressive_df)]},
        "activation_revenue": {s: df[COL_ACT_REV].sum() for s, df in [
            ("conservative", conservative_df), ("balanced", balanced_df),
            ("aggressive", aggressive_df)]},
        "daily_revenue": {s: df[COL_DAILY_REV].sum() for s, df in [
            ("conservative", conservative_df), ("balanced", balanced_df),
            ("aggressive", aggressive_df)]},
    }

    ws = wb[SHEET_SUMMARY]
    _summary_title(ws, styles, first_date, last_date)
    next_row = _summary_battery(ws, styles, start_row=3)
    next_row = _summary_headline(ws, styles, next_row, totals)
    next_row = _summary_monthly(ws, styles, next_row, monthly_df)
    _summary_notes(ws, styles, next_row)
    wb.active = wb.sheetnames.index(SHEET_SUMMARY)
    wb.save(out_path)


def _load_fr_window(fr: FRService) -> Tuple[pd.DataFrame, pd.Timestamp, pd.Timestamp]:
    df = fr._load_fr_data().copy()
    last_date = df["date"].max().date()
    first_date = (pd.Timestamp(last_date) - pd.Timedelta(days=365)).date()
    df = df[(df["date"].dt.date >= first_date) & (df["date"].dt.date <= last_date)]
    return df, first_date, last_date


def main() -> None:
    fr = FRService()
    df, first_date, last_date = _load_fr_window(fr)
    print(f"Window: {first_date} → {last_date}  ({df['date'].dt.date.nunique()} days)")
    print(f"Battery: {POWER_MW} MW / {CAPACITY_MWH} MWh, slot-level real-market simulation")

    balanced_df = _build_strategy_df(df, "balanced")
    aggressive_df = _build_strategy_df(df, "aggressive")
    conservative_df = _build_strategy_df(df, "conservative")

    monthly_df = _build_monthly_comparison({
        "balanced": balanced_df,
        "aggressive": aggressive_df,
        "conservative": conservative_df,
    })

    balanced_out = pd.concat(
        [balanced_df, pd.DataFrame([_total_row(balanced_df)])], ignore_index=True)
    aggressive_out = pd.concat(
        [aggressive_df, pd.DataFrame([_total_row(aggressive_df)])], ignore_index=True)

    out_dir = ROOT / "reports"
    out_dir.mkdir(exist_ok=True)
    out_path = (
        out_dir
        / f"aFRR_daily_real_DAMAS_anchored_bids_{first_date}_to_{last_date}.xlsx"
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        pd.DataFrame().to_excel(xl, sheet_name=SHEET_SUMMARY, index=False)
        balanced_out.to_excel(xl, sheet_name=SHEET_BALANCED, index=False)
        aggressive_out.to_excel(xl, sheet_name=SHEET_AGGRESSIVE, index=False)

    style_workbook(out_path, balanced_df, aggressive_df, conservative_df,
                   monthly_df, first_date, last_date)

    print(f"\nWrote: {out_path}")
    print(f"  Sheet '{SHEET_SUMMARY}'        — battery params + headline + monthly + notes")
    print(f"  Sheet '{SHEET_BALANCED}'    — {len(balanced_df)} daily rows + totals row")
    print(f"  Sheet '{SHEET_AGGRESSIVE}'  — {len(aggressive_df)} daily rows + totals row")
    print("\nHeadline annual totals (real per-slot market simulation):")
    for strat, sdf in [("conservative", conservative_df),
                       ("balanced", balanced_df),
                       ("aggressive", aggressive_df)]:
        cap = sdf[COL_CAP_REV].sum()
        act = sdf[COL_ACT_REV].sum()
        total = sdf[COL_DAILY_REV].sum()
        bound_days = (sdf[COL_BATTERY_BOUND] == "yes").sum()
        print(
            f"  {strat:12s}: capacity €{cap:>10,.0f}  + activation €{act:>10,.0f}  "
            f"= €{total:>10,.0f}   battery cap binding on {bound_days}/{len(sdf)} days",
        )


if __name__ == "__main__":
    main()
