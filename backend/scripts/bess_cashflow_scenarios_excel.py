"""15-year BESS cashflow Excel with 4 scenarios (Goldman-grade format).

Combines aFRR (real DAMAS-anchored) + PZU (real OPCOM-derived) revenue
into a full project finance view: 15-year cashflow, debt service,
augmentation, equity / project IRR, NPV @ 8%, payback, DSCR per year.

Sheet layout:
  1. Executive_Summary           — headline KPIs across 4 scenarios
  2. Inputs_and_Assumptions      — battery, financing, scenario params
  3. Scenario_A_Current          — 15yr cashflow detail
  4. Scenario_B_PICASSO          — 15yr cashflow detail
  5. Scenario_C_Mature           — 15yr cashflow detail
  6. Scenario_D_Bear             — 15yr cashflow detail
  7. Scenarios_Comparison        — side-by-side annual + KPI comparison

Run from `backend/`:
    PYTHONPATH=. arch -arm64 /usr/local/bin/python3 scripts/bess_cashflow_scenarios_excel.py
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

# ===================================================================
#  Project finance inputs (mirror InvestmentParams defaults)
# ===================================================================
EPC_INVESTMENT_EUR = 3_500_000.0          # 10 MW / 20 MWh @ €175/kWh
POWER_MW = 10.0
CAPACITY_MWH = 20.0
EQUITY_PCT = 30.0                          # %
LOAN_PCT = 70.0                            # %
LOAN_RATE = 0.06                           # 6%
LOAN_TERM_YR = 10
OPEX_PCT_OF_EPC = 2.0                      # 2% of EPC = €70k/yr Y1
INSURANCE_PCT_OF_EPC = 0.5                 # 0.5% = €17.5k/yr Y1
OPEX_INFLATION = 0.03                      # 3% annually
DEGRADATION_RATE = 0.025                   # 2.5% capacity fade/yr
AUGMENTATION_PCT_OF_EPC = 10.0             # 10% at Y5 and Y10
AUGMENTATION_YEARS = (5, 10)
PROJECT_LIFE_YR = 15
DISCOUNT_RATE = 0.08                       # 8% for NPV
DSCR_COVENANT = 1.20                       # lender covenant typical
# Tax & depreciation (Romania)
TAX_RATE = 0.16                            # 16% Romanian corporate income tax
DEPRECIATION_LIFE_YR = 8                   # 8-year straight-line for tax (BESS class)

# ===================================================================
#  STACKING FACTORS — physical hardware constraint
# ===================================================================
# A 10 MW / 20 MWh battery has ONE inverter, ONE energy budget per
# cycle, and ONE SOC at any instant. You CANNOT bid the same MW into
# both aFRR and PZU at the same time. Real operators allocate the
# battery between products per-slot, subject to MW + MWh + SOC
# constraints. Empirical European BESS data (Modo Energy UK BESS
# league tables, Magnus Energy NL/DE, Field Energy disclosures) shows
# stacked revenue ≈ 65-75% of the naive sum-of-products. The split
# below applies the stacking constraint per revenue leg:
#
#   - aFRR CAPACITY revenue: 1.00 (FULL). It's pure availability
#     reservation — you commit MW but don't consume MWh, so it sits
#     alongside other products without conflict.
#   - aFRR ACTIVATION revenue: 0.60. Activation consumes MWh from the
#     same daily throughput budget that PZU uses. You can serve
#     ~60% of the activations you'd see in a pure-aFRR commitment
#     because you've reserved part of the cycle for PZU dispatch.
#   - PZU revenue: 0.50. With aFRR holding capacity reservation, you
#     lose dispatch flexibility around peak windows where TSO might
#     activate — typically ~50% of pure-PZU revenue.
#
# Net stacked revenue ≈ 1.0×cap + 0.6×act + 0.5×pzu
# Total stacking efficiency ≈ 0.62 of naive sum (matches the empirical
# 0.65-0.75 European range; conservative side because Romania's
# bid-stack is thinner so capacity reservation costs more dispatch
# flexibility).
#
# To model a SINGLE-PRODUCT operator (no stacking), set all to 1.0
# AND zero the products you don't run. To model a perfect joint
# dispatch optimizer (theoretical max), set all to 1.0 and accept a
# ~10% optimization gap haircut elsewhere.
STACK_AFRR_CAPACITY = 1.00
STACK_AFRR_ACTIVATION = 0.60
STACK_PZU = 0.50


# ===================================================================
#  REALISTIC OPERATOR DRAG MULTIPLIERS (orthogonal to stacking)
# ===================================================================
# Stacking handles the PHYSICAL constraint. These OPERATIONAL drags
# apply on top — they're what real operators consistently lose vs the
# theoretical optimum:
#   - aFRR activation 10% market share assumption → realistic ~2.5%
#     → drag 0.25 (the dominant correction).
#   - PZU forecast error 15-25% vs perfect-foresight backtest → 0.80.
#   - aFRR capacity REMIT/Tg/Td tariff erosion → 0.93.
DRAG_AFRR_CAPACITY = 0.93
DRAG_AFRR_ACTIVATION = 0.25
DRAG_PZU = 0.80


# ===================================================================
#  Scenario base Y1 revenues (from real DAMAS / OPCOM data, balanced strategy)
# ===================================================================
# aFRR Y1 numbers come from afrr_scenarios_excel.py (real-market sim).
# PZU Y1 numbers: scenario A = €835k from production simulate(); other
# scenarios apply progressive compression to that base.
# These are ENGINEERING-OPTIMUM — see DRAG_* multipliers above.
SCENARIOS = [
    {
        "key": "A",
        "label": "A. Current (10% share, pre-PICASSO)",
        "short": "Scenario_A_Current",
        "color": "DDEBF7",
        # Capacity Y1 refreshed against today's canonical resolver
        # (€6.57 aFRRUp + €3.77 aFRRDown, recent-window mean):
        # 10 MW × 24h × ~€5.17 avg × 0.80 acceptance × 366 days ≈ €363k
        # (was €597k under the older €8.72 mean, before recent-window cutoff)
        "afrr_capacity_y1": 363_000,
        "afrr_activation_y1": 2_675_941,  # see DRAG_AFRR_ACTIVATION for realistic
        "pzu_y1": 835_187,
        # Compression schedule for years 2..15
        "compression": {
            "afrr_capacity":   [1.00, 1.00, 0.75, 0.69, 0.63, 0.58, 0.55, 0.55, 0.55, 0.55, 0.55, 0.55, 0.55, 0.55, 0.55],
            "afrr_activation": [1.00, 1.00, 0.75, 0.70, 0.66, 0.62, 0.58, 0.55, 0.55, 0.55, 0.55, 0.55, 0.55, 0.55, 0.55],
            "pzu":             [1.00, 0.97, 0.92, 0.87, 0.82, 0.77, 0.72, 0.68, 0.64, 0.60, 0.57, 0.54, 0.51, 0.48, 0.46],
        },
        "description": (
            "Y1 reflects today's Romanian state: 10% market share, pre-PICASSO. "
            "PICASSO assumed to go live at Y3 → -25% step on capacity + activation. "
            "Standard compression schedule (8%/yr from Y3) bottoms at 55% of Y1 by Y8."
        ),
    },
    {
        "key": "B",
        "label": "B. PICASSO go-live (Y1) (10% share)",
        "short": "Scenario_B_PICASSO",
        "color": "FCE4D6",
        "afrr_capacity_y1": 448_725,       # current × 0.75
        "afrr_activation_y1": 2_006_956,   # current × 0.75
        "pzu_y1": 835_187,                  # PZU largely independent of PICASSO
        "compression": {
            "afrr_capacity":   [1.00, 1.00, 0.95, 0.90, 0.86, 0.82, 0.78, 0.74, 0.71, 0.70, 0.70, 0.70, 0.70, 0.70, 0.70],
            "afrr_activation": [1.00, 1.00, 0.95, 0.90, 0.86, 0.82, 0.78, 0.74, 0.71, 0.70, 0.70, 0.70, 0.70, 0.70, 0.70],
            "pzu":             [1.00, 0.97, 0.92, 0.87, 0.82, 0.77, 0.72, 0.68, 0.64, 0.60, 0.57, 0.54, 0.51, 0.48, 0.46],
        },
        "description": (
            "PICASSO already live at Y1 (alternate timeline if Hungary connects faster). "
            "Y1 base already incorporates -25% from current. Slower compression "
            "(~5%/yr) post-PICASSO since extreme tail already capped. Floor at 70% Y1."
        ),
    },
    {
        "key": "C",
        "label": "C. Mature market (5% share, -40%)",
        "short": "Scenario_C_Mature",
        "color": "FFF2CC",
        "afrr_capacity_y1": 363_911,       # 5% share, -40% prices
        "afrr_activation_y1": 1_836_060,
        "pzu_y1": 626_390,                  # PZU also compressed -25%
        "compression": {
            "afrr_capacity":   [1.00, 0.97, 0.94, 0.91, 0.89, 0.87, 0.85, 0.85, 0.85, 0.85, 0.85, 0.85, 0.85, 0.85, 0.85],
            "afrr_activation": [1.00, 0.97, 0.94, 0.91, 0.89, 0.87, 0.85, 0.85, 0.85, 0.85, 0.85, 0.85, 0.85, 0.85, 0.85],
            "pzu":             [1.00, 0.98, 0.95, 0.92, 0.89, 0.86, 0.83, 0.80, 0.77, 0.74, 0.71, 0.68, 0.65, 0.62, 0.59],
        },
        "description": (
            "Realistic Y3-Y8 lender baseline. Romanian BESS pipeline (1+ GW announced) "
            "materializes 2027-2028. Market share compresses to 5%, capacity prices to "
            "Germany current level (~€5/MW/h). Mild ongoing compression to floor of 85% Y1."
        ),
    },
    {
        "key": "D",
        "label": "D. Bear case (2% share, full saturation)",
        "short": "Scenario_D_Bear",
        "color": "F8CBAD",
        "afrr_capacity_y1": 351_360,       # 2% share, -60% prices
        "afrr_activation_y1": 1_035_773,
        "pzu_y1": 417_594,                  # PZU also compressed -50%
        "compression": {
            "afrr_capacity":   [1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00],
            "afrr_activation": [1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00],
            "pzu":             [1.00, 0.97, 0.94, 0.91, 0.88, 0.85, 0.82, 0.80, 0.78, 0.76, 0.74, 0.72, 0.70, 0.68, 0.66],
        },
        "description": (
            "Long-term steady state stress case (2030+). 2-3 GW BESS deployed Romania. "
            "Already at floor — no further compression. Used by lenders for late-life "
            "(Y8-Y15) DSCR projection."
        ),
    },
]

# ===================================================================
#  Financial calculations
# ===================================================================
def annuity_payment(principal: float, rate: float, n: int) -> float:
    """Standard amortizing loan annual payment (PMT)."""
    if rate == 0:
        return principal / n
    return principal * rate / (1 - (1 + rate) ** -n)


def amortization_schedule(principal: float, rate: float, n: int) -> list[dict]:
    """Standard amortizing loan: list of (interest, principal_paid, balance_end)
    per year. Length = n. After year n, balance = 0."""
    pmt = annuity_payment(principal, rate, n)
    balance = principal
    schedule = []
    for _ in range(n):
        interest = balance * rate
        principal_paid = pmt - interest
        balance = max(0.0, balance - principal_paid)
        schedule.append({
            "interest": interest,
            "principal": principal_paid,
            "payment_total": pmt,
            "balance_end": balance,
        })
    return schedule


def mirr(cashflows: list[float], finance_rate: float, reinvest_rate: float) -> Optional[float]:
    """Modified IRR. Negative cashflows discounted at finance_rate;
    positive cashflows compounded forward at reinvest_rate. Lender-grade
    metric — handles non-conventional cashflows that simple IRR misbehaves on.
    """
    n = len(cashflows) - 1
    if n <= 0:
        return None
    pv_neg = sum(cf / (1 + finance_rate) ** t for t, cf in enumerate(cashflows) if cf < 0)
    fv_pos = sum(cf * (1 + reinvest_rate) ** (n - t) for t, cf in enumerate(cashflows) if cf > 0)
    if pv_neg == 0 or fv_pos <= 0:
        return None
    return (fv_pos / -pv_neg) ** (1 / n) - 1


def discounted_payback(cashflows: list[float], rate: float) -> Optional[float]:
    """Year (with linear interpolation) where DISCOUNTED cumulative CF crosses 0.
    cashflows[0] at Y0, cashflows[1] at Y1, ..."""
    cum = 0.0
    cum_series = []
    for t, cf in enumerate(cashflows):
        cum += cf / (1 + rate) ** t
        cum_series.append(cum)
    return payback_year(cum_series)


def npv(cashflows: list[float], rate: float) -> float:
    """NPV with cashflows[0] at Y0, cashflows[1] at Y1, etc."""
    return sum(cf / (1 + rate) ** t for t, cf in enumerate(cashflows))


def irr(cashflows: list[float], guess: float = 0.10) -> Optional[float]:
    """IRR via Newton-Raphson; returns None if no solution found."""
    if not cashflows or all(cf >= 0 for cf in cashflows) or all(cf <= 0 for cf in cashflows):
        return None
    rate = guess
    for _ in range(200):
        f = sum(cf / (1 + rate) ** t for t, cf in enumerate(cashflows))
        df = sum(-t * cf / (1 + rate) ** (t + 1) for t, cf in enumerate(cashflows))
        if abs(df) < 1e-10:
            break
        new = rate - f / df
        if abs(new - rate) < 1e-7:
            return new
        rate = new
    return rate if -0.99 < rate < 10 else None


def payback_year(cumulative_cashflows: list[float]) -> Optional[float]:
    """First year (with linear interpolation) where cumulative CF crosses 0."""
    for i, cf in enumerate(cumulative_cashflows):
        if cf >= 0:
            if i == 0:
                return 0.0
            prev = cumulative_cashflows[i - 1]
            if prev < 0:
                return (i - 1) + (-prev) / (cf - prev)
            return float(i)
    return None


def build_cashflow(scenario: dict, apply_drag: bool = False) -> pd.DataFrame:
    """Build the 15-year per-year project finance cashflow table.

    Columns produced (Goldman / lender-grade):
      - Revenue: aFRR capacity, aFRR activation, PZU arbitrage, Total
      - OPEX: OPEX, Insurance, Augmentation
      - EBITDA + EBITDA margin %
      - Depreciation (8yr SL on EPC) + EBIT
      - Interest expense + Pre-tax income (EBT) + Tax + Net income
      - Debt: Principal payment + Total debt service + Debt outstanding
      - DSCR (= EBITDA / debt service)
      - Free Cash Flow (= EBITDA - tax - augmentation)
      - Equity FCF (= FCF - debt service)
      - Cumulative equity FCF
      - ROE (= net income / equity)
      - ROIC (= NOPAT / invested capital)
      - Cash-on-cash yield (= equity FCF / equity)

    `apply_drag=True` applies realistic operator-side multipliers
    (DRAG_AFRR_CAPACITY, DRAG_AFRR_ACTIVATION, DRAG_PZU). STACKING
    factors apply unconditionally — physical hardware constraint.
    """
    equity_eur = EPC_INVESTMENT_EUR * EQUITY_PCT / 100.0
    loan_eur = EPC_INVESTMENT_EUR * LOAN_PCT / 100.0
    debt_schedule = amortization_schedule(loan_eur, LOAN_RATE, LOAN_TERM_YR)
    annual_depreciation = EPC_INVESTMENT_EUR / DEPRECIATION_LIFE_YR

    cap_drag = DRAG_AFRR_CAPACITY if apply_drag else 1.0
    act_drag = DRAG_AFRR_ACTIVATION if apply_drag else 1.0
    pzu_drag = DRAG_PZU if apply_drag else 1.0

    rows = []
    cum_equity_fcf = 0.0
    cum_proj_fcf = 0.0
    debt_outstanding = loan_eur

    for y in range(1, PROJECT_LIFE_YR + 1):
        i = y - 1
        deg = (1 - DEGRADATION_RATE) ** i
        afrr_cap = (scenario["afrr_capacity_y1"]
                    * scenario["compression"]["afrr_capacity"][i] * deg
                    * STACK_AFRR_CAPACITY * cap_drag)
        afrr_act = (scenario["afrr_activation_y1"]
                    * scenario["compression"]["afrr_activation"][i] * deg
                    * STACK_AFRR_ACTIVATION * act_drag)
        pzu = (scenario["pzu_y1"]
               * scenario["compression"]["pzu"][i] * deg
               * STACK_PZU * pzu_drag)
        revenue = afrr_cap + afrr_act + pzu

        opex = (EPC_INVESTMENT_EUR * OPEX_PCT_OF_EPC / 100.0) * (1 + OPEX_INFLATION) ** i
        insurance = (EPC_INVESTMENT_EUR * INSURANCE_PCT_OF_EPC / 100.0) * (1 + OPEX_INFLATION) ** i
        augmentation = (EPC_INVESTMENT_EUR * AUGMENTATION_PCT_OF_EPC / 100.0) if y in AUGMENTATION_YEARS else 0.0

        ebitda = revenue - opex - insurance
        ebitda_margin = (ebitda / revenue * 100.0) if revenue > 0 else 0.0

        # Depreciation only for first DEPRECIATION_LIFE_YR years
        depreciation = annual_depreciation if y <= DEPRECIATION_LIFE_YR else 0.0
        ebit = ebitda - depreciation

        # Debt: split annuity into interest + principal
        if y <= LOAN_TERM_YR:
            interest_exp = debt_schedule[i]["interest"]
            principal_payment = debt_schedule[i]["principal"]
            debt_service = debt_schedule[i]["payment_total"]
            debt_outstanding = debt_schedule[i]["balance_end"]
        else:
            interest_exp = 0.0
            principal_payment = 0.0
            debt_service = 0.0
            debt_outstanding = 0.0

        # Income statement
        ebt = ebit - interest_exp
        tax = max(0.0, ebt * TAX_RATE)  # no tax credit for losses (conservative)
        net_income = ebt - tax

        # Cashflow
        # FCF = EBITDA - tax - augmentation (project cashflow available to all)
        project_fcf = ebitda - tax - augmentation
        # Equity FCF = FCF - debt service (what equity sees)
        equity_fcf = project_fcf - debt_service
        cum_equity_fcf += equity_fcf
        cum_proj_fcf += project_fcf

        # Returns ratios
        invested_capital = EPC_INVESTMENT_EUR  # could amortize but standard project finance keeps it static
        nopat = ebit * (1 - TAX_RATE)  # NOPAT = EBIT × (1 - tax)
        roe_pct = (net_income / equity_eur * 100.0) if equity_eur > 0 else 0.0
        roic_pct = (nopat / invested_capital * 100.0) if invested_capital > 0 else 0.0
        coc_yield_pct = (equity_fcf / equity_eur * 100.0) if equity_eur > 0 else 0.0

        # DSCR = EBITDA / debt_service (lender standard)
        dscr = (ebitda / debt_service) if debt_service > 0 else None

        rows.append({
            "Year": y,
            # Revenue
            "aFRR capacity (€)": afrr_cap,
            "aFRR activation (€)": afrr_act,
            "PZU arbitrage (€)": pzu,
            "Total revenue (€)": revenue,
            # Operating expenses
            "OPEX (€)": -opex,
            "Insurance (€)": -insurance,
            # EBITDA
            "EBITDA (€)": ebitda,
            "EBITDA margin (%)": ebitda_margin,
            # Depreciation + EBIT
            "Depreciation (€)": -depreciation,
            "EBIT (€)": ebit,
            # Interest + tax + net income
            "Interest expense (€)": -interest_exp,
            "EBT (€)": ebt,
            "Tax (€)": -tax,
            "Net income (€)": net_income,
            # Debt
            "Principal payment (€)": -principal_payment,
            "Debt service total (€)": -debt_service if debt_service > 0 else 0.0,
            "Debt outstanding YE (€)": debt_outstanding,
            "DSCR": dscr,
            # Augmentation
            "Augmentation (€)": -augmentation if augmentation > 0 else 0.0,
            # Cashflow
            "Project FCF (€)": project_fcf,
            "Equity FCF (€)": equity_fcf,
            "Cumulative equity FCF (€)": cum_equity_fcf,
            # Returns
            "ROE (%)": roe_pct,
            "ROIC (%)": roic_pct,
            "Cash-on-cash yield (%)": coc_yield_pct,
        })

    return pd.DataFrame(rows)


def compute_kpis(scenario: dict, df: pd.DataFrame) -> dict:
    """Goldman-grade KPI suite: profitability, cashflow, returns,
    lender coverage. Computed once per (scenario × view)."""
    equity_eur = EPC_INVESTMENT_EUR * EQUITY_PCT / 100.0
    loan_eur = EPC_INVESTMENT_EUR * LOAN_PCT / 100.0

    # Cashflow series (Y0 to Y15)
    equity_cf = [-equity_eur] + df["Equity FCF (€)"].tolist()
    project_cf = [-EPC_INVESTMENT_EUR] + df["Project FCF (€)"].tolist()

    # IRR / MIRR / NPV
    eq_irr = irr(equity_cf)
    pj_irr = irr(project_cf)
    eq_mirr = mirr(equity_cf, finance_rate=LOAN_RATE, reinvest_rate=DISCOUNT_RATE)
    pj_mirr = mirr(project_cf, finance_rate=LOAN_RATE, reinvest_rate=DISCOUNT_RATE)

    eq_npv_8 = npv(equity_cf, 0.08)
    eq_npv_10 = npv(equity_cf, 0.10)
    eq_npv_12 = npv(equity_cf, 0.12)
    pj_npv_8 = npv(project_cf, 0.08)
    pj_npv_10 = npv(project_cf, 0.10)
    pj_npv_12 = npv(project_cf, 0.12)

    # Paybacks (simple + discounted)
    cum_equity = list(pd.Series(equity_cf).cumsum())
    cum_project = list(pd.Series(project_cf).cumsum())
    eq_payback = payback_year(cum_equity)
    pj_payback = payback_year(cum_project)
    eq_disc_payback = discounted_payback(equity_cf, DISCOUNT_RATE)
    pj_disc_payback = discounted_payback(project_cf, DISCOUNT_RATE)

    # MOIC (Money-on-Invested-Capital) = sum of equity inflows / equity outflow
    total_equity_inflow = sum(cf for cf in equity_cf if cf > 0)
    moic = total_equity_inflow / equity_eur if equity_eur > 0 else None

    # DSCR statistics
    dscrs = [d for d in df["DSCR"] if d is not None]
    avg_dscr = sum(dscrs) / len(dscrs) if dscrs else None
    min_dscr = min(dscrs) if dscrs else None
    dscr_breach_years = [y for y, dscr in zip(df["Year"], df["DSCR"])
                         if dscr is not None and dscr < DSCR_COVENANT]

    # LLCR (Loan Life Coverage Ratio) = NPV(CFADS during debt life) / debt outstanding at start
    # CFADS = EBITDA - Tax - Augmentation (cash available for debt service)
    cfads = (df["EBITDA (€)"] + df["Tax (€)"] + df["Augmentation (€)"]).tolist()
    cfads_during_debt = cfads[:LOAN_TERM_YR]
    llcr_num = sum(cf / (1 + LOAN_RATE) ** (t + 1) for t, cf in enumerate(cfads_during_debt))
    llcr = llcr_num / loan_eur if loan_eur > 0 else None

    # PLCR (Project Life Coverage Ratio) = NPV(CFADS over project life) / debt outstanding at start
    plcr_num = sum(cf / (1 + LOAN_RATE) ** (t + 1) for t, cf in enumerate(cfads))
    plcr = plcr_num / loan_eur if loan_eur > 0 else None

    # Average ROE / ROIC
    avg_roe = float(df["ROE (%)"].mean())
    avg_roic = float(df["ROIC (%)"].mean())

    # Average EBITDA margin
    avg_ebitda_margin = float(df["EBITDA margin (%)"].mean())

    # Profitability totals
    lifetime_revenue = float(df["Total revenue (€)"].sum())
    lifetime_ebitda = float(df["EBITDA (€)"].sum())
    lifetime_net_income = float(df["Net income (€)"].sum())
    lifetime_equity_fcf = float(df["Equity FCF (€)"].sum())
    lifetime_project_fcf = float(df["Project FCF (€)"].sum())

    return {
        # Y1 snapshot
        "y1_revenue_eur": float(df.iloc[0]["Total revenue (€)"]),
        "y1_ebitda_eur": float(df.iloc[0]["EBITDA (€)"]),
        "y1_ebitda_margin_pct": float(df.iloc[0]["EBITDA margin (%)"]),
        "y1_ebit_eur": float(df.iloc[0]["EBIT (€)"]),
        "y1_net_income_eur": float(df.iloc[0]["Net income (€)"]),
        "y1_equity_fcf_eur": float(df.iloc[0]["Equity FCF (€)"]),
        "y1_dscr": (df.iloc[0]["DSCR"] if df.iloc[0]["DSCR"] is not None else None),
        # Lifetime
        "lifetime_revenue_eur": lifetime_revenue,
        "lifetime_ebitda_eur": lifetime_ebitda,
        "lifetime_net_income_eur": lifetime_net_income,
        "lifetime_equity_fcf_eur": lifetime_equity_fcf,
        "lifetime_project_fcf_eur": lifetime_project_fcf,
        "avg_ebitda_margin_pct": avg_ebitda_margin,
        # Returns
        "equity_irr_pct": (eq_irr * 100.0) if eq_irr is not None else None,
        "project_irr_pct": (pj_irr * 100.0) if pj_irr is not None else None,
        "equity_mirr_pct": (eq_mirr * 100.0) if eq_mirr is not None else None,
        "project_mirr_pct": (pj_mirr * 100.0) if pj_mirr is not None else None,
        "equity_npv_8_eur": eq_npv_8,
        "equity_npv_10_eur": eq_npv_10,
        "equity_npv_12_eur": eq_npv_12,
        "project_npv_8_eur": pj_npv_8,
        "project_npv_10_eur": pj_npv_10,
        "project_npv_12_eur": pj_npv_12,
        "moic": moic,
        "avg_roe_pct": avg_roe,
        "avg_roic_pct": avg_roic,
        # Payback
        "equity_payback_years": eq_payback,
        "project_payback_years": pj_payback,
        "equity_discounted_payback_years": eq_disc_payback,
        "project_discounted_payback_years": pj_disc_payback,
        # Lender / coverage
        "min_dscr": min_dscr,
        "avg_dscr": avg_dscr,
        "llcr": llcr,
        "plcr": plcr,
        "dscr_breach_years": dscr_breach_years,
        # Compatibility (legacy callers)
        "equity_npv_eur": eq_npv_8,  # default = NPV @ 8%
        "project_npv_eur": pj_npv_8,
        "lifetime_cf_after_debt_eur": lifetime_equity_fcf,
        "y1_cf_after_debt_eur": float(df.iloc[0]["Equity FCF (€)"]),
    }


# ===================================================================
#  Excel formatting
# ===================================================================
EUR_FMT = '"€"#,##0;[Red]"-€"#,##0'
EUR_K_FMT = '"€"#,##0,"k";[Red]"-€"#,##0,"k"'
PCT_FMT = '0.00%'
PCT_FMT_INT = '0.0%'
DSCR_FMT = '0.00"x"'
INT_FMT = '0'
YR_FMT = '0.0" yr"'


def _styles():
    from openpyxl.styles import Border, Font, PatternFill, Side
    border = Side(border_style="thin", color="BFBFBF")
    thick = Side(border_style="medium", color="2E75B6")
    return {
        "bold": Font(bold=True, size=11),
        "bold_white": Font(bold=True, color="FFFFFF", size=11),
        "bold_big": Font(bold=True, size=15, color="FFFFFF"),
        "bold_huge": Font(bold=True, size=18, color="FFFFFF"),
        "italic": Font(italic=True, size=10, color="595959"),
        "header": PatternFill("solid", fgColor="1F4E78"),
        "title": PatternFill("solid", fgColor="2E75B6"),
        "subtotal": PatternFill("solid", fgColor="BDD7EE"),
        "total": PatternFill("solid", fgColor="FFEB99"),
        "rev": PatternFill("solid", fgColor="E2EFDA"),
        "cost": PatternFill("solid", fgColor="FCE4D6"),
        "profit": PatternFill("solid", fgColor="DDEBF7"),
        "kpi_good": PatternFill("solid", fgColor="C6EFCE"),
        "kpi_bad": PatternFill("solid", fgColor="FFC7CE"),
        "kpi_mid": PatternFill("solid", fgColor="FFEB9C"),
        "box": Border(left=border, right=border, top=border, bottom=border),
        "thick_top": Border(top=thick, left=border, right=border, bottom=border),
    }


def _detail_format(col: str) -> Optional[str]:
    if col == "Year":
        return INT_FMT
    if col == "DSCR":
        return DSCR_FMT
    if "ROI" in col or "(%)" in col:
        return PCT_FMT
    if "(€)" in col:
        return EUR_FMT
    return None


def write_scenario_sheet(ws, scenario: dict, df: pd.DataFrame, kpis: dict, styles):
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    n_cols = len(df.columns)
    last_col_letter = get_column_letter(n_cols)

    # Title row
    ws.merge_cells(f"A1:{last_col_letter}1")
    cell = ws["A1"]
    cell.value = scenario["label"]
    cell.font = styles["bold_big"]
    cell.fill = styles["title"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # KPI strip — 4 rows × 4 cols, comprehensive snapshot
    kpi_strip = [
        ("Y1 revenue",           _fmt_eur_short(kpis["y1_revenue_eur"])),
        ("Y1 EBITDA",            _fmt_eur_short(kpis["y1_ebitda_eur"])),
        ("Y1 EBITDA margin",     _fmt_pct(kpis["y1_ebitda_margin_pct"])),
        ("Y1 Net income",        _fmt_eur_short(kpis["y1_net_income_eur"])),
        ("Y1 Equity FCF",        _fmt_eur_short(kpis["y1_equity_fcf_eur"])),
        ("Equity IRR",           _fmt_pct(kpis["equity_irr_pct"])),
        ("Project IRR",          _fmt_pct(kpis["project_irr_pct"])),
        ("Equity MIRR",          _fmt_pct(kpis["equity_mirr_pct"])),
        ("Equity payback",       _fmt_yr(kpis["equity_payback_years"])),
        ("Disc. payback @ 8%",   _fmt_yr(kpis["equity_discounted_payback_years"])),
        ("MOIC",                 f"{kpis['moic']:.2f}x" if kpis["moic"] is not None else "n/a"),
        ("Avg ROE",              _fmt_pct(kpis["avg_roe_pct"])),
        ("Equity NPV @ 8%",      _fmt_eur_short(kpis["equity_npv_8_eur"])),
        ("Equity NPV @ 10%",     _fmt_eur_short(kpis["equity_npv_10_eur"])),
        ("Project NPV @ 8%",     _fmt_eur_short(kpis["project_npv_8_eur"])),
        ("Project NPV @ 10%",    _fmt_eur_short(kpis["project_npv_10_eur"])),
        ("Min DSCR",             f"{kpis['min_dscr']:.2f}x" if kpis["min_dscr"] is not None else "n/a"),
        ("Avg DSCR",             f"{kpis['avg_dscr']:.2f}x" if kpis["avg_dscr"] is not None else "n/a"),
        ("LLCR",                 f"{kpis['llcr']:.2f}x" if kpis["llcr"] is not None else "n/a"),
        ("PLCR",                 f"{kpis['plcr']:.2f}x" if kpis["plcr"] is not None else "n/a"),
    ]
    # Render in 5 rows × 4 (label,value) pairs = 8 cols
    pairs_per_row = 4
    start_row = 2
    for idx, (label, val) in enumerate(kpi_strip):
        row_offset = idx // pairs_per_row
        col_offset = idx % pairs_per_row
        c_label = get_column_letter(1 + col_offset * 2)
        c_val = get_column_letter(2 + col_offset * 2)
        r = start_row + row_offset
        ws[f"{c_label}{r}"] = label
        ws[f"{c_label}{r}"].font = styles["bold"]
        ws[f"{c_label}{r}"].fill = styles["subtotal"]
        ws[f"{c_label}{r}"].alignment = Alignment(horizontal="right")
        ws[f"{c_label}{r}"].border = styles["box"]
        ws[f"{c_val}{r}"] = val
        ws[f"{c_val}{r}"].font = styles["bold"]
        ws[f"{c_val}{r}"].fill = styles["total"]
        ws[f"{c_val}{r}"].alignment = Alignment(horizontal="left")
        ws[f"{c_val}{r}"].border = styles["box"]

    # Cashflow table — starts after KPI strip
    table_start = start_row + (len(kpi_strip) + pairs_per_row - 1) // pairs_per_row + 2
    cols = list(df.columns)

    # Column groups for fill coloring
    rev_cols = {"aFRR capacity (€)", "aFRR activation (€)", "PZU arbitrage (€)", "Total revenue (€)"}
    cost_cols = {"OPEX (€)", "Insurance (€)", "Augmentation (€)",
                 "Depreciation (€)", "Interest expense (€)", "Tax (€)",
                 "Principal payment (€)", "Debt service total (€)"}
    subtotal_cols = {"EBITDA (€)", "EBIT (€)", "EBT (€)"}
    fcf_cols = {"Project FCF (€)", "Equity FCF (€)", "Cumulative equity FCF (€)"}
    ratio_cols = {"EBITDA margin (%)", "ROE (%)", "ROIC (%)", "Cash-on-cash yield (%)"}

    # Header row
    for col_idx, name in enumerate(cols, start=1):
        cell = ws.cell(row=table_start, column=col_idx, value=name)
        cell.font = styles["bold_white"]
        cell.fill = styles["header"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = styles["box"]
    ws.row_dimensions[table_start].height = 42

    # Body rows
    for i, row in df.iterrows():
        excel_row = table_start + 1 + i
        for col_idx, name in enumerate(cols, start=1):
            val = row[name]
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            nf = _detail_format(name)
            if nf:
                cell.number_format = nf
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles["box"]
            if name == "Total revenue (€)":
                cell.font = styles["bold"]
                cell.fill = styles["rev"]
            elif name in rev_cols:
                cell.fill = styles["rev"]
            elif name in cost_cols:
                cell.fill = styles["cost"]
            elif name in subtotal_cols:
                cell.font = styles["bold"]
                cell.fill = styles["subtotal"]
            elif name == "Net income (€)":
                cell.font = styles["bold"]
                cell.fill = styles["subtotal"]
            elif name in fcf_cols:
                cell.font = styles["bold"]
                cell.fill = styles["total"]
            elif name in ratio_cols:
                cell.fill = styles["profit"]
            elif name == "DSCR":
                if val is not None and val < DSCR_COVENANT:
                    cell.fill = styles["kpi_bad"]
                else:
                    cell.fill = styles["kpi_good"]

    # Freeze first 2 columns (Year + first revenue) for horizontal scrolling
    ws.freeze_panes = f"C{table_start + 1}"

    # Column widths — Year narrow, monetary 14, ratio 11
    widths = []
    for name in cols:
        if name == "Year":
            widths.append(6)
        elif "(%)" in name or name == "DSCR":
            widths.append(11)
        elif "margin" in name or "yield" in name or "ROE" in name or "ROIC" in name:
            widths.append(11)
        else:
            widths.append(14)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Description block below
    desc_row = table_start + len(df) + 2
    ws.merge_cells(start_row=desc_row, start_column=1, end_row=desc_row, end_column=n_cols)
    ws[f"A{desc_row}"] = "Scenario description"
    ws[f"A{desc_row}"].font = styles["bold_white"]
    ws[f"A{desc_row}"].fill = styles["header"]
    ws[f"A{desc_row}"].alignment = Alignment(horizontal="center")
    desc_row += 1
    ws.merge_cells(start_row=desc_row, start_column=1, end_row=desc_row, end_column=n_cols)
    ws[f"A{desc_row}"] = scenario["description"]
    ws[f"A{desc_row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[desc_row].height = 80


def _fmt_eur_short(v: float) -> str:
    if v is None:
        return "n/a"
    if abs(v) >= 1_000_000:
        return f"€{v/1_000_000:.2f}M"
    if abs(v) >= 1_000:
        return f"€{v/1_000:.0f}k"
    return f"€{v:.0f}"


def _fmt_pct(v: Optional[float]) -> str:
    return f"{v:.2f}%" if v is not None else "n/a"


def _fmt_yr(v: Optional[float]) -> str:
    return f"{v:.1f} yr" if v is not None else "never"


def write_executive_summary(ws, all_kpis: list[dict], realistic_kpis: list[dict], styles):
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Executive Summary — 10 MW / 20 MWh BESS, 15-year project finance"
    cell.font = styles["bold_huge"]
    cell.fill = styles["title"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    sub = "Romanian aFRR + PZU arbitrage • Real DAMAS-anchored bids • 4 scenarios across PICASSO + market-share dynamics"
    ws.merge_cells("A2:F2")
    ws["A2"] = sub
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = styles["italic"]

    # Quick project facts
    facts_row = 4
    ws[f"A{facts_row}"] = "Project facts"
    ws[f"A{facts_row}"].font = styles["bold_white"]
    ws[f"A{facts_row}"].fill = styles["header"]
    ws.merge_cells(start_row=facts_row, start_column=1, end_row=facts_row, end_column=6)
    ws[f"A{facts_row}"].alignment = Alignment(horizontal="center")
    facts = [
        ("EPC investment", _fmt_eur_short(EPC_INVESTMENT_EUR), "€175/kWh installed"),
        ("Equity / Loan split", f"{EQUITY_PCT:.0f}% / {LOAN_PCT:.0f}%",
         f"Equity = {_fmt_eur_short(EPC_INVESTMENT_EUR * EQUITY_PCT/100)}, "
         f"Loan = {_fmt_eur_short(EPC_INVESTMENT_EUR * LOAN_PCT/100)}"),
        ("Loan terms", f"{LOAN_RATE*100:.1f}% over {LOAN_TERM_YR} years",
         f"Annuity = {_fmt_eur_short(annuity_payment(EPC_INVESTMENT_EUR*LOAN_PCT/100, LOAN_RATE, LOAN_TERM_YR))}/yr"),
        ("OPEX + Insurance", f"{OPEX_PCT_OF_EPC + INSURANCE_PCT_OF_EPC:.1f}% of EPC",
         f"€{(EPC_INVESTMENT_EUR * (OPEX_PCT_OF_EPC + INSURANCE_PCT_OF_EPC)/100)/1000:.0f}k/yr Y1, +3%/yr inflation"),
        ("Augmentation", f"{AUGMENTATION_PCT_OF_EPC:.0f}% of EPC at Y5 + Y10",
         f"€{EPC_INVESTMENT_EUR * AUGMENTATION_PCT_OF_EPC/100/1000:.0f}k each event"),
        ("Project life / Discount rate", f"{PROJECT_LIFE_YR} years / {DISCOUNT_RATE*100:.0f}%",
         f"DSCR covenant {DSCR_COVENANT:.2f}x (lender standard)"),
        ("Battery", f"{POWER_MW:.0f} MW / {CAPACITY_MWH:.0f} MWh",
         "RTE 0.97, full 0-100% SOC, 1 cycle/day"),
        ("Degradation", f"{DEGRADATION_RATE*100:.1f}%/yr",
         "Capacity fade applied to all revenue legs"),
    ]
    for i, (lbl, val, note) in enumerate(facts, start=1):
        r = facts_row + i
        ws[f"A{r}"] = lbl
        ws[f"B{r}"] = val
        ws[f"C{r}"] = note
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws[f"A{r}"].font = styles["bold"]
        for col in "ABCDEF":
            ws[f"{col}{r}"].border = styles["box"]
            ws[f"{col}{r}"].alignment = Alignment(horizontal="left", vertical="center")

    # KPI table — one row per scenario
    kpi_row = facts_row + len(facts) + 2
    ws[f"A{kpi_row}"] = "Headline KPIs by scenario"
    ws[f"A{kpi_row}"].font = styles["bold_white"]
    ws[f"A{kpi_row}"].fill = styles["header"]
    ws.merge_cells(start_row=kpi_row, start_column=1, end_row=kpi_row, end_column=6)
    ws[f"A{kpi_row}"].alignment = Alignment(horizontal="center")

    hdr = kpi_row + 1
    headers = ["Scenario",
               "Y1 (model)", "IRR (model)",
               "Y1 (realistic)", "IRR (realistic)", "Payback (real)"]
    for col_letter, label in zip("ABCDEF", headers):
        c = ws[f"{col_letter}{hdr}"]
        c.value = label
        c.font = styles["bold_white"]
        c.fill = styles["title"]
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = styles["box"]
    ws.row_dimensions[hdr].height = 30

    for i, (scen, kpis, rkpis) in enumerate(zip(SCENARIOS, all_kpis, realistic_kpis), start=1):
        r = hdr + i
        from openpyxl.styles import PatternFill
        scen_fill = PatternFill("solid", fgColor=scen["color"])
        ws[f"A{r}"] = scen["label"]
        ws[f"B{r}"] = _fmt_eur_short(kpis["y1_revenue_eur"])
        ws[f"C{r}"] = _fmt_pct(kpis["equity_irr_pct"])
        ws[f"D{r}"] = _fmt_eur_short(rkpis["y1_revenue_eur"])
        ws[f"E{r}"] = _fmt_pct(rkpis["equity_irr_pct"])
        ws[f"F{r}"] = _fmt_yr(rkpis["equity_payback_years"])
        for col in "ABCDEF":
            ws[f"{col}{r}"].font = styles["bold"]
            ws[f"{col}{r}"].fill = scen_fill
            ws[f"{col}{r}"].border = styles["box"]
            ws[f"{col}{r}"].alignment = Alignment(horizontal="center")
        ws[f"A{r}"].alignment = Alignment(horizontal="left")
        # Color realistic-IRR cell red if still > 30% (still suspicious)
        ri = rkpis.get("equity_irr_pct") or 0
        if ri > 50:
            ws[f"E{r}"].fill = styles["kpi_bad"]
        elif ri > 25:
            ws[f"E{r}"].fill = styles["kpi_mid"]
        else:
            ws[f"E{r}"].fill = styles["kpi_good"]

    # Decision-quality section
    dec_row = hdr + len(SCENARIOS) + 2
    ws[f"A{dec_row}"] = "How to read these numbers (lender perspective)"
    ws[f"A{dec_row}"].font = styles["bold_white"]
    ws[f"A{dec_row}"].fill = styles["header"]
    ws.merge_cells(start_row=dec_row, start_column=1, end_row=dec_row, end_column=6)
    ws[f"A{dec_row}"].alignment = Alignment(horizontal="center")
    rules = [
        "• Pitch deck Y1 number → Scenario A (current state).",
        "• 5-year average projection → blend of A + B (assumes PICASSO mid-period).",
        "• Lender-grade Y3-Y8 bankability case → Scenario C (mature market).",
        "• Lender stress test Y8-Y15 → Scenario D (bear case).",
        "• If DSCR < 1.20 in any year → lender covenant breach. Either lower gearing (more equity), longer tenor, or equity injection.",
        "• Numbers are SIMULATED on real Romanian DAMAS + OPCOM data. Compliance gates (ANRE/OPCOM/PRE/BRP/BSP/FSE) NOT applied — assumed obtained. Without them, all revenue = €0.",
        "• Realistic operator drag (forecast error, market share guess, Tg/Td tariffs) typically reduces revenue another 30-50% vs the engineering optimum shown here.",
    ]
    for i, txt in enumerate(rules):
        r = dec_row + 1 + i
        ws[f"A{r}"] = txt
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 26

    # Column widths
    widths = [38, 16, 14, 14, 16, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_inputs_sheet(ws, styles):
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Inputs & Assumptions"
    cell.font = styles["bold_huge"]
    cell.fill = styles["title"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Battery + financing block
    sections = [
        ("Battery technical", [
            ("Power", f"{POWER_MW:.0f} MW", "Inverter rating"),
            ("Energy capacity", f"{CAPACITY_MWH:.0f} MWh", "Daily throughput cap"),
            ("RTE", "0.97", "User directive — 3% loss per round-trip"),
            ("SOC band", "0% – 100%", "User directive — full DOD"),
            ("Cycles/day", "1", "Single round-trip per day"),
            ("Degradation", f"{DEGRADATION_RATE*100:.1f}%/yr",
             "Capacity fade applied to all revenue legs"),
        ]),
        ("Financing", [
            ("EPC investment", _fmt_eur_short(EPC_INVESTMENT_EUR), "€175/kWh installed"),
            ("Equity %", f"{EQUITY_PCT:.0f}%", _fmt_eur_short(EPC_INVESTMENT_EUR * EQUITY_PCT/100)),
            ("Loan %", f"{LOAN_PCT:.0f}%", _fmt_eur_short(EPC_INVESTMENT_EUR * LOAN_PCT/100)),
            ("Loan rate", f"{LOAN_RATE*100:.1f}%", "Fixed-rate annuity"),
            ("Loan term", f"{LOAN_TERM_YR} years", "Annuity payment ≈ "
             + _fmt_eur_short(annuity_payment(EPC_INVESTMENT_EUR*LOAN_PCT/100, LOAN_RATE, LOAN_TERM_YR))),
            ("Discount rate (NPV)", f"{DISCOUNT_RATE*100:.0f}%", "Standard EU project finance"),
            ("DSCR covenant", f"{DSCR_COVENANT:.2f}x", "Lender standard"),
        ]),
        ("Operating costs", [
            ("OPEX", f"{OPEX_PCT_OF_EPC:.1f}% of EPC = "
             + _fmt_eur_short(EPC_INVESTMENT_EUR * OPEX_PCT_OF_EPC/100), "Year 1, +3%/yr inflation"),
            ("Insurance", f"{INSURANCE_PCT_OF_EPC:.1f}% of EPC = "
             + _fmt_eur_short(EPC_INVESTMENT_EUR * INSURANCE_PCT_OF_EPC/100), "Year 1, +3%/yr inflation"),
            ("Augmentation", f"{AUGMENTATION_PCT_OF_EPC:.0f}% of EPC at Y5 + Y10",
             f"≈ {_fmt_eur_short(EPC_INVESTMENT_EUR * AUGMENTATION_PCT_OF_EPC/100)} per event"),
        ]),
        ("Physical stacking constraints (applied to BOTH model + realistic)", [
            ("aFRR capacity stacking", f"× {STACK_AFRR_CAPACITY:.2f}",
             "Pure availability reservation — independent of other streams"),
            ("aFRR activation stacking", f"× {STACK_AFRR_ACTIVATION:.2f}",
             "Activation MWh competes with PZU throughput → ~60% of pure-aFRR"),
            ("PZU stacking", f"× {STACK_PZU:.2f}",
             "Capacity reservation costs PZU dispatch flexibility → ~50% of pure-PZU"),
            ("Net stacking efficiency", "≈ 0.62 of naive sum",
             "Matches empirical EU BESS data (Modo Energy, Magnus Energy, Field Energy: 65-75%)"),
        ]),
        ("Operational drag multipliers (realistic mode only, on top of stacking)", [
            ("aFRR capacity drag", f"× {DRAG_AFRR_CAPACITY:.2f}",
             "REMIT/Tg/Td tariff erosion (~5-10%)"),
            ("aFRR activation drag", f"× {DRAG_AFRR_ACTIVATION:.2f}",
             "10% market-share guess → realistic ~2.5% (the BIG haircut)"),
            ("PZU drag", f"× {DRAG_PZU:.2f}",
             "Forecast error vs perfect-foresight backtest"),
        ]),
        ("Revenue scenarios — Y1 base values (BEFORE stacking + drag)", [
            ("Scenario A — aFRR capacity Y1", _fmt_eur_short(SCENARIOS[0]["afrr_capacity_y1"]),
             "10% market share, real DAMAS clearing"),
            ("Scenario A — aFRR activation Y1", _fmt_eur_short(SCENARIOS[0]["afrr_activation_y1"]),
             "Real per-slot DAMAS settlement"),
            ("Scenario A — PZU Y1", _fmt_eur_short(SCENARIOS[0]["pzu_y1"]),
             "Real OPCOM dispatch"),
            ("Scenario B — base Y1 multiplier", "0.75",
             "PICASSO live → -25% on aFRR; PZU unchanged"),
            ("Scenario C — base Y1 multipliers", "share 0.50, prices 0.60",
             "Mature market: 5% share, capacity at €5/MW/h"),
            ("Scenario D — base Y1 multipliers", "share 0.20, prices 0.40",
             "Bear case: 2% share, prices at German current level"),
        ]),
    ]

    r = 3
    for section_label, items in sections:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws[f"A{r}"] = section_label
        ws[f"A{r}"].font = styles["bold_white"]
        ws[f"A{r}"].fill = styles["header"]
        ws[f"A{r}"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 22
        r += 1
        for label, value, note in items:
            ws[f"A{r}"] = label
            ws[f"B{r}"] = value
            ws[f"C{r}"] = note
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            ws[f"A{r}"].font = styles["bold"]
            for col in "ABCDEF":
                ws[f"{col}{r}"].border = styles["box"]
                ws[f"{col}{r}"].alignment = Alignment(horizontal="left", vertical="center")
            r += 1
        r += 1  # spacing

    widths = [34, 22, 22, 22, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_comparison_sheet(ws, dfs: dict[str, pd.DataFrame], all_kpis: list[dict], styles):
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Scenarios Comparison — annual cashflow + KPI summary"
    cell.font = styles["bold_huge"]
    cell.fill = styles["title"]
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 32

    # KPI comparison table (top)
    r = 3
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws[f"A{r}"] = "Scenario KPIs"
    ws[f"A{r}"].font = styles["bold_white"]
    ws[f"A{r}"].fill = styles["header"]
    ws[f"A{r}"].alignment = Alignment(horizontal="center")
    r += 1
    headers = ["KPI"] + [s["label"][:30] for s in SCENARIOS]
    for col_letter, label in zip("ABCDE"[: len(headers)], headers):
        c = ws[f"{col_letter}{r}"]
        c.value = label
        c.font = styles["bold_white"]
        c.fill = styles["title"]
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = styles["box"]
    ws.row_dimensions[r].height = 36

    def _xfmt(v): return f"{v:.2f}x" if v is not None else "n/a"
    kpi_rows = [
        # ---- Profitability snapshot ----
        ("Y1 revenue (€)", lambda k: k["y1_revenue_eur"], _fmt_eur_short),
        ("Y1 EBITDA (€)", lambda k: k["y1_ebitda_eur"], _fmt_eur_short),
        ("Y1 EBITDA margin", lambda k: k["y1_ebitda_margin_pct"], _fmt_pct),
        ("Y1 EBIT (€)", lambda k: k["y1_ebit_eur"], _fmt_eur_short),
        ("Y1 Net Income (€)", lambda k: k["y1_net_income_eur"], _fmt_eur_short),
        ("Y1 Equity FCF (€)", lambda k: k["y1_equity_fcf_eur"], _fmt_eur_short),
        # ---- Lifetime ----
        ("Lifetime revenue (€)", lambda k: k["lifetime_revenue_eur"], _fmt_eur_short),
        ("Lifetime EBITDA (€)", lambda k: k["lifetime_ebitda_eur"], _fmt_eur_short),
        ("Lifetime Net income (€)", lambda k: k["lifetime_net_income_eur"], _fmt_eur_short),
        ("Lifetime Equity FCF (€)", lambda k: k["lifetime_equity_fcf_eur"], _fmt_eur_short),
        ("Lifetime Project FCF (€)", lambda k: k["lifetime_project_fcf_eur"], _fmt_eur_short),
        ("Avg EBITDA margin", lambda k: k["avg_ebitda_margin_pct"], _fmt_pct),
        # ---- Returns ----
        ("Equity IRR", lambda k: k["equity_irr_pct"], _fmt_pct),
        ("Project IRR", lambda k: k["project_irr_pct"], _fmt_pct),
        ("Equity MIRR", lambda k: k["equity_mirr_pct"], _fmt_pct),
        ("Project MIRR", lambda k: k["project_mirr_pct"], _fmt_pct),
        ("Avg ROE", lambda k: k["avg_roe_pct"], _fmt_pct),
        ("Avg ROIC", lambda k: k["avg_roic_pct"], _fmt_pct),
        ("MOIC (money multiple)", lambda k: k["moic"], _xfmt),
        # ---- NPV ----
        ("Equity NPV @ 8%", lambda k: k["equity_npv_8_eur"], _fmt_eur_short),
        ("Equity NPV @ 10%", lambda k: k["equity_npv_10_eur"], _fmt_eur_short),
        ("Equity NPV @ 12%", lambda k: k["equity_npv_12_eur"], _fmt_eur_short),
        ("Project NPV @ 8%", lambda k: k["project_npv_8_eur"], _fmt_eur_short),
        ("Project NPV @ 10%", lambda k: k["project_npv_10_eur"], _fmt_eur_short),
        # ---- Payback ----
        ("Equity payback (simple)", lambda k: k["equity_payback_years"], _fmt_yr),
        ("Equity payback (discounted @ 8%)", lambda k: k["equity_discounted_payback_years"], _fmt_yr),
        ("Project payback (simple)", lambda k: k["project_payback_years"], _fmt_yr),
        ("Project payback (discounted @ 8%)", lambda k: k["project_discounted_payback_years"], _fmt_yr),
        # ---- Lender / coverage ----
        ("Min DSCR", lambda k: k["min_dscr"], _xfmt),
        ("Avg DSCR", lambda k: k["avg_dscr"], _xfmt),
        ("LLCR (Loan Life Coverage)", lambda k: k["llcr"], _xfmt),
        ("PLCR (Project Life Coverage)", lambda k: k["plcr"], _xfmt),
        ("DSCR breach years", lambda k: k["dscr_breach_years"],
         lambda v: ", ".join(str(y) for y in v) if v else "—"),
    ]
    for kpi_label, kpi_get, fmt in kpi_rows:
        r += 1
        ws[f"A{r}"] = kpi_label
        ws[f"A{r}"].font = styles["bold"]
        ws[f"A{r}"].fill = styles["subtotal"]
        ws[f"A{r}"].border = styles["box"]
        ws[f"A{r}"].alignment = Alignment(horizontal="left")
        for i, (scen, kpis) in enumerate(zip(SCENARIOS, all_kpis), start=2):
            col_letter = get_column_letter(i)
            try:
                val = kpi_get(kpis)
                ws[f"{col_letter}{r}"] = fmt(val)
            except Exception:
                ws[f"{col_letter}{r}"] = "n/a"
            ws[f"{col_letter}{r}"].fill = PatternFill("solid", fgColor=scen["color"])
            ws[f"{col_letter}{r}"].border = styles["box"]
            ws[f"{col_letter}{r}"].alignment = Alignment(horizontal="center")

    # Annual cashflow comparison — total revenue per scenario per year
    r += 3
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws[f"A{r}"] = "Annual cashflow after debt by scenario (€)"
    ws[f"A{r}"].font = styles["bold_white"]
    ws[f"A{r}"].fill = styles["header"]
    ws[f"A{r}"].alignment = Alignment(horizontal="center")
    r += 1
    headers = ["Year"] + [s["label"][:30] for s in SCENARIOS]
    for col_letter, label in zip("ABCDE", headers):
        c = ws[f"{col_letter}{r}"]
        c.value = label
        c.font = styles["bold_white"]
        c.fill = styles["title"]
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = styles["box"]
    ws.row_dimensions[r].height = 36

    year_start_row = r + 1
    for y_idx in range(PROJECT_LIFE_YR):
        rr = year_start_row + y_idx
        ws[f"A{rr}"] = y_idx + 1
        ws[f"A{rr}"].font = styles["bold"]
        ws[f"A{rr}"].alignment = Alignment(horizontal="center")
        ws[f"A{rr}"].border = styles["box"]
        for i, scen in enumerate(SCENARIOS, start=2):
            df = dfs[scen["key"]]
            cf = float(df.iloc[y_idx]["Equity FCF (€)"])
            col_letter = get_column_letter(i)
            ws[f"{col_letter}{rr}"] = cf
            ws[f"{col_letter}{rr}"].number_format = EUR_FMT
            ws[f"{col_letter}{rr}"].fill = PatternFill("solid", fgColor=scen["color"])
            ws[f"{col_letter}{rr}"].border = styles["box"]
            ws[f"{col_letter}{rr}"].alignment = Alignment(horizontal="right")

    # TOTAL row
    rr = year_start_row + PROJECT_LIFE_YR
    ws[f"A{rr}"] = "TOTAL"
    ws[f"A{rr}"].font = styles["bold"]
    ws[f"A{rr}"].fill = styles["total"]
    ws[f"A{rr}"].alignment = Alignment(horizontal="center")
    ws[f"A{rr}"].border = styles["box"]
    for i, scen in enumerate(SCENARIOS, start=2):
        df = dfs[scen["key"]]
        total = float(df["Equity FCF (€)"].sum())
        col_letter = get_column_letter(i)
        ws[f"{col_letter}{rr}"] = total
        ws[f"{col_letter}{rr}"].number_format = EUR_FMT
        ws[f"{col_letter}{rr}"].fill = styles["total"]
        ws[f"{col_letter}{rr}"].font = styles["bold"]
        ws[f"{col_letter}{rr}"].border = styles["box"]
        ws[f"{col_letter}{rr}"].alignment = Alignment(horizontal="right")

    widths = [34, 22, 22, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ===================================================================
#  Main
# ===================================================================
def write_consolidated_cashflow_sheet(ws, dfs: dict, kpis_list: list, styles, title_suffix: str):
    """Write all 4 scenarios stacked vertically on one sheet (banking
    layout). Each scenario block: colored header bar, KPI strip,
    15-year cashflow with full PF columns, separator. Compresses what
    used to be 4 separate tabs into 1 dense scrollable sheet.
    """
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    n_cols = 26  # Year + 25 PF columns
    last_col_letter = get_column_letter(n_cols)

    # Title
    ws.merge_cells(f"A1:{last_col_letter}1")
    cell = ws["A1"]
    cell.value = f"BESS 15-year Cashflow — {title_suffix}"
    cell.font = styles["bold_huge"]
    cell.fill = styles["title"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    rev_cols = {"aFRR capacity (€)", "aFRR activation (€)", "PZU arbitrage (€)", "Total revenue (€)"}
    cost_cols = {"OPEX (€)", "Insurance (€)", "Augmentation (€)",
                 "Depreciation (€)", "Interest expense (€)", "Tax (€)",
                 "Principal payment (€)", "Debt service total (€)"}
    subtotal_cols = {"EBITDA (€)", "EBIT (€)", "EBT (€)"}
    fcf_cols = {"Project FCF (€)", "Equity FCF (€)", "Cumulative equity FCF (€)"}
    ratio_cols = {"EBITDA margin (%)", "ROE (%)", "ROIC (%)", "Cash-on-cash yield (%)"}

    cur_row = 3
    for scen, kpis in zip(SCENARIOS, kpis_list):
        df = dfs[scen["key"]]
        # Scenario header band
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
        c = ws.cell(row=cur_row, column=1, value=scen["label"])
        c.font = styles["bold_white"]
        c.fill = PatternFill("solid", fgColor=scen["color"].replace("DDEBF7", "1F4E78")
                             if scen["color"] == "DDEBF7" else "1F4E78")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # Compact KPI strip — one row, the most important numbers
        kpi_strip = [
            ("Y1 rev", _fmt_eur_short(kpis["y1_revenue_eur"])),
            ("Y1 EBITDA", _fmt_eur_short(kpis["y1_ebitda_eur"])),
            ("EBITDA margin", _fmt_pct(kpis["y1_ebitda_margin_pct"])),
            ("Equity IRR", _fmt_pct(kpis["equity_irr_pct"])),
            ("MIRR", _fmt_pct(kpis["equity_mirr_pct"])),
            ("MOIC", f"{kpis['moic']:.2f}x" if kpis["moic"] is not None else "n/a"),
            ("Payback", _fmt_yr(kpis["equity_payback_years"])),
            ("NPV @8%", _fmt_eur_short(kpis["equity_npv_8_eur"])),
            ("Avg ROE", _fmt_pct(kpis["avg_roe_pct"])),
            ("Min DSCR", f"{kpis['min_dscr']:.2f}x" if kpis["min_dscr"] is not None else "n/a"),
            ("LLCR", f"{kpis['llcr']:.2f}x" if kpis["llcr"] is not None else "n/a"),
            ("PLCR", f"{kpis['plcr']:.2f}x" if kpis["plcr"] is not None else "n/a"),
        ]
        # Layout: 12 KPIs across 24 cols (label+value each) — actually
        # use compact label/value pairs in one row across the full width.
        for idx, (label, val) in enumerate(kpi_strip):
            col_label = 1 + idx * 2
            col_val = 2 + idx * 2
            cl = ws.cell(row=cur_row, column=col_label, value=label)
            cl.font = styles["bold"]
            cl.fill = styles["subtotal"]
            cl.alignment = Alignment(horizontal="right")
            cl.border = styles["box"]
            cv = ws.cell(row=cur_row, column=col_val, value=val)
            cv.font = styles["bold"]
            cv.fill = styles["total"]
            cv.alignment = Alignment(horizontal="left")
            cv.border = styles["box"]
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        # Cashflow table header
        cols = list(df.columns)
        for col_idx, name in enumerate(cols, start=1):
            cell = ws.cell(row=cur_row, column=col_idx, value=name)
            cell.font = styles["bold_white"]
            cell.fill = styles["header"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = styles["box"]
        ws.row_dimensions[cur_row].height = 42
        cur_row += 1

        # 15 year rows
        for i in range(len(df)):
            for col_idx, name in enumerate(cols, start=1):
                val = df.iloc[i][name]
                cell = ws.cell(row=cur_row, column=col_idx, value=val)
                nf = _detail_format(name)
                if nf:
                    cell.number_format = nf
                cell.alignment = Alignment(horizontal="center")
                cell.border = styles["box"]
                if name == "Total revenue (€)":
                    cell.font = styles["bold"]
                    cell.fill = styles["rev"]
                elif name in rev_cols:
                    cell.fill = styles["rev"]
                elif name in cost_cols:
                    cell.fill = styles["cost"]
                elif name in subtotal_cols:
                    cell.font = styles["bold"]
                    cell.fill = styles["subtotal"]
                elif name == "Net income (€)":
                    cell.font = styles["bold"]
                    cell.fill = styles["subtotal"]
                elif name in fcf_cols:
                    cell.font = styles["bold"]
                    cell.fill = styles["total"]
                elif name in ratio_cols:
                    cell.fill = styles["profit"]
                elif name == "DSCR":
                    if val is not None and val < DSCR_COVENANT:
                        cell.fill = styles["kpi_bad"]
                    else:
                        cell.fill = styles["kpi_good"]
            cur_row += 1
        cur_row += 1  # blank separator

    # Column widths
    widths = []
    for name in (list(dfs[SCENARIOS[0]["key"]].columns)):
        if name == "Year":
            widths.append(6)
        elif "(%)" in name or name == "DSCR":
            widths.append(11)
        else:
            widths.append(14)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C3"


def write_returns_matrix_sheet(ws, all_kpis: list, realistic_kpis: list, styles):
    """Comprehensive KPI matrix — every metric × 4 scenarios × {Model, Realistic}.
    The single most important sheet for a credit committee."""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = "Returns & Coverage Matrix — every metric, every scenario, modeled vs realistic"
    cell.font = styles["bold_huge"]
    cell.fill = styles["title"]
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 32

    hdr_row = 3
    headers = ["KPI"]
    for scen in SCENARIOS:
        headers.append(f"{scen['key']} — Model")
        headers.append(f"{scen['key']} — Real")
    for col_idx, label in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=col_idx, value=label)
        c.font = styles["bold_white"]
        c.fill = styles["header"]
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = styles["box"]
    ws.row_dimensions[hdr_row].height = 36

    def _xfmt(v): return f"{v:.2f}x" if v is not None else "n/a"

    sections = [
        ("PROFITABILITY (Y1)", [
            ("Y1 revenue", "y1_revenue_eur", _fmt_eur_short),
            ("Y1 EBITDA", "y1_ebitda_eur", _fmt_eur_short),
            ("Y1 EBITDA margin", "y1_ebitda_margin_pct", _fmt_pct),
            ("Y1 EBIT", "y1_ebit_eur", _fmt_eur_short),
            ("Y1 Net Income", "y1_net_income_eur", _fmt_eur_short),
            ("Y1 Equity FCF", "y1_equity_fcf_eur", _fmt_eur_short),
            ("Y1 DSCR", "y1_dscr", _xfmt),
        ]),
        ("LIFETIME (15-yr cumulative)", [
            ("Lifetime revenue", "lifetime_revenue_eur", _fmt_eur_short),
            ("Lifetime EBITDA", "lifetime_ebitda_eur", _fmt_eur_short),
            ("Lifetime Net Income", "lifetime_net_income_eur", _fmt_eur_short),
            ("Lifetime Equity FCF", "lifetime_equity_fcf_eur", _fmt_eur_short),
            ("Lifetime Project FCF", "lifetime_project_fcf_eur", _fmt_eur_short),
            ("Avg EBITDA margin", "avg_ebitda_margin_pct", _fmt_pct),
        ]),
        ("RETURNS (operator)", [
            ("Equity IRR", "equity_irr_pct", _fmt_pct),
            ("Project IRR", "project_irr_pct", _fmt_pct),
            ("Equity MIRR", "equity_mirr_pct", _fmt_pct),
            ("Project MIRR", "project_mirr_pct", _fmt_pct),
            ("Avg ROE", "avg_roe_pct", _fmt_pct),
            ("Avg ROIC", "avg_roic_pct", _fmt_pct),
            ("MOIC (money multiple)", "moic", _xfmt),
        ]),
        ("VALUATION (NPV)", [
            ("Equity NPV @ 8%", "equity_npv_8_eur", _fmt_eur_short),
            ("Equity NPV @ 10%", "equity_npv_10_eur", _fmt_eur_short),
            ("Equity NPV @ 12%", "equity_npv_12_eur", _fmt_eur_short),
            ("Project NPV @ 8%", "project_npv_8_eur", _fmt_eur_short),
            ("Project NPV @ 10%", "project_npv_10_eur", _fmt_eur_short),
        ]),
        ("PAYBACK", [
            ("Equity payback (simple)", "equity_payback_years", _fmt_yr),
            ("Equity payback (discounted @ 8%)", "equity_discounted_payback_years", _fmt_yr),
            ("Project payback (simple)", "project_payback_years", _fmt_yr),
            ("Project payback (discounted @ 8%)", "project_discounted_payback_years", _fmt_yr),
        ]),
        ("LENDER COVERAGE", [
            ("Min DSCR", "min_dscr", _xfmt),
            ("Avg DSCR", "avg_dscr", _xfmt),
            ("LLCR (Loan Life Coverage)", "llcr", _xfmt),
            ("PLCR (Project Life Coverage)", "plcr", _xfmt),
            ("DSCR breach years", "dscr_breach_years",
             lambda v: ", ".join(str(y) for y in v) if v else "—"),
        ]),
    ]

    cur_row = hdr_row + 1
    for section_label, kpi_rows in sections:
        # Section band
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=9)
        c = ws.cell(row=cur_row, column=1, value=section_label)
        c.font = styles["bold_white"]
        c.fill = styles["title"]
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1
        for label, key, fmt in kpi_rows:
            ws.cell(row=cur_row, column=1, value=label)
            ws.cell(row=cur_row, column=1).font = styles["bold"]
            ws.cell(row=cur_row, column=1).fill = styles["subtotal"]
            ws.cell(row=cur_row, column=1).border = styles["box"]
            ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left")
            for s_idx, (scen, kpis, rkpis) in enumerate(zip(SCENARIOS, all_kpis, realistic_kpis)):
                col_m = 2 + s_idx * 2
                col_r = 3 + s_idx * 2
                try:
                    ws.cell(row=cur_row, column=col_m, value=fmt(kpis[key]))
                except Exception:
                    ws.cell(row=cur_row, column=col_m, value="n/a")
                try:
                    ws.cell(row=cur_row, column=col_r, value=fmt(rkpis[key]))
                except Exception:
                    ws.cell(row=cur_row, column=col_r, value="n/a")
                scen_fill = PatternFill("solid", fgColor=scen["color"])
                ws.cell(row=cur_row, column=col_m).fill = scen_fill
                ws.cell(row=cur_row, column=col_r).fill = scen_fill
                ws.cell(row=cur_row, column=col_m).border = styles["box"]
                ws.cell(row=cur_row, column=col_r).border = styles["box"]
                ws.cell(row=cur_row, column=col_m).alignment = Alignment(horizontal="center")
                ws.cell(row=cur_row, column=col_r).alignment = Alignment(horizontal="center")
            cur_row += 1
        cur_row += 1  # blank row between sections

    # Column widths
    ws.column_dimensions["A"].width = 36
    for c in "BCDEFGHI":
        ws.column_dimensions[c].width = 14
    ws.freeze_panes = "B4"


def write_sensitivity_sheet(ws, base_scenario: dict, styles):
    """Single-variable sensitivity on Scenario A: vary each drag lever
    independently and show resulting Equity IRR + Y1 revenue + Min DSCR.
    Lenders use this to identify the IRR's biggest risk drivers."""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Sensitivity Analysis — single-variable swings on Scenario A (Current, 10% share)"
    cell.font = styles["bold_huge"]
    cell.fill = styles["title"]
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 32

    sub = (
        "Each row holds all other defaults constant and varies ONE driver. "
        "Use to identify the IRR's biggest exposure. Bold row = base case."
    )
    ws.merge_cells("A2:F2")
    ws["A2"] = sub
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = styles["italic"]

    # Sensitivity table header
    hdr = 4
    headers = ["Variable", "Value tested", "Y1 revenue", "Equity IRR", "Min DSCR", "vs base case"]
    for col_idx, label in enumerate(headers, start=1):
        c = ws.cell(row=hdr, column=col_idx, value=label)
        c.font = styles["bold_white"]
        c.fill = styles["header"]
        c.alignment = Alignment(horizontal="center")
        c.border = styles["box"]
    ws.row_dimensions[hdr].height = 30

    # Compute base case (Scenario A realistic)
    global DRAG_AFRR_CAPACITY, DRAG_AFRR_ACTIVATION, DRAG_PZU
    global STACK_AFRR_CAPACITY, STACK_AFRR_ACTIVATION, STACK_PZU
    base_cap, base_act, base_pzu = DRAG_AFRR_CAPACITY, DRAG_AFRR_ACTIVATION, DRAG_PZU
    base_s_cap, base_s_act, base_s_pzu = STACK_AFRR_CAPACITY, STACK_AFRR_ACTIVATION, STACK_PZU
    base_df = build_cashflow(base_scenario, apply_drag=True)
    base_kpis = compute_kpis(base_scenario, base_df)
    base_irr = base_kpis["equity_irr_pct"]

    sensitivities = [
        ("aFRR activation drag (share)", "DRAG_AFRR_ACTIVATION",
         [(0.10, "1% market share — strict bear"),
          (0.15, "1.5% — Romanian mature"),
          (0.25, "2.5% — DEFAULT realistic"),
          (0.40, "4% — high optimism"),
          (1.00, "10% — engineering optimum (no drag)")]),
        ("PZU drag (forecast error)", "DRAG_PZU",
         [(0.55, "−45% — full forecast + Tg/Td"),
          (0.65, "−35% — heavy drag"),
          (0.80, "−20% — DEFAULT realistic"),
          (0.90, "−10% — light drag"),
          (1.00, "perfect foresight")]),
        ("aFRR capacity drag (REMIT/Tg/Td)", "DRAG_AFRR_CAPACITY",
         [(0.85, "−15% — heavy tariffs"),
          (0.93, "−7% — DEFAULT realistic"),
          (1.00, "no tariff drag")]),
        ("aFRR-PZU stacking efficiency (PZU)", "STACK_PZU",
         [(0.30, "30% — heavy capacity reservation cost"),
          (0.50, "50% — DEFAULT (industry empirical)"),
          (0.75, "75% — efficient joint dispatch"),
          (1.00, "no stacking constraint (impossible)")]),
        ("aFRR-PZU stacking efficiency (activation)", "STACK_AFRR_ACTIVATION",
         [(0.40, "40% — only some MWh available for activation"),
          (0.60, "60% — DEFAULT"),
          (0.80, "80% — efficient joint dispatch"),
          (1.00, "no stacking constraint (impossible)")]),
    ]

    cur_row = hdr + 1
    import sys as _sys
    mod = _sys.modules[__name__]
    for var_label, var_name, values in sensitivities:
        for value, note in values:
            # Patch the global
            setattr(mod, var_name, value)
            df_s = build_cashflow(base_scenario, apply_drag=True)
            k_s = compute_kpis(base_scenario, df_s)
            irr_s = k_s["equity_irr_pct"]
            y1 = k_s["y1_revenue_eur"]
            mind = k_s["min_dscr"]
            delta = (irr_s - base_irr) if (irr_s is not None and base_irr is not None) else None

            ws.cell(row=cur_row, column=1, value=var_label).font = styles["bold"]
            ws.cell(row=cur_row, column=2, value=f"{value:.2f}  ({note})")
            ws.cell(row=cur_row, column=3, value=_fmt_eur_short(y1))
            ws.cell(row=cur_row, column=4, value=_fmt_pct(irr_s))
            ws.cell(row=cur_row, column=5, value=f"{mind:.2f}x" if mind is not None else "n/a")
            if delta is not None:
                ws.cell(row=cur_row, column=6, value=f"{delta:+.1f}pp")
            else:
                ws.cell(row=cur_row, column=6, value="—")
            for col in range(1, 7):
                ws.cell(row=cur_row, column=col).border = styles["box"]
                ws.cell(row=cur_row, column=col).alignment = Alignment(horizontal="center")
            ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left")
            # Bold the base-case row
            is_base = (
                (var_name == "DRAG_AFRR_ACTIVATION" and value == 0.25) or
                (var_name == "DRAG_PZU" and value == 0.80) or
                (var_name == "DRAG_AFRR_CAPACITY" and value == 0.93) or
                (var_name == "STACK_PZU" and value == 0.50) or
                (var_name == "STACK_AFRR_ACTIVATION" and value == 0.60)
            )
            if is_base:
                for col in range(1, 7):
                    ws.cell(row=cur_row, column=col).font = styles["bold"]
                    ws.cell(row=cur_row, column=col).fill = styles["total"]
            cur_row += 1
        cur_row += 1  # spacer between variables

    # Restore defaults
    setattr(mod, "DRAG_AFRR_CAPACITY", base_cap)
    setattr(mod, "DRAG_AFRR_ACTIVATION", base_act)
    setattr(mod, "DRAG_PZU", base_pzu)
    setattr(mod, "STACK_AFRR_CAPACITY", base_s_cap)
    setattr(mod, "STACK_AFRR_ACTIVATION", base_s_act)
    setattr(mod, "STACK_PZU", base_s_pzu)

    # Column widths
    widths = [40, 38, 18, 14, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Footer note
    cur_row += 1
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
    ws.cell(row=cur_row, column=1, value=(
        "How to read: each variable changes one assumption while the "
        "rest stay at default. Δ vs base = percentage points difference "
        "vs the realistic Scenario A IRR. Variables with the largest "
        "swings are the IRR's biggest risk drivers — typically activation "
        "share assumption is the dominant one."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[cur_row].height = 50


def main() -> None:
    print("Building 15-year cashflow scenarios…")
    dfs = {}
    all_kpis = []
    realistic_dfs = {}
    realistic_kpis = []
    for scen in SCENARIOS:
        df = build_cashflow(scen, apply_drag=False)
        dfs[scen["key"]] = df
        kpis = compute_kpis(scen, df)
        all_kpis.append(kpis)
        rdf = build_cashflow(scen, apply_drag=True)
        realistic_dfs[scen["key"]] = rdf
        rkpis = compute_kpis(scen, rdf)
        realistic_kpis.append(rkpis)
        print(f"  {scen['label']:42s}  Modeled IRR {kpis['equity_irr_pct'] or 0:>6.1f}%  "
              f"Realistic IRR {rkpis['equity_irr_pct'] or 0:>6.1f}%")

    out_dir = ROOT / "reports"
    out_dir.mkdir(exist_ok=True)
    today = "2026-05-04"
    out_path = out_dir / f"BESS_Financial_Model_BANK_6tabs_{today}.xlsx"

    # Banking layout: 6 sheets
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        pd.DataFrame().to_excel(xl, sheet_name="1. Cover & KPIs", index=False)
        pd.DataFrame().to_excel(xl, sheet_name="2. Assumptions", index=False)
        pd.DataFrame().to_excel(xl, sheet_name="3. Cashflow (Modeled)", index=False)
        pd.DataFrame().to_excel(xl, sheet_name="4. Cashflow (Realistic)", index=False)
        pd.DataFrame().to_excel(xl, sheet_name="5. Returns Matrix", index=False)
        pd.DataFrame().to_excel(xl, sheet_name="6. Sensitivity", index=False)

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    styles = _styles()

    write_executive_summary(wb["1. Cover & KPIs"], all_kpis, realistic_kpis, styles)
    write_inputs_sheet(wb["2. Assumptions"], styles)
    write_consolidated_cashflow_sheet(wb["3. Cashflow (Modeled)"], dfs, all_kpis, styles,
                                       "Engineering Optimum (no operator drag)")
    write_consolidated_cashflow_sheet(wb["4. Cashflow (Realistic)"], realistic_dfs, realistic_kpis,
                                       styles, "Realistic — operator drag + tax + depreciation applied")
    write_returns_matrix_sheet(wb["5. Returns Matrix"], all_kpis, realistic_kpis, styles)
    write_sensitivity_sheet(wb["6. Sensitivity"], SCENARIOS[0], styles)

    wb.active = 0
    wb.save(out_path)

    print(f"\nWrote: {out_path}")
    print("\nBanking-grade layout (6 sheets):")
    print("  1. Cover & KPIs           — single-page exec summary + 4-scenario KPI table")
    print("  2. Assumptions            — battery + financing + tax + stacking + drag")
    print("  3. Cashflow (Modeled)     — all 4 scenarios stacked, engineering-optimum view")
    print("  4. Cashflow (Realistic)   — all 4 scenarios stacked, lender-grade view")
    print("  5. Returns Matrix         — every KPI × every scenario × {Model, Real}")
    print("  6. Sensitivity            — single-variable IRR swings (Scenario A)")


if __name__ == "__main__":
    main()
