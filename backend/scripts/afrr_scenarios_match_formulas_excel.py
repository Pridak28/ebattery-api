"""Formula-driven version of aFRR_scenarios_PICASSO_market_share_*.xlsx —
exactly matches the original layout (5 sheets, 14 columns per scenario,
4 scenarios A/B/C/D).

Preserves:
  - Summary sheet structure (battery config + scenario table + descriptions + notes)
  - Per-scenario sheet: 14 columns (Date, Weekday, Selected, Recommended bid,
    Acceptance rate, Capacity revenue, Slots activated, Avg/Max activation
    price, Market activations, Captured energy, Battery cap binding,
    Activation revenue, Daily revenue)

Formula columns (live, recomputable):
  - Recommended bid       — IF(Selected=aFRR+, damas_up × cap_mult_X, damas_down × cap_mult_X)
  - Acceptance rate       — Inputs!acceptance_rate
  - Capacity revenue      — power × hours × bid × acceptance × stack_capacity
  - Captured energy       — MIN(market × share × accept × stack, power × hours, capacity)
  - Battery cap binding   — IF(captured >= capacity, "yes", "no")
  - Activation revenue    — captured × avg_price × activation_mult_X
  - Daily revenue         — capacity + activation
  - Selected              — IF(up_total > down_total, "aFRR+", "aFRR-")

Static columns (from DAMAS_Daily hidden helper sheet):
  - Date, Weekday, Market activations, Avg/Max activation price,
    Slots activated — all measured outputs of the per-15-min-slot
    real-market simulation

Run from `backend/`:
    PYTHONPATH=. arch -arm64 /usr/local/bin/python3 scripts/afrr_scenarios_match_formulas_excel.py
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.config import settings  # noqa: E402
from app.market_data.capacity_prices import get_canonical_capacity_price  # noqa: E402
from app.services.fr_service import FRService  # noqa: E402

# ===================================================================
#  Defaults
# ===================================================================
POWER_MW_DEFAULT = float(settings.DEFAULT_POWER_MW)
CAPACITY_MWH_DEFAULT = float(settings.DEFAULT_CAPACITY_MWH)
ACCEPTANCE_DEFAULT = 0.80
HOURS_PER_DAY = 24
SLOT_HOURS = 0.25
PRICE_VALID_MAX = 500.0
MARKET_SHARE_DEFAULT = 0.10  # before scenario adjustment

# Use ORIGINAL DAMAS values (€8.72 / €6.59) to match the source file exactly.
# User can edit on Inputs sheet.
DAMAS_AFRR_UP_DEFAULT = 8.72
DAMAS_AFRR_DOWN_DEFAULT = 6.59

# Stacking
STACK_CAPACITY_DEFAULT = 1.00
STACK_ACTIVATION_DEFAULT = 0.60

SCENARIOS = [
    {"key": "A", "label": "A. Current (10%, pre-PICASSO)",
     "short": "Scenario_A_Current",
     "color": "DDEBF7",
     "market_share": 0.10, "capacity_mult": 1.00, "activation_mult": 1.00,
     "description": "Today's Romanian aFRR market. PICASSO platform NOT yet connected (Romania waiting on Hungary/MAVIR). Battery operator captures 10% of TSO activations (small new entrant in a sparse BESS market — only ~50 MW deployed Romania-wide). Real DAMAS capacity clearing prices apply. Activation revenue uses real per-slot DAMAS settlement prices."},
    {"key": "B", "label": "B. PICASSO go-live (−25%)",
     "short": "Scenario_B_PICASSO",
     "color": "FCE4D6",
     "market_share": 0.10, "capacity_mult": 0.75, "activation_mult": 0.75,
     "description": "PICASSO platform activates Romania (expected 2026-2027 once Hungary/MAVIR connects). Cross-border aFRR-energy bid stack caps Romanian extreme activation prices toward European average; capacity prices compress in parallel. Conservative step-down applied: −25% on BOTH capacity bid AND per-slot activation revenue."},
    {"key": "C", "label": "C. Mature market (5% share, −40%)",
     "short": "Scenario_C_Mature",
     "color": "FFF2CC",
     "market_share": 0.05, "capacity_mult": 0.60, "activation_mult": 0.65,
     "description": "Romanian BESS pipeline (1+ GW announced) materializes through 2027-2028. Market share compresses from 10% to 5% as new BESS deployments compete for the same activation volumes. aFRR capacity prices compress 40% to ~€5/MW/h (matching Germany's current level). Steady-state realistic case for years Y3-Y8."},
    {"key": "D", "label": "D. Bear case (2%, full saturation)",
     "short": "Scenario_D_Bear",
     "color": "F8CBAD",
     "market_share": 0.02, "capacity_mult": 0.40, "activation_mult": 0.50,
     "description": "Long-term steady state (2030+). Romanian BESS deploys 2-3 GW matching Germany's current density. Market share collapses to 2% per operator. aFRR capacity prices fall 60%. Activation prices fall 50% as PICASSO + thick bid stack normalize Romanian prices to European average. Bankability stress case used by lenders for late-life (Y8-Y15) cashflow projections."},
]


# ===================================================================
#  Pre-compute DAMAS daily aggregates (UP + DOWN separately)
# ===================================================================
def load_damas_daily_directional() -> pd.DataFrame:
    """Per-day aggregates split by direction: market_up_mwh, market_down_mwh,
    avg_price_up/down (from activated slots only), max_price_up/down,
    slots_activated_up/down (count of slots where market was activated)."""
    fr = FRService()
    df = fr._load_fr_data().copy()
    last = df["date"].max().date()
    first = (pd.Timestamp(last) - pd.Timedelta(days=365)).date()
    df = df[(df["date"].dt.date >= first) & (df["date"].dt.date <= last)]
    df["date_only"] = df["date"].dt.date

    daily = pd.DataFrame({"date": sorted(df["date_only"].unique())})
    daily["weekday"] = daily["date"].apply(lambda d: pd.Timestamp(d).day_name())

    by_day = df.groupby("date_only")
    daily["market_up_mwh"] = by_day["afrr_up_activated_mwh"].sum().reindex(daily["date"]).values
    daily["market_down_mwh"] = by_day["afrr_down_activated_mwh"].sum().reindex(daily["date"]).values

    # For avg/max price, restrict to slots where the market was activated
    # AND price is in valid range.
    def _per_dir(df_, price_col, act_col):
        valid = df_[
            (df_[act_col] > 0)
            & (df_[price_col].abs() > 0)
            & (df_[price_col].abs() < PRICE_VALID_MAX)
        ]
        if valid.empty:
            return pd.DataFrame({"avg_price": [], "max_price": [], "slots": []})
        g = valid.groupby("date_only")[price_col]
        return pd.DataFrame({
            "avg_price": g.apply(lambda s: s.abs().mean()),
            "max_price": g.apply(lambda s: s.abs().max()),
            "slots": valid.groupby("date_only").size(),
        })

    up_stats = _per_dir(df, "afrr_up_price_eur", "afrr_up_activated_mwh")
    down_stats = _per_dir(df, "afrr_down_price_eur", "afrr_down_activated_mwh")

    daily["avg_price_up"] = up_stats["avg_price"].reindex(daily["date"]).fillna(0.0).values
    daily["max_price_up"] = up_stats["max_price"].reindex(daily["date"]).fillna(0.0).values
    daily["slots_up"] = up_stats["slots"].reindex(daily["date"]).fillna(0).astype(int).values

    daily["avg_price_down"] = down_stats["avg_price"].reindex(daily["date"]).fillna(0.0).values
    daily["max_price_down"] = down_stats["max_price"].reindex(daily["date"]).fillna(0.0).values
    daily["slots_down"] = down_stats["slots"].reindex(daily["date"]).fillna(0).astype(int).values

    return daily


# ===================================================================
#  Excel formatting
# ===================================================================
def _styles():
    from openpyxl.styles import Border, Font, PatternFill, Side
    border = Side(border_style="thin", color="BFBFBF")
    return {
        "bold": Font(bold=True, size=11),
        "bold_white": Font(bold=True, color="FFFFFF", size=11),
        "bold_huge": Font(bold=True, size=14, color="FFFFFF"),
        "header": PatternFill("solid", fgColor="1F4E78"),
        "title": PatternFill("solid", fgColor="2E75B6"),
        "input_fill": PatternFill("solid", fgColor="FFFFCC"),  # editable
        "subtotal": PatternFill("solid", fgColor="BDD7EE"),
        "total": PatternFill("solid", fgColor="FFEB99"),
        "rev": PatternFill("solid", fgColor="E2EFDA"),
        "box": Border(left=border, right=border, top=border, bottom=border),
    }


EUR_FMT = '"€"#,##0.00'
EUR_BIG = '"€"#,##0;[Red]"-€"#,##0'
EUR_MWH_FMT = '"€"#,##0.00'
EUR_MW_H_FMT = '"€"#,##0.00'
MWH_FMT = '#,##0.00'
PCT_FMT = '0.00%'


# ===================================================================
#  Sheet writers
# ===================================================================
def write_inputs_sheet(ws, styles):
    """Editable inputs with named ranges. Yellow = editable."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Inputs — edit yellow cells to recompute every scenario"
    c.font = styles["bold_huge"]
    c.fill = styles["title"]
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 32

    wb = ws.parent
    cur_row = 3

    def section(label):
        nonlocal cur_row
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
        ws.cell(row=cur_row, column=1, value=label).font = styles["bold_white"]
        ws.cell(row=cur_row, column=1).fill = styles["header"]
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", indent=1)
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

    def input_row(label, value, name, fmt, note):
        nonlocal cur_row
        ws.cell(row=cur_row, column=1, value=label).font = styles["bold"]
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left")
        cell = ws.cell(row=cur_row, column=2, value=value)
        cell.fill = styles["input_fill"]
        cell.font = styles["bold"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = styles["box"]
        if fmt:
            cell.number_format = fmt
        defined = DefinedName(name=name, attr_text=f"Inputs!${get_column_letter(2)}${cur_row}")
        wb.defined_names[name] = defined
        ws.cell(row=cur_row, column=3, value=note).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.merge_cells(start_row=cur_row, start_column=3, end_row=cur_row, end_column=5)
        for col in range(1, 6):
            ws.cell(row=cur_row, column=col).border = styles["box"]
        cur_row += 1

    section("Battery configuration (10 MW / 20 MWh canonical)")
    input_row("Power", POWER_MW_DEFAULT, "power_mw", '0.00" MW"', "Battery rated power")
    input_row("Energy capacity", CAPACITY_MWH_DEFAULT, "capacity_mwh", '0.00" MWh"', "Daily throughput cap")
    input_row("Hours per day", HOURS_PER_DAY, "hours_per_day", '0', "24h availability")
    input_row("Acceptance rate", ACCEPTANCE_DEFAULT, "acceptance_rate", PCT_FMT, "Balanced strategy 80%")

    cur_row += 1
    section("DAMAS clearing — capacity bid base (€/MW/h)")
    input_row("aFRR Up clearing", DAMAS_AFRR_UP_DEFAULT, "damas_afrr_up", EUR_MW_H_FMT,
              "Real DAMAS public-tender sample mean. Edit if you have different data.")
    input_row("aFRR Down clearing", DAMAS_AFRR_DOWN_DEFAULT, "damas_afrr_down", EUR_MW_H_FMT, "ditto")

    cur_row += 1
    section("Stacking factors (physical hardware constraint)")
    input_row("aFRR capacity stacking", STACK_CAPACITY_DEFAULT, "stack_capacity", PCT_FMT,
              "Pure availability — independent of other streams")
    input_row("aFRR activation stacking", STACK_ACTIVATION_DEFAULT, "stack_activation", PCT_FMT,
              "Activation MWh competes with PZU throughput")

    cur_row += 1
    for scen in SCENARIOS:
        section(f"Scenario {scen['key']} — {scen['label']}")
        input_row("  Market share", scen["market_share"], f"market_share_{scen['key']}",
                  PCT_FMT, "Captured share of TSO activations (~10% small entrant)")
        input_row("  Capacity multiplier", scen["capacity_mult"], f"capacity_mult_{scen['key']}",
                  PCT_FMT, "× DAMAS clearing → bid")
        input_row("  Activation multiplier", scen["activation_mult"], f"activation_mult_{scen['key']}",
                  PCT_FMT, "× actual settlement price → activation revenue")

    widths = [32, 16, 22, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_damas_daily_sheet(ws, daily: pd.DataFrame, styles):
    """Hidden helper: per-day per-direction DAMAS aggregates."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    headers = [
        "Date", "Weekday",
        "Market UP MWh", "Market DOWN MWh",
        "Avg price UP", "Avg price DOWN",
        "Max price UP", "Max price DOWN",
        "Slots UP", "Slots DOWN",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = styles["bold_white"]
        c.fill = styles["header"]
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = styles["box"]
    ws.row_dimensions[1].height = 36

    for i, row in daily.iterrows():
        r = 2 + i
        ws.cell(row=r, column=1, value=pd.Timestamp(row["date"]).date()).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=row["weekday"])
        ws.cell(row=r, column=3, value=float(row["market_up_mwh"])).number_format = MWH_FMT
        ws.cell(row=r, column=4, value=float(row["market_down_mwh"])).number_format = MWH_FMT
        ws.cell(row=r, column=5, value=float(row["avg_price_up"])).number_format = EUR_MWH_FMT
        ws.cell(row=r, column=6, value=float(row["avg_price_down"])).number_format = EUR_MWH_FMT
        ws.cell(row=r, column=7, value=float(row["max_price_up"])).number_format = EUR_MWH_FMT
        ws.cell(row=r, column=8, value=float(row["max_price_down"])).number_format = EUR_MWH_FMT
        ws.cell(row=r, column=9, value=int(row["slots_up"]))
        ws.cell(row=r, column=10, value=int(row["slots_down"]))
        for col in range(1, 11):
            ws.cell(row=r, column=col).border = styles["box"]
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")

    for i, w in enumerate([12, 12, 14, 14, 14, 14, 14, 14, 11, 11], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"
    # Hide this sheet — it's helper data
    ws.sheet_state = "hidden"


def write_scenario_sheet(ws, scenario: dict, daily: pd.DataFrame, styles):
    """Per-scenario sheet matching the original 14-column layout, with
    formulas in the computable cells."""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    # Header row (same as original)
    headers = [
        "Date", "Weekday", "Selected", "Recommended bid (€/MW/h)",
        "Acceptance rate", "Capacity revenue (€)", "Slots activated (#/96)",
        "Avg activation price (€/MWh)", "Max activation price (€/MWh)",
        "Market activations (MWh)", "Captured energy (MWh)",
        "Battery cap binding", "Activation revenue (€)", "Daily revenue (€)",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = styles["bold_white"]
        c.fill = styles["header"]
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = styles["box"]
    ws.row_dimensions[1].height = 42

    K = scenario["key"]
    n = len(daily)
    for i in range(n):
        r = 2 + i
        damas_row = 2 + i  # DAMAS_Daily row index matches

        # Capacity revenue helper formulas (used for direction-selection):
        # We need to compute UP and DOWN total revenue, then pick the larger.
        # All inline in the "Selected" formula.
        up_cap = (
            f"power_mw*hours_per_day*damas_afrr_up*"
            f"capacity_mult_{K}*stack_capacity*acceptance_rate"
        )
        down_cap = (
            f"power_mw*hours_per_day*damas_afrr_down*"
            f"capacity_mult_{K}*stack_capacity*acceptance_rate"
        )
        up_captured = (
            f"MIN(DAMAS_Daily!C{damas_row}*market_share_{K}*acceptance_rate*stack_activation,"
            f"power_mw*hours_per_day,capacity_mwh)"
        )
        down_captured = (
            f"MIN(DAMAS_Daily!D{damas_row}*market_share_{K}*acceptance_rate*stack_activation,"
            f"power_mw*hours_per_day,capacity_mwh)"
        )
        up_act = f"({up_captured})*DAMAS_Daily!E{damas_row}*activation_mult_{K}"
        down_act = f"({down_captured})*DAMAS_Daily!F{damas_row}*activation_mult_{K}"
        up_total = f"({up_cap}+{up_act})"
        down_total = f"({down_cap}+{down_act})"

        # Col 1: Date (from DAMAS_Daily)
        ws.cell(row=r, column=1, value=f"=DAMAS_Daily!A{damas_row}").number_format = "yyyy-mm-dd"
        # Col 2: Weekday
        ws.cell(row=r, column=2, value=f"=DAMAS_Daily!B{damas_row}")
        # Col 3: Selected (aFRR+ if up_total > down_total)
        ws.cell(row=r, column=3,
                value=f'=IF({up_total}>={down_total},"aFRR+","aFRR-")')
        # Col 4: Recommended bid
        ws.cell(row=r, column=4,
                value=f'=IF(C{r}="aFRR+",damas_afrr_up*capacity_mult_{K},'
                      f'damas_afrr_down*capacity_mult_{K})').number_format = EUR_MW_H_FMT
        # Col 5: Acceptance rate
        ws.cell(row=r, column=5, value="=acceptance_rate").number_format = PCT_FMT
        # Col 6: Capacity revenue
        ws.cell(row=r, column=6,
                value=f'=IF(C{r}="aFRR+",{up_cap},{down_cap})').number_format = EUR_BIG
        # Col 7: Slots activated
        ws.cell(row=r, column=7,
                value=f'=IF(C{r}="aFRR+",DAMAS_Daily!I{damas_row},DAMAS_Daily!J{damas_row})')
        # Col 8: Avg activation price
        ws.cell(row=r, column=8,
                value=f'=IF(C{r}="aFRR+",DAMAS_Daily!E{damas_row},DAMAS_Daily!F{damas_row})').number_format = EUR_MWH_FMT
        # Col 9: Max activation price
        ws.cell(row=r, column=9,
                value=f'=IF(C{r}="aFRR+",DAMAS_Daily!G{damas_row},DAMAS_Daily!H{damas_row})').number_format = EUR_MWH_FMT
        # Col 10: Market activations
        ws.cell(row=r, column=10,
                value=f'=IF(C{r}="aFRR+",DAMAS_Daily!C{damas_row},DAMAS_Daily!D{damas_row})').number_format = MWH_FMT
        # Col 11: Captured energy
        ws.cell(row=r, column=11,
                value=f'=IF(C{r}="aFRR+",{up_captured},{down_captured})').number_format = MWH_FMT
        # Col 12: Battery cap binding
        ws.cell(row=r, column=12,
                value=f'=IF(K{r}>=capacity_mwh-0.01,"yes","no")')
        # Col 13: Activation revenue
        ws.cell(row=r, column=13,
                value=f'=IF(C{r}="aFRR+",{up_act},{down_act})').number_format = EUR_BIG
        # Col 14: Daily revenue
        ws.cell(row=r, column=14, value=f"=F{r}+M{r}").number_format = EUR_BIG

        # Color the revenue columns
        scen_fill = PatternFill("solid", fgColor=scenario["color"])
        for col in range(1, 15):
            ws.cell(row=r, column=col).border = styles["box"]
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
        for col in (6, 13, 14):
            ws.cell(row=r, column=col).fill = scen_fill
        ws.cell(row=r, column=14).font = styles["bold"]

    # TOTAL row
    total_row = 2 + n
    ws.cell(row=total_row, column=1, value="TOTAL").font = styles["bold"]
    ws.cell(row=total_row, column=2, value=f"{n} days").font = styles["bold"]
    ws.cell(row=total_row, column=3, value="—")
    for col_idx in (6, 7, 11, 13, 14):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx,
                value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})")
        if col_idx in (6, 13, 14):
            ws.cell(row=total_row, column=col_idx).number_format = EUR_BIG
        elif col_idx == 11:
            ws.cell(row=total_row, column=col_idx).number_format = MWH_FMT
    for col in range(1, 15):
        ws.cell(row=total_row, column=col).font = styles["bold"]
        ws.cell(row=total_row, column=col).fill = styles["total"]
        ws.cell(row=total_row, column=col).border = styles["box"]
        ws.cell(row=total_row, column=col).alignment = Alignment(horizontal="center")

    # Column widths
    widths = [11, 11, 10, 14, 12, 14, 12, 14, 14, 14, 14, 12, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"


def write_summary_sheet(ws, daily: pd.DataFrame, styles):
    """Match the original Summary layout: battery config + scenario table
    + descriptions + notes. Annual totals pulled live from scenario sheets."""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = (
        "aFRR scenarios — Current vs PICASSO vs Mature vs Bear  "
        "|  2025-05-01 → 2026-05-01 (formula-driven)"
    )
    c.font = styles["bold_huge"]
    c.fill = styles["title"]
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    # Battery configuration block
    cur_row = 3
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)
    ws.cell(row=cur_row, column=1,
            value="Battery configuration (common to all 4 scenarios — see Inputs sheet)").font = styles["bold_white"]
    ws.cell(row=cur_row, column=1).fill = styles["header"]
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", indent=1)
    cur_row += 1
    rows = [
        ("Power (MW)", "=power_mw", "Battery rated power"),
        ("Energy capacity (MWh)", "=capacity_mwh", "Daily throughput cap (1 cycle/day)"),
        ("Acceptance rate", "=acceptance_rate", "Balanced strategy"),
        ("Real DAMAS aFRR+ (€/MW/h)", "=damas_afrr_up", "Edit on Inputs sheet"),
        ("Real DAMAS aFRR- (€/MW/h)", "=damas_afrr_down", "Edit on Inputs sheet"),
        ("Stacking — capacity", "=stack_capacity", "Physical constraint"),
        ("Stacking — activation", "=stack_activation", "Physical constraint"),
    ]
    for label, formula, note in rows:
        ws.cell(row=cur_row, column=1, value=label).font = styles["bold"]
        c = ws.cell(row=cur_row, column=2, value=formula)
        if "MW)" in label and "(MWh)" not in label:
            c.number_format = '0.00" MW"'
        elif "(MWh)" in label:
            c.number_format = '0.00" MWh"'
        elif "%" in label or "rate" in label or "acking" in label:
            c.number_format = PCT_FMT
        else:
            c.number_format = EUR_MW_H_FMT
        ws.cell(row=cur_row, column=3, value=note)
        ws.merge_cells(start_row=cur_row, start_column=3, end_row=cur_row, end_column=7)
        for col in range(1, 8):
            ws.cell(row=cur_row, column=col).border = styles["box"]
            ws.cell(row=cur_row, column=col).alignment = Alignment(horizontal="left")
        cur_row += 1

    # Scenario table — pulls totals via formulas from each scenario sheet
    cur_row += 1
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)
    ws.cell(row=cur_row, column=1,
            value="Scenario definitions + LIVE annual totals (formulas pull from each scenario sheet's TOTAL row)").font = styles["bold_white"]
    ws.cell(row=cur_row, column=1).fill = styles["header"]
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", indent=1)
    cur_row += 1
    headers = ["Scenario", "Market share", "Capacity ×", "Activation ×",
               "Capacity rev (€)", "Activation rev (€)", "Total annual (€)"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=cur_row, column=col_idx, value=h)
        c.font = styles["bold_white"]
        c.fill = styles["title"]
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = styles["box"]
    ws.row_dimensions[cur_row].height = 30
    cur_row += 1

    n_days = len(daily)
    total_row_idx = 2 + n_days  # TOTAL row index in scenario sheets
    fills = ["DDEBF7", "FCE4D6", "FFF2CC", "F8CBAD"]
    for s_idx, scen in enumerate(SCENARIOS):
        K = scen["key"]
        sheet = scen["short"]
        ws.cell(row=cur_row, column=1, value=scen["label"]).font = styles["bold"]
        ws.cell(row=cur_row, column=2, value=f"=market_share_{K}").number_format = PCT_FMT
        ws.cell(row=cur_row, column=3, value=f"=capacity_mult_{K}").number_format = "0.00"
        ws.cell(row=cur_row, column=4, value=f"=activation_mult_{K}").number_format = "0.00"
        ws.cell(row=cur_row, column=5, value=f"='{sheet}'!F{total_row_idx}").number_format = EUR_BIG
        ws.cell(row=cur_row, column=6, value=f"='{sheet}'!M{total_row_idx}").number_format = EUR_BIG
        ws.cell(row=cur_row, column=7, value=f"='{sheet}'!N{total_row_idx}").number_format = EUR_BIG
        scen_fill = PatternFill("solid", fgColor=fills[s_idx])
        for col in range(1, 8):
            ws.cell(row=cur_row, column=col).fill = scen_fill
            ws.cell(row=cur_row, column=col).border = styles["box"]
            ws.cell(row=cur_row, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=cur_row, column=7).font = styles["bold"]
        cur_row += 1

    # Descriptions block
    cur_row += 1
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)
    ws.cell(row=cur_row, column=1, value="Scenario descriptions").font = styles["bold_white"]
    ws.cell(row=cur_row, column=1).fill = styles["header"]
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", indent=1)
    cur_row += 1
    for s_idx, scen in enumerate(SCENARIOS):
        ws.cell(row=cur_row, column=1, value=scen["label"]).font = styles["bold"]
        ws.cell(row=cur_row, column=1).fill = PatternFill("solid", fgColor=fills[s_idx])
        ws.cell(row=cur_row, column=2, value=scen["description"])
        ws.merge_cells(start_row=cur_row, start_column=2, end_row=cur_row, end_column=7)
        ws.cell(row=cur_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=cur_row, column=2).fill = PatternFill("solid", fgColor=fills[s_idx])
        for col in range(1, 8):
            ws.cell(row=cur_row, column=col).border = styles["box"]
        ws.row_dimensions[cur_row].height = 80
        cur_row += 1

    # Notes block
    cur_row += 1
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)
    ws.cell(row=cur_row, column=1, value="Notes").font = styles["bold_white"]
    ws.cell(row=cur_row, column=1).fill = styles["header"]
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", indent=1)
    cur_row += 1
    notes = [
        "FORMULA-DRIVEN VERSION. Edit any yellow cell on Inputs to recompute every scenario instantly.",
        "All four scenarios run the SAME 10 MW / 20 MWh battery on the SAME real DAMAS market data (366 days × 96 slots/day) — only the market_share, capacity_mult, and activation_mult parameters differ.",
        "Capacity revenue per day = power × 24h × bid × acceptance × stack_capacity, where bid = MAX(damas_up, damas_down) × capacity_mult per scenario.",
        "Activation revenue per day = MIN(market × share × acceptance × stack_activation, power × hours, capacity) × actual_avg_settlement_price × activation_mult.",
        "PICASSO assumption (-25% step-down): conservative read of ENTSO-E PICASSO go-live data from DE/AT/CZ which saw 20-40% reduction in extreme tail prices. Romania expected to connect 2026-2027.",
        "Mature market (5% share, -40% capacity): extrapolation from Germany's BESS build-out 2022-2024 and current ~€3-4/MW/h aFRR clearing.",
        "Bear case (2% share, -60% capacity): long-term steady state matching Germany today after their compression, applied to Romania for Y8-Y15 cashflow stress testing.",
        "Numbers are SIMULATED on real DAMAS data — NOT bankable. Compliance gates (ANRE / OPCOM / PRE / BRP / BSP / FSE) and tariff exemption math apply downstream in investment_service.analyze().",
    ]
    for n in notes:
        ws.cell(row=cur_row, column=1, value="• " + n)
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)
        ws.cell(row=cur_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[cur_row].height = 32
        cur_row += 1

    widths = [38, 18, 14, 14, 18, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ===================================================================
#  Main
# ===================================================================
def main() -> None:
    print("Loading DAMAS daily aggregates…")
    daily = load_damas_daily_directional()
    print(f"  {len(daily)} days, {daily['date'].min()} → {daily['date'].max()}")

    out_dir = ROOT / "reports"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "aFRR_scenarios_PICASSO_market_share_2025-05-01_to_2026-05-01_FORMULAS.xlsx"

    # Build workbook with the EXACT 5-sheet structure of the original
    # PLUS Inputs + DAMAS_Daily helper sheets
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        pd.DataFrame().to_excel(xl, sheet_name="Inputs", index=False)
        pd.DataFrame().to_excel(xl, sheet_name="DAMAS_Daily", index=False)
        pd.DataFrame().to_excel(xl, sheet_name="Summary", index=False)
        for scen in SCENARIOS:
            pd.DataFrame().to_excel(xl, sheet_name=scen["short"], index=False)

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    styles = _styles()

    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

    write_inputs_sheet(wb["Inputs"], styles)
    write_damas_daily_sheet(wb["DAMAS_Daily"], daily, styles)
    write_summary_sheet(wb["Summary"], daily, styles)
    for scen in SCENARIOS:
        write_scenario_sheet(wb[scen["short"]], scen, daily, styles)

    # Reorder: Summary, Scenario_A, Scenario_B, Scenario_C, Scenario_D, Inputs, DAMAS_Daily
    desired_order = (
        ["Summary"]
        + [scen["short"] for scen in SCENARIOS]
        + ["Inputs", "DAMAS_Daily"]
    )
    wb._sheets = [wb[name] for name in desired_order]
    wb.active = 0
    wb.save(out_path)

    print(f"\nWrote: {out_path}")
    print("\nSheets in order (matches original PICASSO file + 2 helper sheets):")
    print("  1. Summary             — battery config + scenario table (LIVE FORMULAS) + descriptions + notes")
    print("  2. Scenario_A_Current  — 366 daily rows × 14 cols, all formulas")
    print("  3. Scenario_B_PICASSO  — same")
    print("  4. Scenario_C_Mature   — same")
    print("  5. Scenario_D_Bear     — same")
    print("  6. Inputs              — yellow editable cells with named ranges")
    print("  7. DAMAS_Daily         — hidden helper (per-day market data)")


if __name__ == "__main__":
    main()
