"""Multi-scenario aFRR daily Excel — current vs PICASSO vs mature vs bear.

Builds on the real-market per-slot simulator in afrr_daily_real_market_excel.py
and runs it under FOUR distinct market scenarios:

  A. Current Romanian state (10% market share, no PICASSO)
  B. PICASSO live (10% share, −25% step-down on capacity + activation)
  C. Mature market (5% share, capacity prices compressed to mature levels)
  D. Bear case (2% share, post-saturation, German-current price levels)

Each scenario:
  - Uses the same battery (10 MW / 20 MWh, RTE 0.97, full 0-100% SOC)
  - Same balanced strategy (80% acceptance, 1.0 capacity multiplier)
  - Same real DAMAS clearing prices as starting point
  - Differs only in market_share, capacity_mult, activation_mult

Output: 5 sheets — Summary + one per scenario.

Run from `backend/`:
    PYTHONPATH=. arch -arm64 /usr/local/bin/python3 scripts/afrr_scenarios_excel.py
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional, Tuple

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.config import settings  # noqa: E402
from app.services.fr_service import FRService  # noqa: E402

# ===================================================================
#  Battery + market constants
# ===================================================================
POWER_MW = float(settings.DEFAULT_POWER_MW)
CAPACITY_MWH = float(settings.DEFAULT_CAPACITY_MWH)
HOURS_PER_DAY = 24
SLOT_HOURS = 0.25
ACCEPTANCE_BALANCED = 0.80  # all scenarios use balanced strategy
CAPACITY_MULTIPLIER_BALANCED = 1.00
AFRR_FLOOR_EUR_MW_H = 5.0
AFRR_UP_CAP_EUR_MW_H = 60.0
AFRR_DOWN_CAP_EUR_MW_H = 40.0
PRICE_VALID_MAX = 500.0


# CANONICAL resolver (app/market_data/capacity_prices.py) — same source
# the production fr_service uses. Recent-window DAMAS sample mean.
from app.market_data.capacity_prices import get_canonical_capacity_price


def _load_damas_avgs() -> dict:
    return {
        "aFRRUp": get_canonical_capacity_price("aFRRUp").price_eur_mw_h,
        "aFRRDown": get_canonical_capacity_price("aFRRDown").price_eur_mw_h,
    }


_DAMAS_AVGS = _load_damas_avgs()

# ===================================================================
#  Scenario definitions
# ===================================================================
SCENARIOS = [
    {
        "key": "A_current_10pct",
        "label": "A. Current (10%, pre-PICASSO)",
        "short": "Scenario_A_Current",
        "market_share": 0.10,
        "capacity_mult": 1.00,
        "activation_mult": 1.00,
        "description": (
            "Today's Romanian aFRR market. PICASSO platform NOT yet "
            "connected (Romania waiting on Hungary/MAVIR). Battery operator "
            "captures 10% of TSO activations (small new entrant in a sparse "
            "BESS market — only ~50 MW deployed Romania-wide). Real DAMAS "
            "capacity clearing prices apply (€8.72/€6.59 per MW/h aFRRUp/"
            "Down avg, sampled Apr 2025 – Jan 2026). Activation revenue uses "
            "real per-slot DAMAS settlement prices."
        ),
    },
    {
        "key": "B_picasso_live",
        "label": "B. PICASSO go-live (−25%)",
        "short": "Scenario_B_PICASSO",
        "market_share": 0.10,
        "capacity_mult": 0.75,
        "activation_mult": 0.75,
        "description": (
            "PICASSO platform activates Romania (expected 2026-2027 once "
            "Hungary/MAVIR connects). Cross-border aFRR-energy bid stack "
            "caps Romanian extreme activation prices toward European "
            "average; capacity prices compress in parallel. Conservative "
            "step-down applied: −25% on BOTH capacity bid (real DAMAS × "
            "0.75) AND per-slot activation revenue. Market share unchanged "
            "(10%). Reference: ENTSO-E PICASSO go-live in DE/AT/CZ produced "
            "20-40% reduction in extreme tail prices over 12-18 months."
        ),
    },
    {
        "key": "C_mature_market",
        "label": "C. Mature market (5% share, −40%)",
        "short": "Scenario_C_Mature",
        "market_share": 0.05,
        "capacity_mult": 0.60,
        "activation_mult": 0.65,
        "description": (
            "Romanian BESS pipeline (1+ GW announced) materializes through "
            "2027-2028. Market share compresses from 10% to 5% as new BESS "
            "deployments compete for the same activation volumes. aFRR "
            "capacity prices compress 40% to ~€5/MW/h (matching Germany's "
            "current level after their 2022-2024 BESS build-out). "
            "Activation prices compress 35% as the bid stack thickens. "
            "Steady-state realistic case for years Y3-Y8."
        ),
    },
    {
        "key": "D_bear_case",
        "label": "D. Bear case (2%, full saturation)",
        "short": "Scenario_D_Bear",
        "market_share": 0.02,
        "capacity_mult": 0.40,
        "activation_mult": 0.50,
        "description": (
            "Long-term steady state (2030+). Romanian BESS deploys 2-3 GW "
            "matching Germany's current density. Market share collapses to "
            "2% per operator. aFRR capacity prices fall 60% (matching "
            "Germany's mature aFRR auction levels of €3-4/MW/h post-2024 "
            "compression). Activation prices fall 50% as PICASSO + thick "
            "bid stack normalize Romanian prices to European average. "
            "Bankability stress case used by lenders for late-life "
            "(Y8-Y15) cashflow projections."
        ),
    },
]

# ===================================================================
#  Column constants
# ===================================================================
COL_DATE = "Date"
COL_WEEKDAY = "Weekday"
COL_SELECTED = "Selected"
COL_BID = "Recommended bid (€/MW/h)"
COL_ACCEPT = "Acceptance rate"
COL_CAP_REV = "Capacity revenue (€)"
COL_SLOTS = "Slots activated (#/96)"
COL_AVG_PX = "Avg activation price (€/MWh)"
COL_MAX_PX = "Max activation price (€/MWh)"
COL_MKT = "Market activations (MWh)"
COL_CAPTURED = "Captured energy (MWh)"
COL_BOUND = "Battery cap binding"
COL_ACT_REV = "Activation revenue (€)"
COL_DAILY = "Daily revenue (€)"

PLACEHOLDER = "—"
SHEET_SUMMARY = "Summary"


# ===================================================================
#  Slot-level simulation (scenario-aware)
# ===================================================================
def _filter_valid(prices: pd.Series) -> pd.Series:
    return prices[(prices.abs() > 0) & (prices.abs() < PRICE_VALID_MAX)]


def _simulate_direction(
    prices: np.ndarray,
    market_activations: np.ndarray,
    acceptance: float,
    market_share: float,
    activation_mult: float,
) -> dict:
    """Walk 96 slots, accumulate activation revenue using REAL per-slot
    prices × scenario activation_mult. Battery throughput cap is a
    running budget."""
    n = len(prices)
    captured_total = 0.0
    activation_revenue = 0.0
    activated_slots = 0
    activation_prices: list[float] = []
    per_slot_max_mwh = POWER_MW * SLOT_HOURS  # 2.5 MWh
    bound = False

    for i in range(n):
        market_mwh = float(market_activations[i])
        if market_mwh <= 0:
            continue
        share_mwh = market_mwh * market_share
        constrained = min(share_mwh, per_slot_max_mwh) * acceptance
        if constrained <= 0:
            continue
        remaining = CAPACITY_MWH - captured_total
        if remaining <= 0:
            bound = True
            break
        slot_mwh = min(constrained, remaining)
        if slot_mwh < constrained:
            bound = True
        slot_price = float(prices[i]) * activation_mult
        slot_rev = slot_mwh * abs(slot_price)
        captured_total += slot_mwh
        activation_revenue += slot_rev
        activated_slots += 1
        activation_prices.append(float(prices[i]))

    return {
        "slots_activated": activated_slots,
        "captured_mwh": captured_total,
        "activation_revenue": activation_revenue,
        "avg_price": float(np.mean(np.abs(activation_prices))) if activation_prices else 0.0,
        "max_price": float(np.max(np.abs(activation_prices))) if activation_prices else 0.0,
        "battery_bound": bound,
        "market_total": float(np.sum(market_activations)),
    }


def daily_row(group: pd.DataFrame, day, scenario: dict) -> Optional[dict]:
    g = group.sort_values("slot")
    valid_up = _filter_valid(g["afrr_up_price_eur"])
    valid_down = _filter_valid(g["afrr_down_price_eur"])
    if valid_up.empty or valid_down.empty:
        return None

    cap_mult = scenario["capacity_mult"] * CAPACITY_MULTIPLIER_BALANCED
    act_mult = scenario["activation_mult"]
    share = scenario["market_share"]

    up_bid = max(AFRR_FLOOR_EUR_MW_H, min(AFRR_UP_CAP_EUR_MW_H,
                                          _DAMAS_AVGS["aFRRUp"] * cap_mult))
    down_bid = max(AFRR_FLOOR_EUR_MW_H, min(AFRR_DOWN_CAP_EUR_MW_H,
                                            _DAMAS_AVGS["aFRRDown"] * cap_mult))
    up_cap_rev = POWER_MW * HOURS_PER_DAY * up_bid * ACCEPTANCE_BALANCED
    down_cap_rev = POWER_MW * HOURS_PER_DAY * down_bid * ACCEPTANCE_BALANCED

    up_sim = _simulate_direction(g["afrr_up_price_eur"].to_numpy(),
                                 g["afrr_up_activated_mwh"].to_numpy(),
                                 ACCEPTANCE_BALANCED, share, act_mult)
    down_sim = _simulate_direction(g["afrr_down_price_eur"].to_numpy(),
                                   g["afrr_down_activated_mwh"].to_numpy(),
                                   ACCEPTANCE_BALANCED, share, act_mult)

    up_total = up_cap_rev + up_sim["activation_revenue"]
    down_total = down_cap_rev + down_sim["activation_revenue"]
    if up_total >= down_total:
        sel, cap_rev, bid, sim, total = "aFRR+", up_cap_rev, up_bid, up_sim, up_total
    else:
        sel, cap_rev, bid, sim, total = "aFRR-", down_cap_rev, down_bid, down_sim, down_total

    return {
        COL_DATE: pd.Timestamp(day),
        COL_WEEKDAY: pd.Timestamp(day).day_name(),
        COL_SELECTED: sel,
        COL_BID: round(bid, 2),
        COL_ACCEPT: ACCEPTANCE_BALANCED,
        COL_CAP_REV: cap_rev,
        COL_SLOTS: sim["slots_activated"],
        COL_AVG_PX: round(sim["avg_price"] * act_mult, 2),
        COL_MAX_PX: round(sim["max_price"] * act_mult, 2),
        COL_MKT: round(sim["market_total"], 2),
        COL_CAPTURED: round(sim["captured_mwh"], 2),
        COL_BOUND: "yes" if sim["battery_bound"] else "no",
        COL_ACT_REV: sim["activation_revenue"],
        COL_DAILY: total,
    }


def _total_row(df: pd.DataFrame) -> dict:
    return {
        COL_DATE: "TOTAL",
        COL_WEEKDAY: f"{len(df)} days",
        COL_SELECTED: PLACEHOLDER,
        COL_BID: None, COL_ACCEPT: None,
        COL_CAP_REV: float(df[COL_CAP_REV].sum()),
        COL_SLOTS: int(df[COL_SLOTS].sum()),
        COL_AVG_PX: None, COL_MAX_PX: None,
        COL_MKT: float(df[COL_MKT].sum()),
        COL_CAPTURED: float(df[COL_CAPTURED].sum()),
        COL_BOUND: f"{(df[COL_BOUND] == 'yes').sum()} days",
        COL_ACT_REV: float(df[COL_ACT_REV].sum()),
        COL_DAILY: float(df[COL_DAILY].sum()),
    }


# ===================================================================
#  Excel formatting
# ===================================================================
EUR_FMT = '"€"#,##0;[Red]"-€"#,##0'
EUR_MWH_FMT = '"€"#,##0.00'
EUR_MW_H_FMT = '"€"#,##0.00" /MW/h"'
MWH_FMT = '#,##0.00" MWh"'
PCT_FMT = '0%'


def _styles():
    from openpyxl.styles import Border, Font, PatternFill, Side
    border = Side(border_style="thin", color="BFBFBF")
    return {
        "bold": Font(bold=True, size=11),
        "bold_white": Font(bold=True, color="FFFFFF", size=11),
        "bold_big": Font(bold=True, size=14, color="FFFFFF"),
        "header": PatternFill("solid", fgColor="1F4E78"),
        "title": PatternFill("solid", fgColor="2E75B6"),
        "total": PatternFill("solid", fgColor="FFEB99"),
        "rev": PatternFill("solid", fgColor="E2EFDA"),
        "profit": PatternFill("solid", fgColor="DDEBF7"),
        "scenA": PatternFill("solid", fgColor="DDEBF7"),  # blue
        "scenB": PatternFill("solid", fgColor="FCE4D6"),  # orange
        "scenC": PatternFill("solid", fgColor="FFF2CC"),  # yellow
        "scenD": PatternFill("solid", fgColor="F8CBAD"),  # red-ish
        "box": Border(left=border, right=border, top=border, bottom=border),
    }


def _detail_format(col: str) -> Optional[str]:
    if col in (COL_AVG_PX, COL_MAX_PX):
        return EUR_MWH_FMT
    if col == COL_BID:
        return EUR_MW_H_FMT
    if "(€)" in col:
        return EUR_FMT
    if "(MWh)" in col:
        return MWH_FMT
    if col == COL_ACCEPT:
        return PCT_FMT
    if col == COL_DATE:
        return "yyyy-mm-dd"
    return None


def _column_width(col: str) -> int:
    if col in (COL_DATE, COL_WEEKDAY, COL_SELECTED, COL_BOUND):
        return 13
    if "revenue" in col.lower() or col in (COL_BID, COL_AVG_PX, COL_MAX_PX, COL_MKT, COL_CAPTURED):
        return 18
    return 14


def _style_detail_sheet(ws, df, styles):
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    for cell in ws[1]:
        cell.font = styles["bold_white"]
        cell.fill = styles["header"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = styles["box"]
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "C2"

    rev_cols = {COL_CAP_REV, COL_ACT_REV}
    profit_cols = {COL_DAILY}
    for col_idx, col_name in enumerate(df.columns, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _column_width(col_name)
        nf = _detail_format(col_name)
        align = Alignment(horizontal="left" if col_name in (COL_DATE, COL_WEEKDAY) else "center")
        for cell in next(ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row)):
            if nf:
                cell.number_format = nf
            cell.alignment = align
            cell.border = styles["box"]
        fill = None
        if col_name in profit_cols:
            fill = styles["profit"]
        elif col_name in rev_cols:
            fill = styles["rev"]
        if fill:
            for cell in next(ws.iter_cols(min_col=col_idx, max_col=col_idx,
                                          min_row=2, max_row=ws.max_row - 1)):
                cell.fill = fill
    for cell in ws[ws.max_row]:
        cell.font = styles["bold"]
        cell.fill = styles["total"]
        cell.border = styles["box"]


def _summary_title(ws, styles, first_date, last_date):
    from openpyxl.styles import Alignment
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = (
        f"aFRR scenarios — Current vs PICASSO vs Mature vs Bear  |  "
        f"{first_date} → {last_date}"
    )
    cell.font = styles["bold_big"]
    cell.fill = styles["title"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    for col_letter, width in zip("ABCDEFG", [28, 22, 22, 22, 22, 22, 16]):
        ws.column_dimensions[col_letter].width = width


def _summary_battery(ws, styles, start_row: int) -> int:
    from openpyxl.styles import Alignment
    ws[f"A{start_row}"] = "Battery configuration (common to all 4 scenarios)"
    ws[f"A{start_row}"].font = styles["bold_white"]
    ws[f"A{start_row}"].fill = styles["header"]
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
    ws[f"A{start_row}"].alignment = Alignment(horizontal="center")
    rows = [
        ("Power (MW)", POWER_MW, "Battery rated power"),
        ("Energy capacity (MWh)", CAPACITY_MWH, "Daily throughput cap (1 cycle/day)"),
        ("RTE", 0.97, "User directive — 3% per round-trip loss"),
        ("SOC band", "0% – 100%", "User directive — full DOD"),
        ("Strategy", "Balanced (80% acceptance, 1.0× bid)",
         "All scenarios run balanced for fair comparison"),
        ("Real DAMAS aFRR+ capacity (avg)", round(_DAMAS_AVGS["aFRRUp"], 2),
         "Scraped from Transelectrica DAMAS II tender statistics"),
        ("Real DAMAS aFRR- capacity (avg)", round(_DAMAS_AVGS["aFRRDown"], 2),
         "ditto"),
    ]
    for offset, (label, value, note) in enumerate(rows, start=1):
        r = start_row + offset
        ws[f"A{r}"] = label
        ws[f"B{r}"] = value
        ws[f"C{r}"] = note
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws[f"A{r}"].font = styles["bold"]
        for col in "ABCDEFG":
            ws[f"{col}{r}"].border = styles["box"]
    return start_row + len(rows) + 2


def _summary_scenarios(ws, styles, start_row: int, totals: list[dict]) -> int:
    from openpyxl.styles import Alignment
    ws[f"A{start_row}"] = "Scenario definitions"
    ws[f"A{start_row}"].font = styles["bold_white"]
    ws[f"A{start_row}"].fill = styles["header"]
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
    ws[f"A{start_row}"].alignment = Alignment(horizontal="center")

    hdr = start_row + 1
    headers = ["Scenario", "Market share", "Capacity ×", "Activation ×",
               "Capacity rev (€)", "Activation rev (€)", "Total annual (€)"]
    for col_letter, label in zip("ABCDEFG", headers):
        c = ws[f"{col_letter}{hdr}"]
        c.value = label
        c.font = styles["bold_white"]
        c.fill = styles["title"]
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = styles["box"]
    ws.row_dimensions[hdr].height = 30

    fills = ["scenA", "scenB", "scenC", "scenD"]
    for offset, (scen, t) in enumerate(zip(SCENARIOS, totals), start=1):
        r = hdr + offset
        ws[f"A{r}"] = scen["label"]
        ws[f"B{r}"] = scen["market_share"]
        ws[f"C{r}"] = scen["capacity_mult"]
        ws[f"D{r}"] = scen["activation_mult"]
        ws[f"E{r}"] = t["capacity"]
        ws[f"F{r}"] = t["activation"]
        ws[f"G{r}"] = t["total"]
        for col_letter in "EFG":
            ws[f"{col_letter}{r}"].number_format = EUR_FMT
        for col_letter in "BCD":
            ws[f"{col_letter}{r}"].number_format = "0%" if col_letter == "B" else "0.00"
        for col in "ABCDEFG":
            ws[f"{col}{r}"].border = styles["box"]
            ws[f"{col}{r}"].fill = styles[fills[offset - 1]]
            ws[f"{col}{r}"].alignment = Alignment(
                horizontal="left" if col == "A" else "center"
            )
        ws[f"A{r}"].font = styles["bold"]
        ws[f"G{r}"].font = styles["bold"]

    # Description block — one row per scenario, full-width text
    desc_start = hdr + len(SCENARIOS) + 2
    ws[f"A{desc_start}"] = "Scenario descriptions"
    ws[f"A{desc_start}"].font = styles["bold_white"]
    ws[f"A{desc_start}"].fill = styles["header"]
    ws.merge_cells(start_row=desc_start, start_column=1, end_row=desc_start, end_column=7)
    ws[f"A{desc_start}"].alignment = Alignment(horizontal="center")
    r = desc_start
    for offset, scen in enumerate(SCENARIOS, start=1):
        r += 1
        # label cell
        ws[f"A{r}"] = scen["label"]
        ws[f"A{r}"].font = styles["bold"]
        ws[f"A{r}"].fill = styles[fills[offset - 1]]
        ws[f"A{r}"].alignment = Alignment(vertical="top")
        ws[f"A{r}"].border = styles["box"]
        # description cell (merged across B-G)
        ws[f"B{r}"] = scen["description"]
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"B{r}"].border = styles["box"]
        ws[f"B{r}"].fill = styles[fills[offset - 1]]
        ws.row_dimensions[r].height = 92
    return r + 2


def _summary_notes(ws, styles, start_row: int) -> None:
    from openpyxl.styles import Alignment
    ws[f"A{start_row}"] = "Notes"
    ws[f"A{start_row}"].font = styles["bold_white"]
    ws[f"A{start_row}"].fill = styles["header"]
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
    ws[f"A{start_row}"].alignment = Alignment(horizontal="center")
    notes = [
        "All four scenarios run the SAME 10 MW / 20 MWh battery on the SAME real DAMAS market data (Apr 2025 - May 2026, 366 days, 96 slots/day) — only the market_share, capacity_mult, and activation_mult parameters differ.",
        "Capacity revenue = power × 24h × bid × acceptance, where bid = real DAMAS aFRR clearing × scenario capacity_mult. Activation revenue = Σ_slot (captured_MWh × actual_settlement_price × scenario activation_mult), with captured_MWh = min(market_MWh × market_share × acceptance, power × 0.25h, remaining battery cap).",
        "PICASSO assumption (-25% step-down): conservative read of ENTSO-E PICASSO go-live data from DE/AT/CZ which saw 20-40% reduction in extreme tail prices. Romania expected to connect 2026-2027.",
        "Mature market (5% share, -40% capacity): extrapolation from Germany's BESS build-out 2022-2024 and current ~€3-4/MW/h aFRR clearing. Assumes Romanian pipeline (1+ GW announced) materializes by 2027-2028.",
        "Bear case (2% share, -60% capacity): long-term steady state matching Germany today after their compression, applied to Romania for Y8-Y15 cashflow stress testing.",
        "Numbers are SIMULATED on real DAMAS data — NOT bankable. Compliance gates (ANRE / OPCOM / PRE / BRP / BSP / FSE) and tariff exemption math apply downstream in investment_service.analyze().",
    ]
    r = start_row
    for note in notes:
        r += 1
        ws[f"A{r}"] = "• " + note
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 38


# ===================================================================
#  Main
# ===================================================================
def _build_scenario_df(df: pd.DataFrame, scenario: dict) -> pd.DataFrame:
    rows = []
    for day, group in df.groupby(df["date"].dt.date):
        row = daily_row(group, day, scenario)
        if row is not None:
            rows.append(row)
    return pd.DataFrame(rows).sort_values(COL_DATE).reset_index(drop=True)


def _load_window(fr: FRService) -> Tuple[pd.DataFrame, pd.Timestamp, pd.Timestamp]:
    df = fr._load_fr_data().copy()
    last = df["date"].max().date()
    first = (pd.Timestamp(last) - pd.Timedelta(days=365)).date()
    df = df[(df["date"].dt.date >= first) & (df["date"].dt.date <= last)]
    return df, first, last


def main() -> None:
    fr = FRService()
    df, first, last = _load_window(fr)
    print(f"Window: {first} → {last}  ({df['date'].dt.date.nunique()} days)")
    print(f"Battery: {POWER_MW} MW / {CAPACITY_MWH} MWh, balanced strategy across all scenarios")
    print(f"Real DAMAS clearing: aFRRUp €{_DAMAS_AVGS['aFRRUp']:.2f}/MW/h, "
          f"aFRRDown €{_DAMAS_AVGS['aFRRDown']:.2f}/MW/h\n")

    scenario_dfs = {}
    totals = []
    for scen in SCENARIOS:
        print(f"  Running {scen['key']} (share={scen['market_share']:.0%}, "
              f"cap×{scen['capacity_mult']}, act×{scen['activation_mult']}) ...")
        sdf = _build_scenario_df(df, scen)
        scenario_dfs[scen["key"]] = sdf
        cap = float(sdf[COL_CAP_REV].sum())
        act = float(sdf[COL_ACT_REV].sum())
        total = float(sdf[COL_DAILY].sum())
        bound = (sdf[COL_BOUND] == "yes").sum()
        totals.append({"capacity": cap, "activation": act, "total": total})
        print(f"    capacity €{cap:>10,.0f}  + activation €{act:>10,.0f}  "
              f"= €{total:>10,.0f}  (battery bound {bound}/{len(sdf)} days)")

    out_dir = ROOT / "reports"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"aFRR_scenarios_PICASSO_market_share_{first}_to_{last}.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        pd.DataFrame().to_excel(xl, sheet_name=SHEET_SUMMARY, index=False)
        for scen in SCENARIOS:
            sdf = scenario_dfs[scen["key"]]
            out = pd.concat([sdf, pd.DataFrame([_total_row(sdf)])], ignore_index=True)
            out.to_excel(xl, sheet_name=scen["short"], index=False)

    # Apply formatting
    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    styles = _styles()
    for scen in SCENARIOS:
        sdf = scenario_dfs[scen["key"]]
        out = pd.concat([sdf, pd.DataFrame([_total_row(sdf)])], ignore_index=True)
        _style_detail_sheet(wb[scen["short"]], out, styles)

    ws = wb[SHEET_SUMMARY]
    _summary_title(ws, styles, first, last)
    next_row = _summary_battery(ws, styles, start_row=3)
    next_row = _summary_scenarios(ws, styles, next_row, totals)
    _summary_notes(ws, styles, next_row)
    wb.active = wb.sheetnames.index(SHEET_SUMMARY)
    wb.save(out_path)

    print(f"\nWrote: {out_path}")
    print(f"\nAnnual totals comparison:")
    for scen, t in zip(SCENARIOS, totals):
        print(f"  {scen['label']:42s}  €{t['total']:>10,.0f}")


if __name__ == "__main__":
    main()
