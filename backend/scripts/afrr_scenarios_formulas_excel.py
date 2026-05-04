"""aFRR scenarios Excel with LIVE FORMULAS — investor can edit assumptions
and see numbers recompute. Banking-grade structure.

Layout (4 sheets, all formula-driven):

  1. Inputs                — editable params: battery, DAMAS clearing,
                              per-scenario market_share / multipliers,
                              stacking factors. Named ranges for clarity.
  2. DAMAS_Daily           — static daily aggregates from real DAMAS data
                              (market activations, avg/max settlement
                              prices). 366 rows × 8 columns.
  3. Scenarios_Daily       — 366 rows × 4 scenarios. Each cell is a
                              FORMULA referencing Inputs + DAMAS_Daily.
                              Editing any input cell recalculates here.
  4. Summary               — monthly + annual aggregates via SUMIFS.
                              Headline KPIs comparing the 4 scenarios.

Run from `backend/`:
    PYTHONPATH=. arch -arm64 /usr/local/bin/python3 scripts/afrr_scenarios_formulas_excel.py
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.config import settings  # noqa: E402
from app.market_data.capacity_prices import get_canonical_capacity_price  # noqa: E402
from app.services.fr_service import FRService  # noqa: E402

# ===================================================================
#  Defaults to embed in the editable Inputs sheet
# ===================================================================
POWER_MW_DEFAULT = float(settings.DEFAULT_POWER_MW)
CAPACITY_MWH_DEFAULT = float(settings.DEFAULT_CAPACITY_MWH)
ACCEPTANCE_DEFAULT = 0.80
HOURS_PER_DAY = 24
PRICE_VALID_MAX = 500.0

# Real DAMAS-derived defaults
DAMAS_AFRR_UP_DEFAULT = get_canonical_capacity_price("aFRRUp").price_eur_mw_h
DAMAS_AFRR_DOWN_DEFAULT = get_canonical_capacity_price("aFRRDown").price_eur_mw_h

# ===================================================================
#  Scenario definitions (defaults; user edits these in the Inputs sheet)
# ===================================================================
SCENARIOS = [
    {"key": "A", "label": "A. Current (10% share, pre-PICASSO)",
     "color": "DDEBF7",
     "market_share": 0.10, "capacity_mult": 1.00, "activation_mult": 1.00,
     "description": "Today's Romania. PICASSO not connected. ~50 MW BESS deployed nationally."},
    {"key": "B", "label": "B. PICASSO go-live (-25%)",
     "color": "FCE4D6",
     "market_share": 0.10, "capacity_mult": 0.75, "activation_mult": 0.75,
     "description": "PICASSO active 2026-2027 → -25% step on capacity + activation."},
    {"key": "C", "label": "C. Mature market (5% share, -40%)",
     "color": "FFF2CC",
     "market_share": 0.05, "capacity_mult": 0.60, "activation_mult": 0.65,
     "description": "Romanian BESS pipeline materializes 2027-2028. Realistic Y3-Y8 lender baseline."},
]


# ===================================================================
#  Compute static DAMAS daily aggregates
# ===================================================================
def load_damas_daily() -> pd.DataFrame:
    """Aggregate the 15-min DAMAS data to daily totals/averages."""
    fr = FRService()
    df = fr._load_fr_data().copy()
    last = df["date"].max().date()
    first = (pd.Timestamp(last) - pd.Timedelta(days=365)).date()
    df = df[(df["date"].dt.date >= first) & (df["date"].dt.date <= last)]

    # Day-level grouping
    df["date_only"] = df["date"].dt.date

    # Filter price outliers for averaging
    up_valid = df[(df["afrr_up_price_eur"].abs() > 0) & (df["afrr_up_price_eur"].abs() < PRICE_VALID_MAX)]
    down_valid = df[(df["afrr_down_price_eur"].abs() > 0) & (df["afrr_down_price_eur"].abs() < PRICE_VALID_MAX)]

    daily = pd.DataFrame({
        "date": sorted(df["date_only"].unique()),
    })
    daily["weekday"] = daily["date"].apply(lambda d: pd.Timestamp(d).day_name())

    # Activation totals (sum across slots) — used for market_share calc
    daily["market_up_mwh"] = (df.groupby("date_only")["afrr_up_activated_mwh"].sum()
                              .reindex(daily["date"]).values)
    daily["market_down_mwh"] = (df.groupby("date_only")["afrr_down_activated_mwh"].sum()
                                 .reindex(daily["date"]).values)

    # Settlement price averages — used for activation revenue
    daily["avg_price_up"] = (up_valid.groupby("date_only")["afrr_up_price_eur"].mean()
                              .reindex(daily["date"]).fillna(0.0).values)
    daily["avg_price_down"] = (down_valid.groupby("date_only")["afrr_down_price_eur"].mean()
                                .reindex(daily["date"]).fillna(0.0).abs().values)

    # Max price for stress visibility
    daily["max_price_up"] = (up_valid.groupby("date_only")["afrr_up_price_eur"].max()
                              .reindex(daily["date"]).fillna(0.0).values)
    daily["max_price_down"] = (down_valid.groupby("date_only")["afrr_down_price_eur"].min()
                                .reindex(daily["date"]).fillna(0.0).abs().values)

    return daily


# ===================================================================
#  Excel writer with formulas
# ===================================================================
def _styles():
    from openpyxl.styles import Border, Font, PatternFill, Side
    border = Side(border_style="thin", color="BFBFBF")
    return {
        "bold": Font(bold=True, size=11),
        "bold_white": Font(bold=True, color="FFFFFF", size=11),
        "bold_huge": Font(bold=True, size=15, color="FFFFFF"),
        "italic": Font(italic=True, size=10, color="595959"),
        "header": PatternFill("solid", fgColor="1F4E78"),
        "title": PatternFill("solid", fgColor="2E75B6"),
        "input_fill": PatternFill("solid", fgColor="FFFFCC"),  # yellow = editable
        "subtotal": PatternFill("solid", fgColor="BDD7EE"),
        "total": PatternFill("solid", fgColor="FFEB99"),
        "rev": PatternFill("solid", fgColor="E2EFDA"),
        "box": Border(left=border, right=border, top=border, bottom=border),
    }


EUR_FMT = '"€"#,##0;[Red]"-€"#,##0'
EUR_MWH_FMT = '"€"#,##0.00'
EUR_MW_H_FMT = '"€"#,##0.00" /MW/h"'
MWH_FMT = '#,##0.00" MWh"'
PCT_FMT = '0.00%'


def write_inputs_sheet(ws, styles):
    """Sheet 1: editable inputs with named ranges. Yellow-fill cells are
    user-editable; everything else is locked metadata."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Inputs — edit the yellow cells to recompute every scenario"
    c.font = styles["bold_huge"]
    c.fill = styles["title"]
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 32

    # Helper to write input rows + register named ranges
    wb = ws.parent
    cur_row = 3

    def section_header(label: str):
        nonlocal cur_row
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
        ws.cell(row=cur_row, column=1, value=label).font = styles["bold_white"]
        ws.cell(row=cur_row, column=1).fill = styles["header"]
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", indent=1)
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

    def input_row(label: str, value, name: str, fmt: Optional[str], note: str):
        nonlocal cur_row
        ws.cell(row=cur_row, column=1, value=label).font = styles["bold"]
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left")
        cell = ws.cell(row=cur_row, column=2, value=value)
        cell.fill = styles["input_fill"]
        cell.font = styles["bold"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = styles["box"]
        if fmt:
            cell.number_format = fmt
        # Register named range
        defined = DefinedName(name=name, attr_text=f"Inputs!${get_column_letter(2)}${cur_row}")
        wb.defined_names[name] = defined
        ws.cell(row=cur_row, column=3, value=note).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.merge_cells(start_row=cur_row, start_column=3, end_row=cur_row, end_column=5)
        for col in range(1, 6):
            ws.cell(row=cur_row, column=col).border = styles["box"]
        cur_row += 1

    section_header("Battery (10 MW / 20 MWh canonical)")
    input_row("Power", POWER_MW_DEFAULT, "power_mw", '0.00" MW"',
              "Inverter rating")
    input_row("Energy capacity", CAPACITY_MWH_DEFAULT, "capacity_mwh", '0.00" MWh"',
              "Daily throughput cap (1 cycle/day)")
    input_row("Hours per day", HOURS_PER_DAY, "hours_per_day", '0',
              "Capacity availability window")
    input_row("Acceptance rate", ACCEPTANCE_DEFAULT, "acceptance_rate", PCT_FMT,
              "Balanced strategy (80% default)")

    cur_row += 1
    section_header("DAMAS clearing (capacity €/MW/h) — recent-window mean from canonical resolver")
    input_row("aFRR Up clearing", DAMAS_AFRR_UP_DEFAULT, "damas_afrr_up", EUR_MW_H_FMT,
              "Real DAMAS public-tender sample mean")
    input_row("aFRR Down clearing", DAMAS_AFRR_DOWN_DEFAULT, "damas_afrr_down", EUR_MW_H_FMT,
              "Real DAMAS public-tender sample mean")

    cur_row += 1
    section_header("Stacking factors (physical hardware constraint)")
    input_row("aFRR capacity stacking", 1.00, "stack_capacity", PCT_FMT,
              "Pure availability — independent of other streams")
    input_row("aFRR activation stacking", 0.60, "stack_activation", PCT_FMT,
              "Activation MWh competes with PZU throughput")

    cur_row += 1
    section_header("Per-scenario parameters (edit to test alternative cases)")
    for scen in SCENARIOS:
        section_label = f"Scenario {scen['key']} — {scen['label']}"
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
        ws.cell(row=cur_row, column=1, value=section_label).font = styles["bold"]
        ws.cell(row=cur_row, column=1).fill = styles["subtotal"]
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", indent=1)
        cur_row += 1
        input_row(f"  Market share", scen["market_share"], f"market_share_{scen['key']}", PCT_FMT,
                  scen["description"])
        input_row(f"  Capacity multiplier", scen["capacity_mult"], f"capacity_mult_{scen['key']}", PCT_FMT,
                  "× DAMAS clearing → bid")
        input_row(f"  Activation multiplier", scen["activation_mult"], f"activation_mult_{scen['key']}", PCT_FMT,
                  "× actual settlement price → activation revenue")

    # Column widths
    widths = [32, 16, 22, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_damas_daily_sheet(ws, daily: pd.DataFrame, styles):
    """Sheet 2: static DAMAS daily aggregates. Referenced by formula
    rows on the Scenarios sheet."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "DAMAS Daily Aggregates (static — from real Romanian TSO data, 2025-05-01 → 2026-05-01)"
    c.font = styles["bold_huge"]
    c.fill = styles["title"]
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    headers = ["Date", "Weekday", "Market aFRR+ activated (MWh)", "Market aFRR- activated (MWh)",
               "Avg aFRR+ settlement price (€/MWh)", "Avg aFRR- settlement price (€/MWh)",
               "Max aFRR+ price (€/MWh)", "Max aFRR- price (€/MWh)"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = styles["bold_white"]
        c.fill = styles["header"]
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = styles["box"]
    ws.row_dimensions[3].height = 36

    for i, row in daily.iterrows():
        r = 4 + i
        ws.cell(row=r, column=1, value=pd.Timestamp(row["date"]).date()).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=row["weekday"])
        ws.cell(row=r, column=3, value=float(row["market_up_mwh"])).number_format = MWH_FMT
        ws.cell(row=r, column=4, value=float(row["market_down_mwh"])).number_format = MWH_FMT
        ws.cell(row=r, column=5, value=float(row["avg_price_up"])).number_format = EUR_MWH_FMT
        ws.cell(row=r, column=6, value=float(row["avg_price_down"])).number_format = EUR_MWH_FMT
        ws.cell(row=r, column=7, value=float(row["max_price_up"])).number_format = EUR_MWH_FMT
        ws.cell(row=r, column=8, value=float(row["max_price_down"])).number_format = EUR_MWH_FMT
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = styles["box"]
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")

    # Widths
    for i, w in enumerate([12, 12, 18, 18, 20, 20, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C4"


def write_scenarios_daily_sheet(ws, daily: pd.DataFrame, styles):
    """Sheet 3: 366 daily rows × 4 scenarios. Every revenue cell is a
    FORMULA referencing Inputs + DAMAS_Daily."""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "Scenarios Daily — formulas reference Inputs + DAMAS_Daily (edit Inputs to recompute)"
    c.font = styles["bold_huge"]
    c.fill = styles["title"]
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 32

    # Column layout: Date + (per scenario × 3 columns: capacity_rev, activation_rev, daily_total)
    base_headers = ["Date", "Weekday"]
    for scen in SCENARIOS:
        base_headers.append(f"{scen['key']} cap rev (€)")
        base_headers.append(f"{scen['key']} act rev (€)")
        base_headers.append(f"{scen['key']} total (€)")

    for col_idx, h in enumerate(base_headers, start=1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = styles["bold_white"]
        c.fill = styles["header"]
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = styles["box"]
    ws.row_dimensions[3].height = 42

    for i in range(len(daily)):
        excel_row = 4 + i
        damas_row = 4 + i  # DAMAS_Daily row index matches
        ws.cell(row=excel_row, column=1, value=f"=DAMAS_Daily!A{damas_row}").number_format = "yyyy-mm-dd"
        ws.cell(row=excel_row, column=2, value=f"=DAMAS_Daily!B{damas_row}")

        for s_idx, scen in enumerate(SCENARIOS):
            k = scen["key"]
            cap_col = 3 + s_idx * 3
            act_col = 4 + s_idx * 3
            tot_col = 5 + s_idx * 3

            # Capacity revenue formula (always same per day per scenario):
            #  power × hours × max(damas_up, damas_down) × cap_mult × stack_cap × acceptance
            # We use the higher of the two clearings (operator picks more profitable direction).
            cap_formula = (
                f"=power_mw*hours_per_day*"
                f"MAX(damas_afrr_up,damas_afrr_down)*"
                f"capacity_mult_{k}*stack_capacity*acceptance_rate"
            )
            ws.cell(row=excel_row, column=cap_col, value=cap_formula).number_format = EUR_FMT

            # Activation revenue formula:
            #  captured_mwh = MIN(market_total × share × acceptance × stack, power × hours, capacity)
            #  rev = captured × avg_price × act_mult
            # Pick the more profitable direction (UP vs DOWN) using IF.
            up_rev = (
                f"MIN("
                f"DAMAS_Daily!C{damas_row}*market_share_{k}*acceptance_rate*stack_activation,"
                f"power_mw*hours_per_day,"
                f"capacity_mwh"
                f")*DAMAS_Daily!E{damas_row}*activation_mult_{k}"
            )
            down_rev = (
                f"MIN("
                f"DAMAS_Daily!D{damas_row}*market_share_{k}*acceptance_rate*stack_activation,"
                f"power_mw*hours_per_day,"
                f"capacity_mwh"
                f")*DAMAS_Daily!F{damas_row}*activation_mult_{k}"
            )
            act_formula = f"=MAX({up_rev},{down_rev})"
            ws.cell(row=excel_row, column=act_col, value=act_formula).number_format = EUR_FMT

            # Daily total = capacity + activation
            tot_formula = f"={get_column_letter(cap_col)}{excel_row}+{get_column_letter(act_col)}{excel_row}"
            ws.cell(row=excel_row, column=tot_col, value=tot_formula).number_format = EUR_FMT

            # Color-band by scenario
            scen_fill = PatternFill("solid", fgColor=scen["color"])
            for col in (cap_col, act_col, tot_col):
                ws.cell(row=excel_row, column=col).fill = scen_fill
                ws.cell(row=excel_row, column=col).border = styles["box"]
                ws.cell(row=excel_row, column=col).alignment = Alignment(horizontal="center")
            ws.cell(row=excel_row, column=tot_col).font = styles["bold"]

        for col in (1, 2):
            ws.cell(row=excel_row, column=col).border = styles["box"]

    # TOTAL row
    total_row = 4 + len(daily)
    ws.cell(row=total_row, column=1, value="TOTAL").font = styles["bold"]
    ws.cell(row=total_row, column=1).fill = styles["total"]
    for s_idx, scen in enumerate(SCENARIOS):
        for offset in (0, 1, 2):
            col = 3 + s_idx * 3 + offset
            col_letter = get_column_letter(col)
            ws.cell(row=total_row, column=col,
                    value=f"=SUM({col_letter}4:{col_letter}{total_row - 1})").number_format = EUR_FMT
            ws.cell(row=total_row, column=col).font = styles["bold"]
            ws.cell(row=total_row, column=col).fill = styles["total"]
            ws.cell(row=total_row, column=col).border = styles["box"]

    # Column widths
    widths = [12, 11] + [16, 16, 18] * 4
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C4"


def write_summary_sheet(ws, daily: pd.DataFrame, styles):
    """Sheet 4: monthly + annual totals via SUMIFS / SUM. All formulas."""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Summary — annual + monthly totals (formulas pull from Scenarios_Daily)"
    c.font = styles["bold_huge"]
    c.fill = styles["title"]
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 32

    # Annual totals — formulas pulling from Scenarios_Daily TOTAL row
    cur_row = 3
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
    ws.cell(row=cur_row, column=1, value="Annual totals (live formulas)").font = styles["bold_white"]
    ws.cell(row=cur_row, column=1).fill = styles["header"]
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center")
    cur_row += 1

    total_row_idx = 4 + len(daily)  # the TOTAL row in Scenarios_Daily
    headers = ["Scenario", "Capacity revenue (€)", "Activation revenue (€)", "Total annual (€)", "vs A %"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=cur_row, column=col_idx, value=h).font = styles["bold_white"]
        ws.cell(row=cur_row, column=col_idx).fill = styles["title"]
        ws.cell(row=cur_row, column=col_idx).alignment = Alignment(horizontal="center")
        ws.cell(row=cur_row, column=col_idx).border = styles["box"]
    cur_row += 1

    for s_idx, scen in enumerate(SCENARIOS):
        cap_col_letter = get_column_letter(3 + s_idx * 3)
        act_col_letter = get_column_letter(4 + s_idx * 3)
        tot_col_letter = get_column_letter(5 + s_idx * 3)
        ws.cell(row=cur_row, column=1, value=scen["label"]).font = styles["bold"]
        ws.cell(row=cur_row, column=2, value=f"=Scenarios_Daily!{cap_col_letter}{total_row_idx}").number_format = EUR_FMT
        ws.cell(row=cur_row, column=3, value=f"=Scenarios_Daily!{act_col_letter}{total_row_idx}").number_format = EUR_FMT
        ws.cell(row=cur_row, column=4, value=f"=Scenarios_Daily!{tot_col_letter}{total_row_idx}").number_format = EUR_FMT
        # vs A: divide own total by Scenario A total
        a_tot_letter = get_column_letter(5)  # column 5 is A's total
        ws.cell(row=cur_row, column=5, value=f"=Scenarios_Daily!{tot_col_letter}{total_row_idx}/Scenarios_Daily!{a_tot_letter}{total_row_idx}").number_format = "0.0%"
        scen_fill = PatternFill("solid", fgColor=scen["color"])
        for col in range(1, 6):
            ws.cell(row=cur_row, column=col).border = styles["box"]
            ws.cell(row=cur_row, column=col).fill = scen_fill
            ws.cell(row=cur_row, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=cur_row, column=4).font = styles["bold"]
        cur_row += 1

    # Monthly comparison — for each scenario, monthly sums via SUMIFS
    cur_row += 2
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
    ws.cell(row=cur_row, column=1, value="Monthly totals — Scenario A (live SUMIFS)").font = styles["bold_white"]
    ws.cell(row=cur_row, column=1).fill = styles["header"]
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center")
    cur_row += 1
    headers = ["Month", "A total (€)", "B total (€)", "C total (€)", "D total (€)", "Spread (€)"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=cur_row, column=col_idx, value=h).font = styles["bold_white"]
        ws.cell(row=cur_row, column=col_idx).fill = styles["title"]
        ws.cell(row=cur_row, column=col_idx).alignment = Alignment(horizontal="center")
        ws.cell(row=cur_row, column=col_idx).border = styles["box"]
    cur_row += 1

    months = sorted(set(pd.Timestamp(d).strftime("%Y-%m") for d in daily["date"]))
    daily_first_row = 4
    daily_last_row = 4 + len(daily) - 1

    for month in months:
        # Use SUMPRODUCT to filter by month string match
        ws.cell(row=cur_row, column=1, value=month).font = styles["bold"]
        for s_idx, scen in enumerate(SCENARIOS):
            tot_col_letter = get_column_letter(5 + s_idx * 3)
            # SUMIFS: sum totals where TEXT(date,"yyyy-mm") = month
            formula = (
                f"=SUMPRODUCT("
                f"(TEXT(Scenarios_Daily!A{daily_first_row}:A{daily_last_row},\"yyyy-mm\")=\"{month}\")*"
                f"Scenarios_Daily!{tot_col_letter}{daily_first_row}:{tot_col_letter}{daily_last_row}"
                f")"
            )
            ws.cell(row=cur_row, column=2 + s_idx, value=formula).number_format = EUR_FMT
        # Spread: max - min of A,B,C,D
        ws.cell(row=cur_row, column=6,
                value=f"=MAX(B{cur_row}:E{cur_row})-MIN(B{cur_row}:E{cur_row})").number_format = EUR_FMT
        for col in range(1, 7):
            ws.cell(row=cur_row, column=col).border = styles["box"]
            ws.cell(row=cur_row, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left")
        cur_row += 1

    # Notes
    cur_row += 2
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
    ws.cell(row=cur_row, column=1, value="How to use this workbook").font = styles["bold_white"]
    ws.cell(row=cur_row, column=1).fill = styles["header"]
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center")
    cur_row += 1
    notes = [
        "• Edit any YELLOW cell on the Inputs sheet — every formula recalculates instantly.",
        "• Power, capacity, acceptance, DAMAS clearing rates, stacking factors, and per-scenario multipliers are all editable.",
        "• DAMAS_Daily holds the static market data (366 days × 8 cols). Don't edit; replace by re-running scrape_damas_activations.py.",
        "• Scenarios_Daily computes daily revenue per scenario — every cell is a live formula.",
        "• Capacity revenue uses MAX(aFRR+, aFRR-) clearing — operator picks the more profitable direction.",
        "• Activation revenue uses MIN(market×share×acceptance×stack, power×hours, capacity) — battery throughput cap.",
        "• Numbers are SIMULATED, NOT bankable. DAMAS source = unverified_snapshot. Participant settlement export required for bankable mode.",
    ]
    for n in notes:
        ws.cell(row=cur_row, column=1, value=n)
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
        ws.cell(row=cur_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[cur_row].height = 28
        cur_row += 1

    widths = [42, 18, 18, 18, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ===================================================================
#  Main
# ===================================================================
def main() -> None:
    print("Loading DAMAS daily aggregates…")
    daily = load_damas_daily()
    print(f"  {len(daily)} days from {daily['date'].min()} → {daily['date'].max()}")
    print(f"  DAMAS canonical: aFRR+ €{DAMAS_AFRR_UP_DEFAULT:.2f}/MW/h, aFRR- €{DAMAS_AFRR_DOWN_DEFAULT:.2f}/MW/h")

    out_dir = ROOT / "reports"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"aFRR_scenarios_PICASSO_market_share_{daily['date'].min()}_to_{daily['date'].max()}_FORMULAS.xlsx"

    # Build empty workbook with 4 sheets
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        pd.DataFrame().to_excel(xl, sheet_name="Inputs", index=False)
        pd.DataFrame().to_excel(xl, sheet_name="DAMAS_Daily", index=False)
        pd.DataFrame().to_excel(xl, sheet_name="Scenarios_Daily", index=False)
        pd.DataFrame().to_excel(xl, sheet_name="Summary", index=False)

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    styles = _styles()

    # Clear pandas-written rows so we start clean
    for sn in ("Inputs", "DAMAS_Daily", "Scenarios_Daily", "Summary"):
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

    write_inputs_sheet(wb["Inputs"], styles)
    write_damas_daily_sheet(wb["DAMAS_Daily"], daily, styles)
    write_scenarios_daily_sheet(wb["Scenarios_Daily"], daily, styles)
    write_summary_sheet(wb["Summary"], daily, styles)

    wb.active = wb.sheetnames.index("Summary")
    wb.save(out_path)

    print(f"\nWrote: {out_path}")
    print("\nLayout (4 sheets, all formula-driven):")
    print("  1. Inputs            — yellow cells are editable (battery, DAMAS, scenario params)")
    print("  2. DAMAS_Daily       — static daily market data (366 rows)")
    print("  3. Scenarios_Daily   — 366 days × 4 scenarios × 3 cols, every cell is a formula")
    print("  4. Summary           — annual + monthly totals via formulas")


if __name__ == "__main__":
    main()
